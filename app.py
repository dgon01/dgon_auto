import streamlit as st
import os
import re
import math
from io import BytesIO
from datetime import datetime, date
import base64

# 현재 실행 디렉토리를 기준으로 경로 설정
APP_ROOT = os.path.dirname(os.path.abspath(__file__))

# =============================================================================
# 0. 페이지 설정 및 커스텀 디자인
# =============================================================================

st.set_page_config(
    layout="wide", 
    page_title="DG-Form | 전자설정 자동화시스템",
    page_icon=os.path.join(APP_ROOT, "my_icon.ico"),
    initial_sidebar_state="collapsed"
)

# 로고 이미지를 base64로 인코딩하는 함수
def get_base64_image(image_path):
    try:
        with open(image_path, "rb") as f:
            data = f.read()
        return base64.b64encode(data).decode()
    except:
        return None

# 로고 이미지 경로
LOGO_PATH = os.path.join(APP_ROOT, "my_icon.ico")
logo_base64 = get_base64_image(LOGO_PATH)

# 💡 등기온 공식 브랜드 컬러 및 스타일 적용
st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;600;700&display=swap');
    .stApp {{ font-family: 'Noto Sans KR', sans-serif !important; }}
    input, textarea, select, button {{ font-family: 'Noto Sans KR', sans-serif !important; }}
    
    .header-container {{
        background: white; border: 3px solid #00428B; padding: 20px 40px;
        border-radius: 15px; margin-bottom: 20px;
        box-shadow: 0 4px 15px rgba(0, 66, 139, 0.2);
        display: flex; align-items: center; justify-content: space-between;
    }}
    .logo-title-container {{ display: flex; align-items: center; gap: 20px; }}
    .header-logo {{ width: 120px; height: auto; }}
    .header-title {{ margin: 0; font-size: 2.5rem; font-weight: 700; }}
    .title-dg {{ color: #00428B; }}
    .title-form {{ color: #FDD000; }}
    .header-subtitle {{ color: #00428B; font-size: 1.2rem; font-weight: 500; margin: 0; }}
    
    /* 📱 반응형 CSS - 모바일 전체 대응 */
    @media (max-width: 768px) {{
        .header-container {{ 
            padding: 15px 20px !important; 
            flex-direction: column !important; 
            gap: 10px !important;
        }}
        .logo-title-container {{ flex-direction: column !important; text-align: center !important; }}
        .header-logo {{ width: 80px !important; }}
        .header-title {{ font-size: 1.8rem !important; }}
        .header-subtitle {{ font-size: 0.9rem !important; }}
        .header-right p {{ font-size: 0.85rem !important; }}
        h3 {{ font-size: 1.2rem !important; }}
        .total-amount {{ font-size: 1.5rem !important; }}
        
        /* 모든 columns를 세로 배치 */
        [data-testid="stHorizontalBlock"] {{
            flex-direction: column !important;
            gap: 10px !important;
        }}
        
        /* 각 컬럼 전체 너비 */
        [data-testid="stHorizontalBlock"] > div {{
            width: 100% !important;
            flex: none !important;
            min-width: 0 !important;
        }}
        
        /* 입력 필드 전체 너비 */
        .stNumberInput, .stTextInput, .stSelectbox, .stTextArea {{
            width: 100% !important;
        }}
        
        /* 버튼 전체 너비 */
        .stButton > button {{
            width: 100% !important;
            font-size: 0.85rem !important;
            padding: 8px 12px !important;
        }}
        
        /* 탭 버튼 축소 */
        .stTabs [data-baseweb="tab"] {{
            padding: 8px 12px !important;
            font-size: 0.8rem !important;
        }}
    }}
    
    /* 세로 모니터 대응 (height > width) */
    @media (max-aspect-ratio: 1/1) {{
        [data-testid="stHorizontalBlock"] {{
            flex-direction: column !important;
        }}
        [data-testid="stHorizontalBlock"] > div {{
            width: 100% !important;
            flex: none !important;
        }}
    }}
    
    .stTabs [data-baseweb="tab-list"] {{ gap: 10px; background-color: #ffffff; padding: 10px; border-radius: 12px; box-shadow: 0 2px 5px rgba(0,0,0,0.05); }}
    .stTabs [data-baseweb="tab"] {{ background-color: #f8f9fa; border-radius: 8px; padding: 10px 20px; font-weight: 600; color: #495057; border: 1px solid #dee2e6; }}
    .stTabs [aria-selected="true"] {{ background-color: #00428B; color: white; border-color: #00428B; }}

    .stTextInput > div > div > input, .stNumberInput > div > div > input, .stSelectbox > div > div > select {{
        border-radius: 6px; border: 1px solid #ced4da; padding: 8px 12px; font-size: 0.95rem;
    }}
    .stTextInput > div > div > input:focus {{ border-color: #00428B; box-shadow: 0 0 0 0.2rem rgba(0, 66, 139, 0.15); }}

    /* 헤더 버튼 스타일 - 한 줄 유지 */
    .header-btn-container {{
        display: flex;
        gap: 8px;
        flex-wrap: nowrap;
        white-space: nowrap;
    }}
    
    /* 헤더 영역 컬럼 한 줄 유지 */
    [data-testid="stHorizontalBlock"] {{
        flex-wrap: nowrap !important;
        align-items: center;
    }}
    
    /* 버튼 텍스트 줄바꿈 방지 */
    .stButton > button {{
        white-space: nowrap !important;
        min-width: fit-content !important;
    }}
    
    /* 3탭 커스텀 레이아웃 */
    .section-header {{ font-size: 1.1rem; font-weight: 700; margin-bottom: 15px; padding-bottom: 5px; border-bottom: 2px solid; }}
    .income-header {{ color: #28a745; border-color: #28a745; }}
    .tax-header {{ color: #fd7e14; border-color: #fd7e14; }}
    .total-header {{ color: #dc3545; border-color: #dc3545; }}
    .row-label {{ font-weight: 500; color: #495057; display: flex; align-items: center; height: 100%; font-size: 0.9rem; }}
    .total-box {{ background-color: #ff0033; color: white; padding: 20px; text-align: center; border-radius: 8px; margin: 15px 0; box-shadow: 0 4px 6px rgba(220, 53, 69, 0.3); }}
    .total-amount {{ font-size: 2rem; font-weight: 800; }}
    [data-testid="stContainer"] {{ background-color: white; padding: 20px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.05); border: 1px solid #e9ecef; }}
    
    /* 3탭 컬럼 높이 동일 - 강화 */
    [data-testid="stHorizontalBlock"]:has([data-testid="stColumn"]) {{
        align-items: stretch !important;
    }}
    [data-testid="stColumn"] > [data-testid="stVerticalBlockBorderWrapper"] {{
        height: 100% !important;
    }}
    [data-testid="stColumn"] > [data-testid="stVerticalBlockBorderWrapper"] > div {{
        height: 100% !important;
    }}
</style>
""", unsafe_allow_html=True)

# 헤더 섹션
if logo_base64:
    st.markdown(f"""
    <div class="header-container">
        <div class="logo-title-container">
            <img src="data:image/x-icon;base64,{logo_base64}" class="header-logo" alt="DG-ON Logo">
            <div>
                <h1 class="header-title"><span class="title-dg">DG</span><span class="title-form">-Form</span></h1>
                <p class="header-subtitle">전자설정 자동화시스템 | <span style="color: #FDD000;">등기온</span></p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
else:
    st.markdown("""
    <div class="header-container">
        <div>
            <h1 class="header-title">🏠 <span class="title-dg">DG</span><span class="title-form">-Form</span></h1>
            <p class="header-subtitle">전자설정 자동화시스템 | <span style="color: #FDD000;">등기온</span></p>
        </div>
    </div>
    """, unsafe_allow_html=True)

# =============================================================================
# 1. 라이브러리 및 환경 설정
# =============================================================================

# Excel (영수증)
try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    EXCEL_OK = True
except Exception:
    openpyxl = None
    MergedCell = None
    EXCEL_OK = False

# 계약서/자필서명정보 PDF (템플릿 위에 오버레이)
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from PyPDF2 import PdfReader, PdfWriter
    PDF_OK = True
except Exception:
    canvas = None
    A4 = None
    pdfmetrics = None
    TTFont = None
    PdfReader = None
    PdfWriter = None
    PDF_OK = False

# 비용내역 PDF (FPDF)
try:
    from fpdf import FPDF
    FPDF_OK = True
except Exception:
    FPDF = None
    FPDF_OK = False

# 등기부 PDF 파싱 (pdfplumber)
try:
    import pdfplumber
    PDFPLUMBER_OK = True
except Exception:
    pdfplumber = None
    PDFPLUMBER_OK = False

# 위택스 API 호출 (requests)
try:
    import requests
    REQUESTS_OK = True
except Exception:
    requests = None
    REQUESTS_OK = False

LIBS_OK = PDF_OK

# =============================================================================
# 2. 상수 및 데이터
# =============================================================================
TEMPLATE_FILENAMES = {
    "개인": "1.pdf",
    "3자담보": "2.pdf",
    "공동담보": "3.pdf",
    "자필": "자필서명정보 템플릿.pdf",
    "영수증": "receipt_template.xlsx",
    "영수증1": "receipt_template1.xlsx"
}

CREDITORS = {
    "(주)티플레인대부 대표이사 윤웅원": {"addr": "서울특별시 마포구 삼개로16, 2신관1층103호(도화동,근신빌딩)", "corp_num": "110111-7350161", "fee": {"제증명": 50000, "교통비": 100000, "원인증서": 50000}},
    "(주)유노스프레스티지대부 사내이사 한은수": {"addr": "서울특별시 강남구 압구정로28길24, 5층 501호(신사동,디앤씨빌딩)", "corp_num": "110111-4138560", "fee": {"제증명": 20000, "교통비": 0, "원인증서": 0, "확인서면": 0, "선순위 말소": 0}},
    "(주)파트너스대부 사내이사 허성": {"addr": "부산광역시 부산진구 서면문화로 43, 2층(부전동)", "corp_num": "180111-1452175", "fee": {"제증명": 50000, "교통비": 100000, "원인증서": 50000}},
    "(주)드림앤캐쉬대부 대표이사 김재섭": {"addr": "서울특별시 강남구 압구정로28길24, 6층 601호(신사동,디앤씨빌딩)", "corp_num": "110111-4176552", "fee": {"제증명": 20000, "교통비": 0, "원인증서": 0}},
    "(주)마젤란트러스트대부 대표이사 김병수": {"addr": "서울특별시 서초구 강남대로34길 7, 7층(양재동,이안빌딩)", "corp_num": "110111-6649979", "fee": {"제증명": 50000, "교통비": 100000, "원인증서": 50000}},
    "(주)하이클래스대부 사내이사 성윤호": {"addr": "서울특별시 강남구 도곡로 188, 3층 4호(도곡동,도곡스퀘어)", "corp_num": "110111-0933512", "fee": {"제증명": 50000, "교통비": 100000, "원인증서": 50000}},
    "㈜엘하비스트대부 대표이사 김상수": {"addr": "서울특별시 중구 무교로 15, 16층(무교동,남강건설회관빌딩)", "corp_num": "110111-3648627", "fee": {"제증명": 20000, "교통비": 0, "원인증서": 0}}
}

# =============================================================================
# 등기의무자(근저당권자) 목록 - 말소용
# =============================================================================
# name: 법인명, addr: 주소, corp_num: 법인등록번호, rep: 대표자직위+성명, branch: 기본 취급지점
OBLIGORS = {
    # === 직접입력 ===
    "직접입력": {
        "addr": "",
        "corp_num": "",
        "rep": "",
        "branch": ""
    },
    # === 1금융권 ===
    "주식회사 하나은행": {
        "addr": "서울특별시 중구 을지로 35(을지로1가)",
        "corp_num": "110111-0672538",
        "rep": "지배인 강석윤",
        "branch": "업무혁신부"
    },
    "주식회사 국민은행": {
        "addr": "서울특별시 영등포구 국제금융로8길 26(여의도동)",
        "corp_num": "110111-2365321",
        "rep": "지배인 김동선",
        "branch": "대출실행센터"
    },
    "주식회사 우리은행": {
        "addr": "서울특별시 중구 소공로51(회현동1가)",
        "corp_num": "110111-0023393",
        "rep": "지배인 김미정",
        "branch": "본점"
    },
    "주식회사 신한은행": {
        "addr": "서울특별시 중구 세종대로9길 20(태평로2가)",
        "corp_num": "110111-0012809",
        "rep": "지배인 홍길동",
        "branch": ""
    },
    "농협은행주식회사": {
        "addr": "서울특별시 중구 통일로 120(충정로1가)",
        "corp_num": "110111-4809385",
        "rep": "지배인 전경호",
        "branch": "가양중앙지점"
    },
    "북파주농업협동조합": {
        "addr": "경기도 파주시 문산읍 문향로 75",
        "corp_num": "115636-0000297",
        "rep": "조합장 이갑영",
        "branch": "적성지점"
    },
    "삼성화재해상보험주식회사": {
        "addr": "서울특별시 서초구 서초대로74길 14(서초동)",
        "corp_num": "110111-0005078",
        "rep": "대표이사 이문화",
        "branch": ""
    },
    # === 대부업체 ===
    "(주)티플레인대부": {
        "addr": "서울특별시 마포구 삼개로16, 2신관1층103호(도화동,근신빌딩)",
        "corp_num": "110111-7350161",
        "rep": "대표이사 윤웅원",
        "branch": ""
    },
    "(주)유노스프레스티지대부": {
        "addr": "서울특별시 강남구 압구정로28길24, 5층 501호(신사동,디앤씨빌딩)",
        "corp_num": "110111-4138560",
        "rep": "사내이사 한은수",
        "branch": ""
    },
    "(주)파트너스대부": {
        "addr": "부산광역시 부산진구 서면문화로 43, 2층(부전동)",
        "corp_num": "180111-1452175",
        "rep": "사내이사 허성",
        "branch": ""
    },
    "(주)드림앤캐쉬대부": {
        "addr": "서울특별시 강남구 압구정로28길24, 6층 601호(신사동,디앤씨빌딩)",
        "corp_num": "110111-4176552",
        "rep": "대표이사 김재섭",
        "branch": ""
    },
    "(주)마젤란트러스트대부": {
        "addr": "서울특별시 서초구 강남대로34길 7, 7층(양재동,이안빌딩)",
        "corp_num": "110111-6649979",
        "rep": "대표이사 김병수",
        "branch": ""
    },
    "(주)하이클래스대부": {
        "addr": "서울특별시 강남구 도곡로 188, 3층 4호(도곡동,도곡스퀘어)",
        "corp_num": "110111-0933512",
        "rep": "사내이사 성윤호",
        "branch": ""
    },
    "(주)엘하비스트대부": {
        "addr": "서울특별시 중구 무교로 15, 16층(무교동,남강건설회관빌딩)",
        "corp_num": "110111-3648627",
        "rep": "대표이사 김상수",
        "branch": ""
    }
}

# =============================================================================
# 1금융권 좌표 설정 (부동산표시 영역)
# =============================================================================
# 좌표: x=시작X, y_top=PDF상단기준Y, box_height=박스높이
BANK_COORDS = {
    "하나은행": {
        "설정계약서": {"x": 42, "y_top": 114, "box_height": 140},
        "위임장": {"x": 89, "y_top": 116, "box_height": 289}
    },
    "신한은행": {
        "설정계약서": {"x": 49, "y_top": 273, "box_height": 177},
        "위임장": {"x": 100, "y_top": 127, "box_height": 251}
    },
    "우리은행": {
        "설정계약서": {"x": 82, "y_top": 84, "box_height": 252},
        "위임장": {"x": 85, "y_top": 118, "box_height": 228}
    }
}

def resource_path(relative_path):
    return os.path.join(APP_ROOT, relative_path)

FONT_PATH = resource_path("Malgun.ttf") 

# 템플릿 파일 상태 확인
if 'template_status' not in st.session_state:
    st.session_state['template_status'] = {}
    missing_files = []
    for key, filename in TEMPLATE_FILENAMES.items():
        path = resource_path(filename)
        st.session_state['template_status'][key] = path if os.path.exists(path) else None
        if not st.session_state['template_status'][key]:
            missing_files.append(filename)
    st.session_state['missing_templates'] = missing_files

# =============================================================================
# 3. 유틸리티 및 계산 로직
# =============================================================================
def format_date_korean(date_obj):
    if isinstance(date_obj, date): return f"{date_obj.year}년 {date_obj.month:02d}월 {date_obj.day:02d}일"
    return str(date_obj)

def format_number_with_comma(num_str):
    if num_str is None: return ""
    if isinstance(num_str, (int, float)): return "{:,}".format(int(num_str))
    numbers = re.sub(r'[^\d]', '', str(num_str))
    if not numbers: return ""
    try: return "{:,}".format(int(numbers))
    except ValueError: return num_str

def remove_commas(v):
    if v is None: return ""
    if isinstance(v, (int, float)): return str(int(v))
    return v.replace(',', '') if isinstance(v, str) else str(v)

def floor_10(v): return math.floor(v / 10) * 10

def lookup_base_fee(amount):
    """법무사 보수표 기준 기본료 계산 (2024.9.12 시행)"""
    if amount <= 50_000_000:
        return 210_000
    elif amount <= 100_000_000:
        return 210_000 + int((amount - 50_000_000) * 10 / 10000)
    elif amount <= 300_000_000:
        return 260_000 + int((amount - 100_000_000) * 9 / 10000)
    elif amount <= 500_000_000:
        return 440_000 + int((amount - 300_000_000) * 8 / 10000)
    elif amount <= 1_000_000_000:
        return 600_000 + int((amount - 500_000_000) * 7 / 10000)
    elif amount <= 2_000_000_000:
        return 950_000 + int((amount - 1_000_000_000) * 5 / 10000)
    elif amount <= 20_000_000_000:
        return 1_450_000 + int((amount - 2_000_000_000) * 4 / 10000)
    else:
        return 8_650_000 + int((amount - 20_000_000_000) * 1 / 10000)

def get_rate():
    try:
        import requests
        url = "https://lawss.co.kr/lawpro/homepage/siga/auto_siga_kjaa.php"
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=3)
        response.encoding = 'EUC-KR'
        match = re.search(r"오늘 채권할인율\s*=\s*([\d\.]+) %", response.text)
        if match: return math.ceil(float(match.group(1)) * 10) / 10 / 100
    except: pass
    return 0.0913459

def number_to_korean(num_str):
    if not num_str: return ""
    try: num = int(re.sub(r'[^\d]', '', num_str))
    except: return ""
    units = ['', '만', '억', '조']; digits = ['', '일', '이', '삼', '사', '오', '육', '칠', '팔', '구']
    if num == 0: return "영원정"
    result = []; unit_idx = 0
    while num > 0:
        part = num % 10000
        if part > 0:
            part_str = ""
            if part >= 1000: part_str += digits[part // 1000] + "천"; part %= 1000
            if part >= 100: part_str += digits[part // 100] + "백"; part %= 100
            if part >= 10: part_str += digits[part // 10] + "십"; part %= 10
            if part > 0: part_str += digits[part]
            result.append(part_str + units[unit_idx])
        num //= 10000; unit_idx += 1
    return ''.join(reversed(result)) + "원정"

def convert_multiple_amounts_to_korean(amount_str):
    if not amount_str: return ""
    if '/' in amount_str:
        return ', '.join([number_to_korean(p.strip()) for p in amount_str.split('/') if number_to_korean(p.strip())])
    return number_to_korean(amount_str)

def extract_address_from_estate(estate_text):
    if not estate_text: return ""
    lines = [line.strip() for line in estate_text.strip().split('\n')]
    for line in lines:
        if "1동의 건물의 표시" in line or "건물의 표시" in line: continue
        if any(region in line for region in ['특별시', '광역시', '시 ', '군 ', '구 ']):
            if '대 ' not in line and '도로명주소' not in line and '[' not in line:
                return line.strip()
    return ""

# PDF 관련 클래스 및 함수 생략 (위와 동일)
if FPDF_OK:
    class PDFConverter(FPDF):
        def __init__(self, show_fee=True):
            super().__init__()
            self.show_fee = show_fee
            self.line_height = 6.5
            self.col_width1 = 150; self.col_width2 = 30
            if FONT_PATH and os.path.exists(FONT_PATH):
                try:
                    self.add_font('Malgun', '', FONT_PATH, uni=True)
                    self.add_font('Malgun', 'B', FONT_PATH, uni=True)
                    self.set_font('Malgun', '', 11)
                except: self.set_font('Arial', '', 11)
            else: self.set_font('Arial', '', 11)
        
        def draw_labelframe_box(self, title, content_func):
            self.set_font(self.font_family, 'B', 11)
            start_y = self.get_y(); start_x = self.l_margin
            box_width = self.w - self.l_margin * 2
            self.set_y(start_y + self.line_height)
            content_func()
            content_end_y = self.get_y()
            box_height = (content_end_y - start_y) + self.line_height + 4
            self.set_draw_color(211, 211, 211)
            self.rect(start_x, start_y + self.font_size / 2, box_width, box_height)
            title_width = self.get_string_width(title)
            self.set_fill_color(255, 255, 255)
            self.rect(start_x + 9, start_y, title_width + 4, self.font_size, 'F')
            self.set_xy(start_x + 11, start_y)
            self.cell(0, self.font_size, title)
            self.set_y(content_end_y + 4)
            
        def output_pdf(self, data):
            self.add_page(); self.set_font(self.font_family, 'B', 20)
            self.cell(0, 12, "근저당권설정 비용내역", ln=True, align="C"); self.ln(2)
            self.set_font(self.font_family, '', 9)
            self.cell(0, 5, f"작성일: {data['date_input']}", ln=True, align="R"); self.ln(2)
            self.set_font(self.font_family, '', 10)
            client = data['client']
            self.cell(95, self.line_height, f"채권최고액: {client['채권최고액']} 원")
            self.cell(0, self.line_height, f"|  필지수: {client['필지수']}", ln=True)
            if client.get('금융사'): self.cell(0, self.line_height, f"금  융  사: {client['금융사']}", ln=1)
            if client.get('채무자'): self.cell(0, self.line_height, f"채  무  자: {client['채무자']}", ln=1)
            if client.get('물건지'): self.multi_cell(0, self.line_height, f"물  건  지: {client['물건지']}")
            self.ln(3)
            if self.show_fee:
                def fee_content():
                    self.set_font(self.font_family, '', 10); items = data['fee_items']
                    subtotal = items.get('기본료', 0) + items.get('추가보수', 0) + items.get('기타보수', 0)
                    self.set_x(self.l_margin + 5); self.cell(self.col_width1, self.line_height, "보수액 소계"); self.cell(self.col_width2, self.line_height, f"{subtotal:,} 원", ln=1, align="R")
                    self.set_x(self.l_margin + 5); self.cell(self.col_width1, self.line_height, "할인금액"); self.cell(self.col_width2, self.line_height, f"{items.get('할인금액', 0):,} 원", ln=1, align="R")
                    self.ln(1); self.line(self.get_x() + 5, self.get_y(), self.w - self.r_margin - 5, self.get_y()); self.ln(1)
                    self.set_font(self.font_family, 'B', 10); self.set_x(self.l_margin + 5); self.cell(self.col_width1, self.line_height, "보수 소계"); self.cell(self.col_width2, self.line_height, f"{data['fee_totals']['보수총액']:,} 원", ln=1, align="R")
                self.draw_labelframe_box("1. 보수액", fee_content); self.ln(5)
            def costs_content():
                self.set_font(self.font_family, '', 10); items = data['cost_items']
                for name, val in items.items():
                    if val != 0: self.set_x(self.l_margin + 5); self.cell(self.col_width1, self.line_height, name); self.cell(self.col_width2, self.line_height, f"{int(val):,} 원", ln=1, align="R")
                self.ln(1); self.line(self.get_x() + 5, self.get_y(), self.w - self.r_margin - 5, self.get_y()); self.ln(1)
                self.set_font(self.font_family, 'B', 10); self.set_x(self.l_margin + 5); self.cell(self.col_width1, self.line_height, "공과금소계"); self.cell(self.col_width2, self.line_height, f"{data['cost_totals']['공과금 총액']:,} 원", ln=1, align="R")
            self.draw_labelframe_box(data['cost_section_title'], costs_content); self.ln(5)
            self.ln(3)  # 등기비용 합계를 더 아래로 이동
            self.set_font(self.font_family, 'B', 12); self.cell(self.col_width1 - 10, 10, "등기비용 합계"); self.cell(self.col_width2 + 10, 10, f"{data['grand_total']:,} 원", ln=True, align="R"); self.ln(5)
            def notes_content():
                self.set_font(self.font_family, '', 10); self.set_x(self.l_margin + 5); self.cell(0, self.line_height, "• 원활한 확인을 위해 입금자는 소유자명(또는 채무자명)으로 기재해 주세요.", ln=1)
                self.set_x(self.l_margin + 5); self.cell(0, self.line_height, "• 입금 완료 후, 메시지를 남겨주시면 더욱 빠르게 처리됩니다.", ln=1)
                self.set_x(self.l_margin + 5); self.cell(0, self.line_height, "• 업무는 입금이 확인된 후에 진행됩니다.", ln=1)
            self.draw_labelframe_box("안내사항", notes_content); self.ln(5)
            def bank_content():
                self.set_font(self.font_family, '', 10)
                self.set_x(self.l_margin + 5)
                self.cell(0, self.line_height, "• 신한은행 100-035-852291", ln=1)
                self.set_x(self.l_margin + 5)
                self.cell(0, self.line_height, "• 예금주 : 법무법인 시화", ln=1)
            self.draw_labelframe_box("입금 계좌 정보", bank_content)
            pdf_buffer = BytesIO(); pdf_bytes = self.output(dest='S')
            if isinstance(pdf_bytes, str): pdf_buffer.write(pdf_bytes.encode('latin-1'))
            else: pdf_buffer.write(pdf_bytes)
            pdf_buffer.seek(0); return pdf_buffer
else: PDFConverter = None

def draw_fit_text(canvas_obj, text, x, y, max_width, font_name, font_size, min_font_size=5):
    """텍스트를 max_width에 맞춰 폰트 크기 자동 조절 (1줄로 맞춤)"""
    if not text:
        return
    
    current_font_size = font_size
    
    # 폰트 크기를 줄여가며 한 줄에 맞추기
    while current_font_size >= min_font_size:
        canvas_obj.setFont(font_name, current_font_size)
        
        # 한 줄에 들어가는지 확인
        if canvas_obj.stringWidth(text, font_name, current_font_size) <= max_width:
            canvas_obj.drawString(x, y, text)
            canvas_obj.setFont(font_name, font_size)  # 원래 폰트 크기로 복원
            return
        
        # 안 맞으면 폰트 크기 줄이기
        current_font_size -= 0.5
    
    # 최소 폰트로도 안 맞으면 잘라서 표시
    canvas_obj.setFont(font_name, min_font_size)
    truncated = text
    while canvas_obj.stringWidth(truncated + "...", font_name, min_font_size) > max_width and len(truncated) > 10:
        truncated = truncated[:-1]
    if len(truncated) < len(text):
        truncated += "..."
    canvas_obj.drawString(x, y, truncated)
    canvas_obj.setFont(font_name, font_size)  # 원래 폰트 크기로 복원

def create_overlay_pdf(data, font_path):
    packet = BytesIO(); c = canvas.Canvas(packet, pagesize=A4); width, height = A4
    try: pdfmetrics.registerFont(TTFont('Korean', font_path)); font_name = 'Korean'
    except: font_name = 'Helvetica'
    font_size = 11; c.setFont(font_name, font_size); c.setFillColorRGB(0, 0, 0)
    MAX_TEXT_WIDTH = 380
    if data.get("date"): c.drawString(480, height - 85, data["date"])
    if data.get("creditor_name"): c.drawString(157, height - 134, data["creditor_name"])
    if data.get("creditor_addr"): draw_fit_text(c, data["creditor_addr"], 157, height - 150, MAX_TEXT_WIDTH, font_name, font_size)
    if data.get("debtor_name"): c.drawString(157, height - 172, data["debtor_name"])
    if data.get("debtor_addr"): draw_fit_text(c, data["debtor_addr"], 157, height - 190, MAX_TEXT_WIDTH, font_name, font_size)
    if data.get("owner_name"): c.drawString(157, height - 212, data["owner_name"])
    if data.get("owner_addr"): draw_fit_text(c, data["owner_addr"], 157, height - 230, MAX_TEXT_WIDTH, font_name, font_size)
    if data.get("guarantee_type"): c.drawString(65, height - 343, data["guarantee_type"])
    if data.get("claim_amount"): c.drawString(150, height - 535, data["claim_amount"])
    c.showPage(); c.setFont(font_name, font_size)
    if data.get("date"): c.drawString(180, height - 270, data["date"])
    contract_type = data.get("contract_type", "3자담보")
    if contract_type == "개인": 
        if data.get("debtor_name"): c.drawString(450, height - 270, data["debtor_name"])
    elif contract_type == "3자담보": 
        if data.get("owner_name"): c.drawString(490, height - 270, data["owner_name"])
    elif contract_type == "공동담보": 
        if data.get("debtor_name"): c.drawString(450, height - 270, data["debtor_name"])
        if data.get("owner_name"): c.drawString(490, height - 270, data["owner_name"])
    c.showPage(); c.setFont(font_name, font_size)
    base_x = 35; base_y = height - 80; gap = 16
    for i, line in enumerate(data.get("estate_list", [])):
        if line.strip(): c.drawString(base_x, base_y - (i * gap), line)
    c.showPage(); c.save(); packet.seek(0)
    return packet

def make_pdf(template_path, data):
    overlay_packet = create_overlay_pdf(data, FONT_PATH)
    overlay_pdf = PdfReader(overlay_packet); template_pdf = PdfReader(template_path); writer = PdfWriter()
    output_buffer = BytesIO() 
    for page_num in range(min(len(template_pdf.pages), len(overlay_pdf.pages))):
        template_page = template_pdf.pages[page_num]; overlay_page = overlay_pdf.pages[page_num]
        template_page.merge_page(overlay_page); writer.add_page(template_page)
    writer.write(output_buffer)
    output_buffer.seek(0)
    return output_buffer

def make_signature_pdf(template_path, data):
    packet = BytesIO(); c = canvas.Canvas(packet, pagesize=A4); width, height = A4
    try: pdfmetrics.registerFont(TTFont('Korean', FONT_PATH)); font_name = 'Korean'
    except: font_name = 'Helvetica'
    c.setFont(font_name, 10); estate_x = 150; estate_y = height - 170; line_h = 14
    if data.get("estate_text"):
        for i, line in enumerate(str(data["estate_text"]).split("\n")[:17]):
            c.drawString(estate_x, estate_y - (i * line_h), line)
    # 등기원인 - 박스 좌표 (139.5, 557.5, 340.5, 581.0)
    if data.get("purpose"):
        c.setFont(font_name, 11)
        c.drawString(145, 569, str(data["purpose"]))
        c.setFont(font_name, 10)
    if data.get("debtor_name"): c.drawString(250, 322, str(data["debtor_name"]))
    if data.get("debtor_rrn"): c.drawString(250, 298, str(data["debtor_rrn"]))
    if data.get("owner_name"): c.drawString(400, 322, str(data["owner_name"]))
    if data.get("owner_rrn"): c.drawString(400, 298, str(data["owner_rrn"]))
    if data.get("date"):
        c.setFont(font_name, 11); text = str(data["date"]); tw = c.stringWidth(text, font_name, 11)
        c.drawString((width - tw) / 2, 150, text)
    c.showPage(); c.save(); packet.seek(0)
    overlay_pdf = PdfReader(packet); template_pdf = PdfReader(template_path); writer = PdfWriter()
    output_buffer = BytesIO()
    template_page = template_pdf.pages[0]; overlay_page = overlay_pdf.pages[0]
    template_page.merge_page(overlay_page); writer.add_page(template_page)
    writer.write(output_buffer); output_buffer.seek(0)
    return output_buffer

# =============================================================================
# 1금융권 PDF 생성 함수들
# =============================================================================

def make_bank_estate_pdf(bank_name, doc_type, estate_text):
    """1금융권 설정계약서/위임장 PDF 생성 (부동산표시만, 빈 PDF, 자동 크기 조절)"""
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=A4)
    width, height = A4
    
    try:
        pdfmetrics.registerFont(TTFont('Korean', FONT_PATH))
        font_name = 'Korean'
    except:
        font_name = 'Helvetica'
    
    coords = BANK_COORDS.get(bank_name, {}).get(doc_type, {})
    x = coords.get("x", 50)
    y_top = coords.get("y_top", 100)
    box_height = coords.get("box_height", 200)
    
    # y 좌표 변환 (PDF 상단 기준 → reportlab 좌표)
    y_start = height - y_top
    
    lines = [line for line in (estate_text.split("\n") if estate_text else []) if line.strip()]
    num_lines = len(lines)
    
    if num_lines == 0:
        c.showPage()
        c.save()
        packet.seek(0)
        return packet
    
    # 폰트 크기와 줄 간격 자동 계산 (박스 높이 내에 맞추기)
    # 기본: 폰트 9pt, 줄간격 12pt
    # 텍스트가 많으면 폰트와 줄간격을 줄임
    max_font = 9
    min_font = 6
    
    # 필요한 총 높이 계산 (줄 수 * 줄간격)
    # 줄간격은 폰트 크기의 약 1.3배
    for font_size in [9, 8.5, 8, 7.5, 7, 6.5, 6]:
        line_h = font_size * 1.35
        total_height = num_lines * line_h
        if total_height <= box_height:
            break
    else:
        # 최소 폰트로도 안 맞으면 줄 간격 더 줄임
        font_size = min_font
        line_h = box_height / num_lines if num_lines > 0 else 10
    
    c.setFont(font_name, font_size)
    
    for i, line in enumerate(lines):
        c.drawString(x, y_start - (i * line_h), line)
    
    c.showPage()
    c.save()
    packet.seek(0)
    return packet

def make_bank_signature_pdf(template_path, data):
    """1금융권 자필서명정보 PDF 생성"""
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=A4)
    width, height = A4
    
    try:
        pdfmetrics.registerFont(TTFont('Korean', FONT_PATH))
        font_name = 'Korean'
    except:
        font_name = 'Helvetica'
    
    # 부동산표시 (기존 템플릿 좌표 사용)
    c.setFont(font_name, 10)
    estate_x = 150
    estate_y = height - 170
    line_h = 14
    
    estate_text = data.get("estate_text", "")
    if estate_text:
        for i, line in enumerate(str(estate_text).split("\n")[:17]):
            if line.strip():
                c.drawString(estate_x, estate_y - (i * line_h), line)
    
    # 등기원인 - 박스 좌표 (139.5, 557.5, 340.5, 581.0)
    if data.get("purpose"):
        c.setFont(font_name, 11)
        c.drawString(145, 569, str(data["purpose"]))
    
    # 채무자 정보
    c.setFont(font_name, 10)
    if data.get("debtor_name"):
        c.drawString(250, 322, str(data["debtor_name"]))
    if data.get("debtor_rrn"):
        c.drawString(250, 298, str(data["debtor_rrn"]))
    
    # 소유자 정보
    if data.get("owner_name"):
        c.drawString(400, 322, str(data["owner_name"]))
    if data.get("owner_rrn"):
        c.drawString(400, 298, str(data["owner_rrn"]))
    
    # 날짜 (중앙)
    if data.get("date"):
        c.setFont(font_name, 11)
        text = str(data["date"])
        tw = c.stringWidth(text, font_name, 11)
        c.drawString((width - tw) / 2, 150, text)
    
    c.showPage()
    c.save()
    packet.seek(0)
    
    # 템플릿과 병합
    overlay_pdf = PdfReader(packet)
    template_pdf = PdfReader(template_path)
    writer = PdfWriter()
    output_buffer = BytesIO()
    
    template_page = template_pdf.pages[0]
    overlay_page = overlay_pdf.pages[0]
    template_page.merge_page(overlay_page)
    writer.add_page(template_page)
    
    writer.write(output_buffer)
    output_buffer.seek(0)
    return output_buffer

# =============================================================================
# 말소 문서 PDF 생성 함수들
# =============================================================================

def make_malso_signature_pdf(template_path, data):
    """말소용 자필서명정보 PDF 생성 (탭2와 유사)"""
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=A4)
    width, height = A4
    
    try:
        pdfmetrics.registerFont(TTFont('Korean', FONT_PATH))
        font_name = 'Korean'
    except:
        font_name = 'Helvetica'
    
    c.setFont(font_name, 10)
    
    # 부동산 표시
    estate_x = 150
    estate_y = height - 170
    line_h = 14
    estate_list = data.get('estate_list', [])
    for i, line in enumerate(estate_list[:17]):
        if line.strip():
            c.drawString(estate_x, estate_y - (i * line_h), line)
    
    # 등기원인 - 박스 좌표 (139.5, 557.5, 340.5, 581.0)
    if data.get("purpose"):
        c.setFont(font_name, 11)
        c.drawString(145, 569, str(data["purpose"]))
        c.setFont(font_name, 10)
    
    # 권리자 정보 (최대 2명)
    holders = data.get('holders', [])
    if len(holders) >= 1:
        c.drawString(250, 322, str(holders[0].get('name', '')))
        c.drawString(250, 298, str(holders[0].get('rrn', '')))
    if len(holders) >= 2:
        c.drawString(400, 322, str(holders[1].get('name', '')))
        c.drawString(400, 298, str(holders[1].get('rrn', '')))
    
    # 날짜 (중앙)
    if data.get("date"):
        c.setFont(font_name, 11)
        text = str(data["date"])
        tw = c.stringWidth(text, font_name, 11)
        c.drawString((width - tw) / 2, 150, text)
    
    c.showPage()
    c.save()
    packet.seek(0)
    
    # 템플릿과 병합
    overlay_pdf = PdfReader(packet)
    template_pdf = PdfReader(template_path)
    writer = PdfWriter()
    output_buffer = BytesIO()
    
    template_page = template_pdf.pages[0]
    overlay_page = overlay_pdf.pages[0]
    template_page.merge_page(overlay_page)
    writer.add_page(template_page)
    
    writer.write(output_buffer)
    output_buffer.seek(0)
    return output_buffer

def make_malso_power_pdf(template_path, data):
    """위임장 PDF 생성 (템플릿 오버레이)"""
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=A4)
    width, height = A4
    
    try:
        pdfmetrics.registerFont(TTFont('Korean', FONT_PATH))
        font_name = 'Korean'
    except:
        font_name = 'Helvetica'
    
    # 위임장 좌표 (분석 결과 기반)
    # 부동산 표시: (102.9, 93.1) ~ (529.7, 344.6), RL Y: 497.4 ~ 748.9
    # 등기원인/목적: X=172.0, RL Y: 450~491
    # 말소할 사항: (172.0, 397.1) ~ (530.3, 455.4), RL Y: 386.6 ~ 444.9
    # 의무자/권리자: (65.1, 590.3) ~ (415.4, 778.9), RL Y: 63.1 ~ 251.7
    
    c.setFont(font_name, 9)
    
    # 부동산 표시 (Box 1)
    estate_x = 105
    estate_y = 745
    line_h = 12
    estate_lines = data.get('estate_text', '').split('\n')
    for i, line in enumerate(estate_lines[:20]):
        if line.strip():
            c.drawString(estate_x, estate_y - (i * line_h), line)
    
    # 등기원인과 그 년월일 (Box 2) - 상하 중앙정렬, RL Y: 470.6 ~ 491.1, 중앙 480.85
    cause_date = data.get('date', '')
    c.drawString(175, 478, f"{cause_date} 해지")
    
    # 등기목적 (Box 3) - 상하 중앙정렬, RL Y: 450.6 ~ 471.1, 중앙 460.85
    malso_type = data.get('malso_type', '근저당권')
    c.drawString(175, 458, f"{malso_type}말소")
    
    # 말소할 사항 (Box 4) - 상하 중앙정렬, RL Y: 386.6 ~ 444.9, 중앙 415.75
    cancel_text = data.get('cancel_text', '')
    c.setFont(font_name, 8)
    # 긴 텍스트 줄바꿈 처리 (폭 넓힘)
    if len(cancel_text) > 75:
        c.drawString(175, 422, cancel_text[:75])
        c.drawString(175, 410, cancel_text[75:])
    else:
        c.drawString(175, 416, cancel_text)
    
    c.setFont(font_name, 9)
    
    # 의무자 (소유자) - 왼쪽 상단 (Box 5 영역), RL Y: 63.1 ~ 251.7
    holder1_name = data.get('holder1_name', '')
    holder1_addr = data.get('holder1_addr', '')
    holder2_name = data.get('holder2_name', '')
    holder2_addr = data.get('holder2_addr', '')
    
    # 등기의무자 라벨
    c.setFont(font_name, 8)
    c.drawString(70, 248, "등기의무자")
    c.setFont(font_name, 9)
    
    # 의무자 내용
    c.drawString(70, 232, holder1_name)
    # 주소 (여러 줄 처리)
    addr_lines = holder1_addr.split('\n') if holder1_addr else []
    for i, line in enumerate(addr_lines[:2]):
        c.drawString(70, 217 - (i * 12), line)
    
    if holder2_name:
        c.drawString(70, 180, holder2_name)
        addr2_lines = holder2_addr.split('\n') if holder2_addr else []
        for i, line in enumerate(addr2_lines[:2]):
            c.drawString(70, 165 - (i * 12), line)
    
    # 권리자 (채권자) - 하단
    obligor_name = data.get('obligor_name', '')
    obligor_id = data.get('obligor_id', '')
    obligor_addr = data.get('obligor_addr', '')
    obligor_rep = data.get('obligor_rep', '')
    obligor_branch = data.get('obligor_branch', '')
    
    # 등기권리자 라벨
    c.setFont(font_name, 8)
    c.drawString(70, 118, "등기권리자")
    c.setFont(font_name, 9)
    
    # 위임장: 법인등록번호 없이 출력
    c.drawString(70, 102, obligor_name)
    c.drawString(70, 87, obligor_addr)
    
    # 취급지점 + 대표자 형식
    if obligor_branch and obligor_rep:
        rep_text = f"(취급지점:{obligor_branch}) {obligor_rep}"
    elif obligor_rep:
        rep_text = obligor_rep
    else:
        rep_text = ""
    
    if rep_text:
        c.drawString(70, 72, rep_text)
    
    c.showPage()
    c.save()
    packet.seek(0)
    
    # 템플릿과 병합 (1페이지만)
    overlay_pdf = PdfReader(packet)
    template_pdf = PdfReader(template_path)
    writer = PdfWriter()
    output_buffer = BytesIO()
    
    template_page = template_pdf.pages[0]
    overlay_page = overlay_pdf.pages[0]
    template_page.merge_page(overlay_page)
    writer.add_page(template_page)
    
    writer.write(output_buffer)
    output_buffer.seek(0)
    return output_buffer

def make_malso_termination_pdf(data):
    """해지증서 PDF 생성 (백지에서 생성)"""
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=A4)
    width, height = A4
    
    try:
        pdfmetrics.registerFont(TTFont('Korean', FONT_PATH))
        font_name = 'Korean'
    except:
        font_name = 'Helvetica'
    
    # 페이지 설정: 좌측 X=50, 우측 X=545 (여백 줄임)
    left_x = 50
    right_x = 545
    center_x = (left_x + right_x) / 2
    content_width = right_x - left_x
    
    # 제목: 해 지 증 서 (중앙, 상단)
    c.setFont(font_name, 18)
    title = "해 지 증 서"
    title_width = c.stringWidth(title, font_name, 18)
    c.drawString(center_x - title_width/2, 750, title)
    
    # 부제목: (부동산의표시) (중앙, 2줄)
    c.setFont(font_name, 11)
    subtitle = "(부동산의표시)"
    subtitle_width = c.stringWidth(subtitle, font_name, 11)
    c.drawString(center_x - subtitle_width/2, 720, subtitle)
    
    # 부동산 표시 내용
    c.setFont(font_name, 10)
    estate_text = data.get('estate_text', '')
    estate_lines = estate_text.split('\n')
    estate_y = 695
    line_h = 13
    for i, line in enumerate(estate_lines[:22]):
        if line.strip():
            c.drawString(left_x, estate_y - (i * line_h), line)
    
    # 내용 영역
    c.setFont(font_name, 10)
    cancel_text = data.get('cancel_text', '')
    content_y = 395
    
    # 전체 내용 한 문장으로 구성
    full_content = f"위 부동산에 관하여 {cancel_text} (을)를 해지한다."
    
    # 폭 기반 줄바꿈 (약 90자 또는 폭 495pt 기준)
    max_chars = 90
    if len(full_content) > max_chars:
        # 첫 줄
        c.drawString(left_x, content_y, full_content[:max_chars])
        # 두번째 줄
        c.drawString(left_x, content_y - 16, full_content[max_chars:])
    else:
        c.drawString(left_x, content_y, full_content)
    
    # 작성일자 (중앙)
    date_text = data.get('date', '')
    c.setFont(font_name, 11)
    date_width = c.stringWidth(date_text, font_name, 11)
    c.drawString(center_x - date_width/2, 320, date_text)
    
    # 의무자 영역 - 중앙정렬, 라벨 좌측
    obligor_label = data.get('obligor_label', '근저당권자')
    obligor_name = data.get('obligor_name', '')
    obligor_id = data.get('obligor_id', '')
    obligor_addr = data.get('obligor_addr', '')
    obligor_rep = data.get('obligor_rep', '')
    obligor_branch = data.get('obligor_branch', '')
    
    # 라벨 (좌측)
    c.setFont(font_name, 10)
    c.drawString(left_x, 280, obligor_label)
    
    # 내용 (중앙) - 해지증서는 법인등록번호 포함
    c.setFont(font_name, 10)
    if obligor_id:
        obligor_display = f"{obligor_name}({obligor_id})"
    else:
        obligor_display = obligor_name
    
    text_width = c.stringWidth(obligor_display, font_name, 10)
    c.drawString(center_x - text_width/2, 260, obligor_display)
    
    c.setFont(font_name, 9)
    addr_width = c.stringWidth(obligor_addr, font_name, 9)
    c.drawString(center_x - addr_width/2, 245, obligor_addr)
    
    # 취급지점 + 대표자 형식
    if obligor_branch and obligor_rep:
        rep_text = f"(취급지점:{obligor_branch}) {obligor_rep}"
    elif obligor_rep:
        rep_text = obligor_rep
    else:
        rep_text = ""
    
    if rep_text:
        rep_width = c.stringWidth(rep_text, font_name, 10)
        c.setFont(font_name, 10)
        c.drawString(center_x - rep_width/2, 225, rep_text)
    
    # 권리자 (우측정렬)
    holder1_name = data.get('holder1_name', '')
    holder2_name = data.get('holder2_name', '')
    
    c.setFont(font_name, 10)
    if holder2_name:
        holder_text = f"{holder1_name},{holder2_name} 귀하"
    else:
        holder_text = f"{holder1_name} 귀하"
    
    holder_width = c.stringWidth(holder_text, font_name, 10)
    c.drawString(right_x - holder_width, 160, holder_text)
    
    c.showPage()
    c.save()
    packet.seek(0)
    return packet

def make_malso_transfer_pdf(data):
    """이관증명서 PDF 생성 (백지에서 생성)"""
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=A4)
    width, height = A4
    
    try:
        pdfmetrics.registerFont(TTFont('Korean', FONT_PATH))
        font_name = 'Korean'
    except:
        font_name = 'Helvetica'
    
    # 페이지 설정: 좌측 X=50, 우측 X=545 (여백 줄임)
    left_x = 50
    right_x = 545
    center_x = (left_x + right_x) / 2
    
    # 제목: 이 관 증 명 서 (중앙)
    c.setFont(font_name, 18)
    title = "이 관 증 명 서"
    title_width = c.stringWidth(title, font_name, 18)
    c.drawString(center_x - title_width/2, 750, title)
    
    # 부제목
    c.setFont(font_name, 11)
    subtitle = "(부동산의표시)"
    subtitle_width = c.stringWidth(subtitle, font_name, 11)
    c.drawString(center_x - subtitle_width/2, 720, subtitle)
    
    # 부동산 표시
    c.setFont(font_name, 10)
    estate_text = data.get('estate_text', '')
    estate_lines = estate_text.split('\n')
    estate_y = 695
    line_h = 13
    for i, line in enumerate(estate_lines[:22]):
        if line.strip():
            c.drawString(left_x, estate_y - (i * line_h), line)
    
    # 내용
    c.setFont(font_name, 10)
    cancel_text = data.get('cancel_text', '')
    from_branch = data.get('from_branch', '')
    to_branch = data.get('to_branch', '')
    
    content_y = 395
    
    # 전체 내용
    full_content1 = f"위 부동산에 관하여 {cancel_text}"
    full_content2 = f"업무일체가 {from_branch}에서 {to_branch}(으)로 이관되었음을 확인합니다."
    
    # 폭 기반 줄바꿈 (약 90자 기준)
    max_chars = 90
    if len(full_content1) > max_chars:
        c.drawString(left_x, content_y, full_content1[:max_chars])
        c.drawString(left_x, content_y - 16, full_content1[max_chars:])
        c.drawString(left_x, content_y - 32, full_content2)
    else:
        c.drawString(left_x, content_y, full_content1)
        c.drawString(left_x, content_y - 16, full_content2)
    
    # 작성일자 (중앙)
    date_text = data.get('date', '')
    c.setFont(font_name, 11)
    date_width = c.stringWidth(date_text, font_name, 11)
    c.drawString(center_x - date_width/2, 320, date_text)
    
    # 의무자 (중앙, 라벨 좌측)
    obligor_label = data.get('obligor_label', '근저당권자')
    obligor_name = data.get('obligor_name', '')
    obligor_id = data.get('obligor_id', '')
    obligor_addr = data.get('obligor_addr', '')
    obligor_rep = data.get('obligor_rep', '')
    obligor_branch = data.get('obligor_branch', '')
    
    c.setFont(font_name, 10)
    c.drawString(left_x, 280, obligor_label)
    
    # 이관증명서도 법인등록번호 포함
    if obligor_id:
        obligor_display = f"{obligor_name}({obligor_id})"
    else:
        obligor_display = obligor_name
    
    text_width = c.stringWidth(obligor_display, font_name, 10)
    c.drawString(center_x - text_width/2, 260, obligor_display)
    
    c.setFont(font_name, 9)
    addr_width = c.stringWidth(obligor_addr, font_name, 9)
    c.drawString(center_x - addr_width/2, 245, obligor_addr)
    
    # 취급지점 + 대표자 형식
    if obligor_branch and obligor_rep:
        rep_text = f"(취급지점:{obligor_branch}) {obligor_rep}"
    elif obligor_rep:
        rep_text = obligor_rep
    else:
        rep_text = ""
    
    if rep_text:
        rep_width = c.stringWidth(rep_text, font_name, 10)
        c.setFont(font_name, 10)
        c.drawString(center_x - rep_width/2, 225, rep_text)
    
    c.showPage()
    c.save()
    packet.seek(0)
    return packet

# =============================================================================
# 5. Streamlit UI 및 상태 관리
# =============================================================================

# 상태 변수 초기화
keys_to_init = [
    'base_fee_val', 'add_fee_val', 'etc_fee_val', 'disc_fee_val', 
    'cost_manual_제증명', 'cost_manual_교통비', 'cost_manual_원인증서', 
    'cost_manual_주소변경', 'cost_manual_확인서면', 'cost_manual_선순위 말소',
    'tax_등록면허세', 'tax_지방교육세', 'tax_증지대', 'tax_채권할인'
]
for k in keys_to_init:
    if k not in st.session_state: st.session_state[k] = "0"

if 'use_address_change' not in st.session_state: st.session_state['use_address_change'] = False
if 'address_change_count' not in st.session_state: st.session_state['address_change_count'] = 0
if 'addr_count_input' not in st.session_state: st.session_state['addr_count_input'] = 1

if 'calc_data' not in st.session_state:
    st.session_state['calc_data'] = {}
    st.session_state['show_fee'] = True
    st.session_state['input_amount'] = ""
    st.session_state['amount_raw_input'] = ""
    st.session_state['input_parcels'] = 1
    st.session_state['input_rate'] = f"{get_rate()*100:.5f}"
    st.session_state['input_debtor'] = "" # Tab 1과 동기화 위해 존재하지만 초기값은 빈값
    st.session_state['input_creditor'] = list(CREDITORS.keys())[0]
    st.session_state['input_creditor_name'] = ""
    st.session_state['input_creditor_corp_num'] = ""
    st.session_state['input_creditor_addr'] = ""
    st.session_state['input_collateral_addr'] = ""
    st.session_state['input_debtor_addr'] = ""
    st.session_state['input_owner'] = ""
    st.session_state['input_owner_addr'] = ""
    st.session_state['guarantee'] = "한정근담보"
    st.session_state['contract_type'] = "개인"
    st.session_state['input_date'] = datetime.now().date()
    st.session_state['estate_text'] = """[토지]\n서울특별시 강남구 대치동 123번지\n대 300㎡\n\n[건물]\n서울특별시 강남구 대치동 123번지\n철근콘크리트조 슬래브지붕 5층 주택\n1층 100㎡\n2층 100㎡"""
    st.session_state['input_debtor_rrn'] = ""
    st.session_state['input_owner_rrn'] = ""

# 3탭 수기 입력값 초기 상태 (금융사 기본값 로드)
manual_keys = ["cost_manual_제증명", "cost_manual_교통비", "cost_manual_원인증서", "cost_manual_확인서면", "cost_manual_선순위 말소"]
for key in manual_keys:
    if key not in st.session_state:
        # 이미 0으로 초기화했지만 금융사 변경 시 값 덮어쓰기 위해 여기서 체크
        pass # 아래 handle_creditor_change 등에서 처리

# =============================================================================
# 등기부 PDF 파싱 함수
# =============================================================================
def parse_registry_pdf(uploaded_file):
    """집합건물 등기부 PDF에서 부동산표시 추출 - 삭선 감지 포함"""
    
    # 행정구역 변환
    행정구역_변환 = {"전라북도": "전북특별자치도", "강원도": "강원특별자치도"}
    def convert_region(text):
        for old, new in 행정구역_변환.items():
            text = text.replace(old, new)
        return text
    
    def get_red_lines_from_page(page):
        """페이지에서 빨간 삭선 y좌표 수집"""
        red_ys = []
        for line in page.lines:
            color = line.get('stroking_color')
            if color and isinstance(color, (list, tuple)) and len(color) >= 3:
                r, g, b = color[0], color[1], color[2]
                if r > 0.9 and g < 0.1 and b < 0.1:
                    width = line['x1'] - line['x0']
                    if width > 30:
                        red_ys.append(line['top'])
        return red_ys
    
    def is_번지_strikethrough(번지, page_words, red_ys):
        """번지의 모든 위치에서 삭선 여부 확인"""
        if not red_ys or not 번지:
            return False
        
        found_positions = []
        for word in page_words:
            if 번지 in word['text']:
                word_y = word['top']
                # 삭선이 단어 위나 아래에 있을 수 있으므로 양방향 체크
                has_strike = any(abs(word_y - ry) < 15 for ry in red_ys)
                found_positions.append({'y': word_y, 'strike': has_strike})
        
        if not found_positions:
            return False
        
        # 삭선 없는 위치가 하나라도 있으면 False
        return all(pos['strike'] for pos in found_positions)
    
    result = {
        "1동건물표시": "",
        "아파트명": "",
        "동명칭": "",
        "도로명주소": "",
        "건물번호": "",
        "고유번호": "",
        "구조": "",
        "면적": "",
        "토지": [],
        "대지권종류": "",
        "대지권비율": ""
    }
    
    debug = {
        "errors": [],
        "warnings": [],
        "info": []
    }
    
    if not PDFPLUMBER_OK:
        debug["errors"].append("pdfplumber 라이브러리가 설치되지 않았습니다.")
        return result, debug
    
    try:
        with pdfplumber.open(uploaded_file) as pdf:
            debug["info"].append(f"PDF 페이지 수: {len(pdf.pages)}")
            
            # 고유번호 추출
            first_page_text = pdf.pages[0].extract_text() or ""
            고유번호_match = re.search(r'고유번호\s*(\d{4}-\d{4}-\d{6})', first_page_text)
            if 고유번호_match:
                result["고유번호"] = 고유번호_match.group(1)
            
            # 페이지별 빨간선/텍스트 정보 수집
            page_info = {}
            for page_idx, page in enumerate(pdf.pages):
                page_info[page_idx] = {
                    'red_ys': get_red_lines_from_page(page),
                    'words': page.extract_words()
                }
            
            # 컬러 PDF 여부
            has_color = any(len(info['red_ys']) > 0 for info in page_info.values())
            if has_color:
                debug["info"].append("컬러 PDF 감지 - 삭선 기반 필터링")
            else:
                debug["info"].append("흑백 PDF - 번지 기반 필터링")
            
            # 페이지별 테이블 수집
            all_tables = []
            for page_idx, page in enumerate(pdf.pages):
                for table in page.extract_tables():
                    all_tables.append((page_idx, table))
            
            # 섹션별 데이터 행 수집
            sections = {"1동건물": [], "토지": [], "전유부분": [], "대지권": []}
            current_section = None
            대지권_header = []
            
            for page_idx, table in all_tables:
                if not table:
                    continue
                
                for row in table:
                    row_text = str(row[0]) if row[0] else ""
                    
                    # 섹션 헤더 감지
                    if "1동의 건물의 표시" in row_text:
                        current_section = "1동건물"
                        continue
                    elif "대지권의 목적인 토지의 표시" in row_text:
                        current_section = "토지"
                        continue
                    elif "전유부분의 건물의 표시" in row_text:
                        current_section = "전유부분"
                        continue
                    elif "대지권의 표시" in row_text and "목적인 토지" not in row_text:
                        current_section = "대지권"
                        continue
                    elif "갑 구" in row_text or "을 구" in row_text:
                        current_section = None
                        continue
                    
                    # 대지권 헤더 저장
                    if current_section == "대지권" and row_text.strip() == "표시번호":
                        대지권_header = row
                        continue
                    
                    # 컬럼 헤더 스킵
                    if row_text.strip() in ["표시번호", "순위번호"]:
                        continue
                    
                    # 현행 데이터: "1", "2" 또는 "1\n(전 1)" 형태
                    if current_section and row[0]:
                        row0_clean = str(row[0]).strip()
                        if re.match(r'^\d+$', row0_clean) or re.match(r'^\d+\n\(전', row0_clean):
                            
                            # 토지 섹션: 삭선 감지 적용
                            if current_section == "토지":
                                소재지 = (row[1] or "") if len(row) > 1 else ""
                                
                                # 주소 패턴 체크
                                if not re.search(r'(시|군|구|동|리|읍|면)\s', 소재지):
                                    continue
                                
                                # 번지 추출
                                소재지_clean = 소재지.replace('\n', ' ').strip()
                                번지_match = re.search(r'(\d+(-\d+)?)$', 소재지_clean)
                                번지 = 번지_match.group(1) if 번지_match else None
                                
                                # 컬러 PDF면 삭선 체크
                                if has_color and 번지:
                                    info = page_info[page_idx]
                                    if is_번지_strikethrough(번지, info['words'], info['red_ys']):
                                        continue  # 말소 스킵
                                
                                sections["토지"].append((번지, row))
                            else:
                                sections[current_section].append(row)
            
            # ===== 1동건물: 마지막 유효 행 =====
            if sections["1동건물"]:
                row = sections["1동건물"][-1]
                col2 = (row[2] or "") if len(row) > 2 else ""
                
                # 워터마크 제거
                col2 = re.sub(r'열\s*람\s*용', '', col2)
                col2 = re.sub(r'(?<=[가-힣])(열|람|용)(?=[가-힣])', '', col2)
                
                lines = col2.split('\n')
                
                # [도로명주소] 위치 찾기
                road_idx = -1
                for i, line in enumerate(lines):
                    if '[도로명주소]' in line:
                        road_idx = i
                        break
                
                # 도로명주소 추출
                if road_idx > 0:
                    road_lines = []
                    for i in range(road_idx + 1, len(lines)):
                        line = lines[i].strip()
                        if line and line not in ['열', '람', '용']:
                            road_lines.append(line)
                    result["도로명주소"] = convert_region(' '.join(road_lines))
                
                # [도로명주소] 앞부분만 사용
                content_lines = lines[:road_idx] if road_idx > 0 else lines
                content_lines = [l.strip() for l in content_lines if l.strip() and l.strip() not in ['열', '람', '용']]
                
                if content_lines:
                    # 번지(숫자)로 끝나는 마지막 줄 = 지번 끝
                    지번_end_idx = -1
                    for i, line in enumerate(content_lines):
                        if re.search(r'\d+(-\d+)?$', line.strip()):
                            지번_end_idx = i
                    
                    if 지번_end_idx >= 0:
                        result["1동건물표시"] = convert_region(' '.join(content_lines[:지번_end_idx+1]))
                        
                        건물명_lines = content_lines[지번_end_idx+1:]
                        if 건물명_lines:
                            건물명_text = ' '.join(건물명_lines)
                            동_match = re.search(r'(제[가-힣\d]+동)$', 건물명_text)
                            if 동_match:
                                result["동명칭"] = 동_match.group(1)
                                result["아파트명"] = 건물명_text[:동_match.start()].strip()
                            else:
                                result["아파트명"] = 건물명_text
                    else:
                        result["1동건물표시"] = convert_region(' '.join(content_lines))
            
            # ===== 토지: 같은 번지면 마지막만 =====
            토지_by_번지 = {}
            for 번지, row in sections["토지"]:
                토지_by_번지[번지] = row
            
            # 번지 숫자순 정렬
            토지_items = sorted(토지_by_번지.items(), key=lambda x: (int(re.search(r'^(\d+)', x[0]).group(1)) if x[0] and re.search(r'^(\d+)', x[0]) else 0))
            
            for idx, (번지, row) in enumerate(토지_items, 1):
                소재지_raw = (row[1] or "").replace('\n', ' ').strip()
                
                # 지목과 면적: row[2:]에서 패턴으로 찾기 (pdfplumber 파싱 차이 대응)
                지목_raw = ""
                면적_raw = ""
                for col in row[2:]:
                    col_str = (col or "").strip()
                    if col_str:
                        # 지목: 대, 전, 답 등으로만 구성
                        if re.match(r'^(대|전|답|임야|잡종지|도로|하천)(\n(대|전|답|임야|잡종지|도로|하천))*$', col_str):
                            지목_raw = col_str
                        # 면적: ㎡ 포함
                        elif '㎡' in col_str:
                            면적_raw = col_str
                
                # 여러 필지가 한 행에 있는 경우 분리 (1. xxx 2. xxx 3. xxx 형태)
                필지_matches = re.split(r'(?=\d+\.\s*[가-힣])', 소재지_raw)
                필지_matches = [p.strip() for p in 필지_matches if p.strip()]
                
                if len(필지_matches) > 1:
                    # 지목과 면적도 분리
                    지목_list = re.findall(r'(대|전|답|임야|잡종지|도로|하천|공장용지|학교용지|주차장|창고용지|목장용지|광천지|염전|유지|양어장|수도용지|공원|체육용지|유원지|종교용지|사적지|묘지|주유소용지)', 지목_raw)
                    면적_list = re.findall(r'([\d.]+㎡)', 면적_raw)
                    
                    for i, 필지 in enumerate(필지_matches):
                        소재지 = re.sub(r'^\d+\.\s*', '', 필지).strip()
                        지목 = 지목_list[i] if i < len(지목_list) else (지목_list[0] if 지목_list else "")
                        면적 = 면적_list[i] if i < len(면적_list) else ""
                        
                        result["토지"].append({
                            "번호": str(len(result["토지"]) + 1),
                            "소재지": convert_region(소재지),
                            "지목": 지목,
                            "면적": 면적
                        })
                else:
                    # 단일 필지
                    소재지 = re.sub(r'^\d+\.\s*', '', 소재지_raw)
                    지목_match = re.search(r'(대|전|답|임야|잡종지)', 지목_raw)
                    지목 = 지목_match.group(1) if 지목_match else 지목_raw
                    면적_match = re.search(r'([\d.]+㎡)', 면적_raw)
                    면적 = 면적_match.group(1) if 면적_match else 면적_raw
                    
                    result["토지"].append({
                        "번호": str(len(result["토지"]) + 1),
                        "소재지": convert_region(소재지),
                        "지목": 지목,
                        "면적": 면적
                    })
            
            # ===== 전유부분: 마지막 유효 행 =====
            if sections["전유부분"]:
                row = sections["전유부분"][-1]
                건물번호 = (row[2] or "").replace('\n', ' ').strip() if len(row) > 2 else ""
                건물내역 = (row[3] or "").replace('\n', ' ').strip() if len(row) > 3 else ""
                
                if result["동명칭"] and result["동명칭"] not in 건물번호:
                    건물번호 = f"{result['동명칭']} {건물번호}"
                
                result["건물번호"] = 건물번호
                
                구조_match = re.search(r'([가-힣]+조|[가-힣]+구조)', 건물내역)
                면적_match = re.search(r'([\d.]+㎡)', 건물내역)
                result["구조"] = 구조_match.group(1) if 구조_match else ""
                result["면적"] = 면적_match.group(1) if 면적_match else ""
            
            # ===== 대지권: 소유권/지상권/전세권이 있는 마지막 유효 행 =====
            if sections["대지권"]:
                # 대지권종류가 있는 행 찾기
                valid_대지권_row = None
                for row in sections["대지권"]:
                    종류_raw = (row[1] or "").replace('\n', ' ').strip() if len(row) > 1 else ""
                    if re.search(r'(소유권|지상권|전세권)', 종류_raw):
                        valid_대지권_row = row
                
                if valid_대지권_row:
                    row = valid_대지권_row
                    
                    종류_raw = (row[1] or "").replace('\n', ' ').strip() if len(row) > 1 else ""
                    # "1, 2, 3 소유권대지권" → "소유권" 으로 단순화
                    종류_match = re.search(r'(소유권|지상권|전세권)', 종류_raw)
                    result["대지권종류"] = 종류_match.group(1) if 종류_match else 종류_raw
                    
                    # 대지권비율: "분의" 패턴이 있는 컬럼 찾기
                    for col in row[2:]:
                        col_str = (col or "").replace('\n', ' ').strip()
                        if "분의" in col_str:
                            result["대지권비율"] = col_str
                            break
            
            # 아파트명 없으면 갑구에서 찾기
            if not result["아파트명"]:
                full_text = "\n".join([p.extract_text() or "" for p in pdf.pages])
                갑구_match = re.search(r'【\s*갑\s*구\s*】(.+?)【\s*을\s*구\s*】', full_text, re.DOTALL)
                if 갑구_match:
                    아파트_match = re.search(r'([가-힣A-Za-z0-9]+(?:아파트|빌라|오피스텔|주상복합|타운|파크|힐스|뷰|애비뉴|타워|팰리스|하이츠))', 갑구_match.group(1))
                    if 아파트_match:
                        result["아파트명"] = 아파트_match.group(1)
    
    except Exception as e:
        debug["errors"].append(f"PDF 파싱 오류: {str(e)}")
        return result, debug
    
    return result, debug

def format_estate_text(data):
    """부동산 표시 포맷팅"""
    # 행정구역 변환
    행정구역_변환 = {"전라북도": "전북특별자치도", "강원도": "강원특별자치도"}
    def convert_region(text):
        for old, new in 행정구역_변환.items():
            text = text.replace(old, new)
        return text
    
    lines = []
    
    # 1동의 건물의 표시
    lines.append("1동의 건물의 표시")
    lines.append(f"   {data['1동건물표시']}")
    
    # 아파트명/동명칭
    건물명칭_parts = []
    if data["아파트명"]:
        건물명칭_parts.append(data["아파트명"])
    if data["동명칭"]:
        건물명칭_parts.append(data["동명칭"])
    if 건물명칭_parts:
        lines.append(f"   {' '.join(건물명칭_parts)}")
    
    lines.append("")  # 빈 줄
    
    # 전유부분의 건물의 표시
    lines.append("전유부분의 건물의 표시")
    lines.append(f"  1. 건물의 번호 : {data['건물번호']}[고유번호:{data['고유번호']}]")
    lines.append(f"      구조 및 면적 : {data['구조']} {data['면적']}")
    
    lines.append("")  # 빈 줄
    
    # 전유부분의 대지권의 표시
    lines.append("전유부분의 대지권의 표시")
    lines.append("      토지의 표시")
    
    # 토지는 최신 것만 (행정구역 변경된 경우 마지막 것만)
    if data["토지"]:
        # 마지막 토지만 사용 (최신)
        t = data["토지"][-1]
        소재지 = convert_region(t['소재지'])
        lines.append(f"       1.{소재지}")
        lines.append(f"           {t['지목']}  {t['면적']}")
    
    lines.append(f"          대지권의 종류: {data['대지권종류']}")
    lines.append(f"          대지권의 비율: {data['대지권비율']}")
    
    return "\n".join(lines)


def show_debug(debug):
    """디버깅 정보 표시 - 접이식"""
    total_errors = len(debug["errors"])
    total_warnings = len(debug["warnings"])
    
    if total_errors > 0:
        st.error(f"❌ 오류 {total_errors}건 발생")
    elif total_warnings > 0:
        st.warning(f"⚠️ 경고 {total_warnings}건 (일부 항목 추출 실패)")
    else:
        st.success("✅ 모든 항목 추출 성공")
    
    with st.expander("🔍 상세 파싱 결과 보기"):
        if debug["errors"]:
            st.markdown("**🔴 오류:**")
            for e in debug["errors"]:
                st.markdown(f"- {e}")
        
        if debug["warnings"]:
            st.markdown("**🟡 경고:**")
            for w in debug["warnings"]:
                st.markdown(f"- {w}")
        
        if debug["info"]:
            st.markdown("**🟢 추출 성공:**")
            for i in debug["info"]:
                st.markdown(f"- {i}")


# =============================================================================
# 위택스 API 호출 함수
# =============================================================================
WETAX_API_URL_DEFAULT = "http://localhost:8000"

def call_wetax_api(cases, base_url=None):
    """위택스 API 호출"""
    if not REQUESTS_OK:
        return None, "requests 라이브러리가 설치되지 않았습니다."
    
    # URL 결정
    if base_url:
        api_url = base_url.rstrip('/') + "/wetax/submit"
    else:
        api_url = WETAX_API_URL_DEFAULT + "/wetax/submit"
    
    try:
        response = requests.post(api_url, json={"cases": cases}, timeout=120)
        if response.status_code == 200:
            return response.json(), None
        else:
            return None, f"API 오류: {response.status_code}"
    except requests.exceptions.ConnectionError:
        return None, "위택스 서버에 연결할 수 없습니다. 서버가 실행 중인지 확인하세요."
    except Exception as e:
        return None, str(e)


def parse_corp_num(corp_num_str):
    """법인번호 분리 (110111-4138560 → 앞6자리, 뒤7자리)"""
    clean = re.sub(r'[^0-9]', '', str(corp_num_str))
    if len(clean) >= 13:
        return clean[:6], clean[6:13]
    elif len(clean) >= 6:
        return clean[:6], clean[6:]
    return clean, ""


def parse_rrn(rrn_str):
    """주민번호 분리 (800101-1234567 → 앞6자리, 뒤7자리)"""
    clean = re.sub(r'[^0-9]', '', str(rrn_str))
    if len(clean) >= 13:
        return clean[:6], clean[6:13]
    elif len(clean) >= 6:
        return clean[:6], clean[6:]
    return clean, ""


def extract_road_address(full_address):
    """전체 주소에서 도로명 추출 (상세주소 제외)"""
    if not full_address:
        return "", ""
    
    # 쉼표로 분리
    parts = full_address.split(',')
    if len(parts) >= 2:
        return parts[0].strip(), parts[1].strip()
    
    # 숫자 뒤 공백으로 분리 시도
    match = re.match(r'(.+?(?:로|길)\s*\d+(?:-\d+)?)\s*(.*)$', full_address)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    
    return full_address, ""

def parse_int_input(text_input):
    try:
        if isinstance(text_input, int): return text_input
        return int(remove_commas(text_input or "0"))
    except ValueError: return 0

def handle_creditor_change():
    creditor_key = st.session_state.get('t1_creditor_select', list(CREDITORS.keys())[0])
    if creditor_key == "🖊️ 직접입력":
        # 수기입력 항목들 0으로 초기화
        for k in manual_keys: st.session_state[k] = "0"
        st.session_state['cost_manual_주소변경'] = "0"
        st.session_state['input_creditor_name'] = ""
        st.session_state['input_creditor_corp_num'] = ""
        st.session_state['input_creditor_addr'] = ""
    else:
        # 유노스프레스티지일 경우 제증명만 20,000원, 교통비/원인증서는 0
        if "(주)유노스프레스티지대부" in creditor_key:
            st.session_state['cost_manual_제증명'] = "20,000"
            st.session_state['cost_manual_교통비'] = "0"
            st.session_state['cost_manual_원인증서'] = "0"
            st.session_state['cost_manual_확인서면'] = "0"
            st.session_state['cost_manual_선순위 말소'] = "0"
        else:
            # 기타 금융사: 제증명 50,000, 교통비 100,000, 원인증서 50,000
            st.session_state['cost_manual_제증명'] = "50,000"
            st.session_state['cost_manual_교통비'] = "100,000"
            st.session_state['cost_manual_원인증서'] = "50,000"
            st.session_state['cost_manual_확인서면'] = "0"
            st.session_state['cost_manual_선순위 말소'] = "0"
        st.session_state['cost_manual_주소변경'] = "0" # 주소변경은 체크박스로만 제어
    st.session_state.calc_data['creditor_key_check'] = creditor_key

MANUAL_COST_NAMES = ["제증명", "교통비", "원인증서", "주소변경", "확인서면", "선순위 말소"]

def calculate_all(data):
    amount = parse_int_input(data.get('채권최고액')) 
    parcels = parse_int_input(data.get('필지수'))
    try: rate = float(remove_commas(data.get('채권할인율', '0'))) / 100
    except ValueError: rate = 0 
    
    # 원본 데이터 보존
    data['input_amount'] = data.get('채권최고액', '')
    
    # 금융사 확인 (엘하비스트대부 체크)
    creditor = st.session_state.get('tab3_creditor_select', '')
    is_elharvest = (creditor == "㈜엘하비스트대부 대표이사 김상수")
    
    # 기본료 계산
    if is_elharvest:
        # 엘하비스트대부는 항상 70,000원 고정
        base_fee = 70000
        data['기본료_자동'] = 70000
    else:
        # 일반 금융사: 수기입력 우선, 없으면 자동계산
        manual_base_fee = parse_int_input(data.get('기본료_val', 0))
        if manual_base_fee > 0:
            base_fee = manual_base_fee
        else:
            base_fee = lookup_base_fee(amount)
        data['기본료_자동'] = lookup_base_fee(amount)
    data['기본료'] = base_fee
    
    add_fee = parse_int_input(data.get('추가보수_val'))
    etc_fee = parse_int_input(data.get('기타보수_val'))
    disc_fee = parse_int_input(data.get('할인금액'))

    fee_total = 0
    if st.session_state['show_fee']:
        supply_val = base_fee + add_fee + etc_fee - disc_fee
        vat = math.floor(max(0, supply_val) * 0.1)
        fee_total = supply_val + vat
        data['공급가액'] = supply_val
        data['부가세'] = vat
        data['보수총액'] = fee_total
    else:
        data['공급가액'] = 0; data['부가세'] = 0; data['보수총액'] = 0
    
    # 공과금 계산
    addr_count = int(st.session_state.get('address_change_count', 0))  # 계산 로직에서 미리 설정됨
    
    addr_reg = 0; addr_edu = 0; addr_jeungji = 0
    if addr_count > 0:
        addr_reg = 6000 * addr_count
        addr_edu = 1200 * addr_count
        addr_jeungji = 3000 * addr_count
    
    basic_reg = floor_10(amount * 0.002)
    basic_edu = floor_10(basic_reg * 0.2)
    final_reg = basic_reg + addr_reg
    final_edu = basic_edu + addr_edu
    jeungji = (18000 * parcels) + addr_jeungji 

    bond = 0
    if amount >= 20_000_000: bond = math.ceil(amount * 0.01 / 10000) * 10000
    bond_disc = floor_10(bond * rate)
    
    data["등록면허세"] = final_reg
    data["지방교육세"] = final_edu
    data["증지대"] = jeungji
    data["채권할인금액"] = bond_disc
    
    cost_total = final_reg + final_edu + jeungji + bond_disc
    for k in MANUAL_COST_NAMES:
        cost_total += parse_int_input(st.session_state.get('cost_manual_' + k, 0))
    
    data['공과금 총액'] = cost_total
    data['총 합계'] = fee_total + cost_total
    return data

def create_receipt_excel(data, template_path=None):
    """영수증 Excel 파일 생성 - 템플릿 기반"""
    if not EXCEL_OK:
        return None
    
    # 기본 정보 입력
    client = data.get('client', {})
    full_creditor = client.get('금융사', '')
    
    # ===== 엘하비스트대부 전용 처리 =====
    if full_creditor == "㈜엘하비스트대부 대표이사 김상수":
        # receipt_template1.xlsx 사용
        template_dir = os.path.dirname(template_path) if template_path else APP_ROOT
        elharvest_template = os.path.join(template_dir, "receipt_template1.xlsx")
        
        if os.path.exists(elharvest_template):
            try:
                workbook = openpyxl.load_workbook(elharvest_template)
                ws = workbook.active
                
                # 채권최고액 (숫자만 추출)
                amount_str = client.get('채권최고액', '0')
                amount_val = int(re.sub(r'[^\d]', '', amount_str)) if amount_str else 0
                
                # 공과금 항목
                cost_items = data.get('cost_items', {})
                cost_total = int(data.get('cost_totals', {}).get('공과금 총액', 0))
                
                # 보수료 고정값
                fee_amount = 70000        # 보수액
                vat_amount = 7000         # 부가가치세
                fee_total = fee_amount + vat_amount  # 보수료 합계
                
                # ===== 기본 정보 =====
                ws['C5'] = client.get('채무자', '')       # 채무자
                ws['D7'] = client.get('물건지', '')       # 소재지
                ws['G8'] = amount_val                     # 채권최고액
                
                # ===== 보수료 항목 =====
                ws['D11'] = fee_amount                    # 보수료
                ws['D21'] = fee_amount                    # 보수료소계
                ws['D22'] = vat_amount                    # 부가가치세
                ws['D23'] = fee_total                     # 보수료소계(합계)
                
                # ===== 공과금 항목 =====
                ws['G11'] = int(cost_items.get('등록면허세', 0))      # 등록면허세
                ws['G12'] = int(cost_items.get('지방교육세', 0))      # 지방교육세
                ws['G13'] = int(cost_items.get('증지대', 0))          # 등기신청수수료
                ws['G14'] = int(cost_items.get('채권할인', 0))        # 채권할인액
                ws['G23'] = cost_total                                # 공과금합계
                
                # ===== 총합계 및 작성일자 =====
                ws['F24'] = fee_total + cost_total                    # 총합계
                ws['C27'] = data.get('date_input', '')                # 작성일자
                
                output = BytesIO()
                workbook.save(output)
                output.seek(0)
                return output
                
            except Exception as e:
                # 템플릿 사용 실패 시 기본 템플릿으로 폴백
                pass
    
    # ===== 기존 템플릿 처리 (일반 금융사) =====
    if template_path and os.path.exists(template_path):
        try:
            workbook = openpyxl.load_workbook(template_path)
            ws = workbook.active
            
            # 금융사명에서 회사명만 추출 (대표이사/사내이사 앞까지)
            if '대표이사' in full_creditor:
                company_name = full_creditor.split('대표이사')[0].strip()
            elif '사내이사' in full_creditor:
                company_name = full_creditor.split('사내이사')[0].strip()
            else:
                company_name = full_creditor
            
            # ===== 좌측 영수증 =====
            ws['B4'] = company_name                      # 금융사 (회사명만)
            
            # 채권최고액 (숫자만 추출)
            amount_str = client.get('채권최고액', '0')
            amount_val = int(re.sub(r'[^\d]', '', amount_str)) if amount_str else 0
            ws['M5'] = amount_val                        # 좌측 채권최고액
            
            ws['E7'] = client.get('물건지', '')           # 좌측 물건지
            
            # ===== 우측 영수증 =====
            ws['V4'] = client.get('채무자', '')           # 채무자
            ws['AG5'] = amount_val                       # 우측 채권최고액
            ws['Y7'] = client.get('물건지', '')           # 우측 물건지
            
            # ===== 작성일자 (좌/우 둘 다) =====
            date_str = data.get('date_input', '')
            ws['A24'] = date_str                         # 좌측 날짜
            ws['U24'] = date_str                         # 우측 날짜
            
            # ===== 공과금 항목 입력 (웹에서 계산된 값 직접 입력) =====
            cost_items = data.get('cost_items', {})
            
            # AH11~AH18: 고정 항목
            ws['AH11'] = int(cost_items.get('등록면허세', 0))
            ws['AH12'] = int(cost_items.get('지방교육세', 0))
            ws['AH13'] = int(cost_items.get('증지대', 0))          # 등기신청수수료
            ws['AH14'] = int(cost_items.get('채권할인', 0))        # 채권할인액
            ws['AH15'] = int(cost_items.get('제증명', 0))          # 등본/제증명
            ws['AH16'] = int(cost_items.get('원인증서', 0))
            ws['AH17'] = int(cost_items.get('주소변경', 0))
            ws['AH18'] = int(cost_items.get('선순위말소', 0))      # 선순위 말소
            
            # AD19/AH19: 교통비 (값이 있을 때만 라벨+금액 입력)
            traffic_fee = int(cost_items.get('교통비', 0))
            if traffic_fee > 0:
                ws['AD19'] = '교통비'
                ws['AH19'] = traffic_fee
            else:
                ws['AD19'] = None
                ws['AH19'] = None
            
            # AD20/AH20: 확인서면 (값이 있을 때만 라벨+금액 입력)
            confirm_fee = int(cost_items.get('확인서면', 0))
            if confirm_fee > 0:
                ws['AD20'] = '확인서면'
                ws['AH20'] = confirm_fee
            else:
                ws['AD20'] = None
                ws['AH20'] = None
            
            # ===== 소계/총계 (웹에서 계산된 값 직접 입력) =====
            cost_total = int(data.get('cost_totals', {}).get('공과금 총액', 0))
            ws['AH21'] = cost_total                      # 우측 공과금 소계
            ws['Y22'] = cost_total                       # 우측 총계
            
        except Exception as e:
            # 템플릿 사용 실패 시 새로 생성
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "영수증"
            _create_simple_receipt(ws, data)
    else:
        # 템플릿 없이 새로 생성
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "영수증"
        _create_simple_receipt(ws, data)
    
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def _create_simple_receipt(sheet, data):
    """간단한 영수증 시트 생성"""
    from openpyxl.styles import Font, Alignment, Border, Side
    
    # 제목
    sheet['A1'] = '근저당권설정 영수증'
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A1:C1')
    
    # 날짜
    sheet['A3'] = '작성일:'
    sheet['B3'] = data.get('date_input', '')
    
    # 고객 정보
    client = data.get('client', {})
    sheet['A5'] = '채무자:'
    sheet['B5'] = client.get('채무자', '')
    sheet['A6'] = '물건지:'
    sheet['B6'] = client.get('물건지', '')
    sheet['A7'] = '채권최고액:'
    sheet['B7'] = client.get('채권최고액', '')
    
    # 비용 항목
    row = 9
    sheet[f'A{row}'] = '항목'
    sheet[f'B{row}'] = '금액'
    sheet[f'A{row}'].font = Font(bold=True)
    sheet[f'B{row}'].font = Font(bold=True)
    
    row += 1
    cost_items = data.get('cost_items', {})
    for name, value in cost_items.items():
        if value != 0:
            sheet[f'A{row}'] = name
            sheet[f'B{row}'] = f"{int(value):,} 원"
            row += 1
    
    # 합계
    row += 1
    sheet[f'A{row}'] = '총 합계'
    sheet[f'B{row}'] = f"{data.get('grand_total', 0):,} 원"
    sheet[f'A{row}'].font = Font(bold=True, size=12)
    sheet[f'B{row}'].font = Font(bold=True, size=12)
    
    # 열 너비 조정
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 15


# =============================================================================
# UI 구현
# =============================================================================

tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(["🏦 시중은행", "📄 대부업(전자설정)", "🗑️ 말소문서", "✍️ 자필서명정보", "🧾 비용계산(대부업)", "🏛️ 위택스신고", "🏢 법인등기"])

# Tab 2: 대부업(전자설정)
with tab2:
    col_header = st.columns([6, 1])
    col_header[0].markdown("### 📝 대부업(전자설정) 작성")
    with col_header[1]:
        if st.button("🔄 초기화", type="secondary", key="reset_tab1", use_container_width=True, help="모든 입력 초기화"):
            # 날짜
            st.session_state['input_date'] = datetime.now().date()
            st.session_state['date_picker'] = datetime.now().date()
            
            # 채권자 관련
            st.session_state['input_creditor'] = list(CREDITORS.keys())[0]
            st.session_state['t1_creditor_select'] = list(CREDITORS.keys())[0]
            st.session_state['input_creditor_name'] = ''
            st.session_state['input_creditor_corp_num'] = ''
            st.session_state['input_creditor_addr'] = ''
            st.session_state['direct_creditor_name'] = ''
            st.session_state['direct_corp_num'] = ''
            st.session_state['direct_creditor_addr'] = ''
            
            # 채무자 관련 (위젯 키 포함)
            st.session_state['t1_debtor_name'] = ''
            st.session_state['t1_debtor_addr'] = ''
            st.session_state['t1_debtor_rrn'] = ''
            st.session_state['input_debtor_rrn'] = ''
            
            # 소유자 관련 (위젯 키 포함)
            st.session_state['t1_owner_name'] = ''
            st.session_state['t1_owner_addr'] = ''
            st.session_state['t1_owner_rrn'] = ''
            st.session_state['input_owner_rrn'] = ''
            
            # 계약 유형 (위젯 키 포함)
            st.session_state['contract_type'] = '개인'
            st.session_state['contract_type_radio'] = '개인'
            st.session_state['guarantee'] = '한정근담보'
            
            # 금액 (위젯 키 포함)
            st.session_state['amount_raw_input'] = ''
            st.session_state['input_amount'] = ''
            
            # 물건지 (위젯 키 포함)
            st.session_state['input_collateral_addr'] = ''
            st.session_state['collateral_addr_input'] = ''
            st.session_state['collateral_addr_widget'] = ''
            default_estate = """[토지]\n서울특별시 강남구 대치동 123번지\n대 300㎡\n\n[건물]\n서울특별시 강남구 대치동 123번지\n철근콘크리트조 슬래브지붕 5층 주택\n1층 100㎡\n2층 100㎡"""
            st.session_state['estate_text'] = default_estate
            st.session_state['estate_text_area'] = default_estate
            
            # 2탭 관련 session_state 초기화
            st.session_state['tab2_owner1_name'] = ''
            st.session_state['tab2_owner1_name_input'] = ''
            st.session_state['tab2_owner1_rrn'] = ''
            st.session_state['tab2_owner1_rrn_input'] = ''
            st.session_state['tab2_owner2_name'] = ''
            st.session_state['tab2_owner2_name_input'] = ''
            st.session_state['tab2_owner2_rrn'] = ''
            st.session_state['tab2_owner2_rrn_input'] = ''
            st.session_state['tab2_estate'] = ''
            st.session_state['tab2_estate_input'] = ''
            
            # 3탭 관련 session_state 초기화
            st.session_state['tab3_debtor_input'] = ''
            st.session_state['tab3_estate_input'] = ''
            st.session_state['calc_amount_input'] = ''
            st.session_state['calc_debtor_view'] = ''
            st.session_state['calc_estate_view'] = ''
            st.session_state['calc_creditor_view'] = ''
            st.session_state['calc_data'] = {}
            st.session_state['input_parcels'] = 1
            
            # 수기입력 비용 초기화
            st.session_state['cost_manual_제증명'] = '0'
            st.session_state['cost_manual_교통비'] = '0'
            st.session_state['cost_manual_원인증서'] = '0'
            st.session_state['cost_manual_주소변경'] = '0'
            st.session_state['cost_manual_확인서면'] = '0'
            st.session_state['cost_manual_선순위 말소'] = '0'
            st.session_state['use_address_change'] = False
            st.session_state['address_change_count'] = 1
            st.session_state['add_fee_val'] = '0'
            st.session_state['etc_fee_val'] = '0'
            st.session_state['disc_fee_val'] = '0'
            
            st.success("✅ 초기화되었습니다!")
            st.rerun()
    st.markdown("---")
    
    with st.expander("📌 기본 정보", expanded=True):
        current_date = st.session_state.get('input_date')
        if not isinstance(current_date, date): current_date = datetime.now().date()
        st.session_state['input_date'] = st.date_input("작성일자", value=current_date, key='date_picker')

    with st.expander("👤 당사자 정보", expanded=True):
        creditor_list = list(CREDITORS.keys()) + ["🖊️ 직접입력"]
        current_creditor = st.session_state.get('input_creditor', creditor_list[0])
        default_index = creditor_list.index(current_creditor) if current_creditor in creditor_list else 0
        selected_creditor = st.selectbox("채권자 선택", options=creditor_list, index=default_index, key='t1_creditor_select', on_change=handle_creditor_change)
        st.session_state['input_creditor'] = selected_creditor
        
        if selected_creditor == "🖊️ 직접입력":
            def on_direct_name_change():
                st.session_state['input_creditor_name'] = st.session_state.get('direct_creditor_name', '')
            def on_direct_corp_change():
                st.session_state['input_creditor_corp_num'] = st.session_state.get('direct_corp_num', '')
            def on_direct_addr_change():
                st.session_state['input_creditor_addr'] = st.session_state.get('direct_creditor_addr', '')
            
            st.text_input("채권자 성명/상호", value=st.session_state.get('input_creditor_name', ''), key='direct_creditor_name', on_change=on_direct_name_change)
            st.text_input("법인번호", value=st.session_state.get('input_creditor_corp_num', ''), key='direct_corp_num', on_change=on_direct_corp_change)
            st.text_area("채권자 주소", value=st.session_state.get('input_creditor_addr', ''), key='direct_creditor_addr', height=100, on_change=on_direct_addr_change)
            
            # 현재 값도 동기화
            st.session_state['input_creditor_name'] = st.session_state.get('direct_creditor_name', st.session_state.get('input_creditor_name', ''))
            st.session_state['input_creditor_corp_num'] = st.session_state.get('direct_corp_num', st.session_state.get('input_creditor_corp_num', ''))
            st.session_state['input_creditor_addr'] = st.session_state.get('direct_creditor_addr', st.session_state.get('input_creditor_addr', ''))
        else:
            creditor_info = CREDITORS.get(selected_creditor, {})
            st.text_input("법인번호", value=creditor_info.get('corp_num', ''), disabled=False)
            st.text_area("채권자 주소", value=creditor_info.get('addr', ''), disabled=False)
            st.session_state['input_creditor_name'] = selected_creditor
            st.session_state['input_creditor_corp_num'] = creditor_info.get('corp_num', '')
            st.session_state['input_creditor_addr'] = creditor_info.get('addr', '')

        # 주민번호 자동 하이픈 삽입 함수
        def auto_format_rrn_input(key):
            """6자리 입력 시 자동으로 '-' 삽입"""
            if key in st.session_state:
                val = st.session_state[key]
                # 숫자만 추출
                clean_val = re.sub(r'[^0-9]', '', str(val))
                # 6자리 이상이면 하이픈 삽입
                if len(clean_val) >= 6 and '-' not in val:
                    st.session_state[key] = f"{clean_val[:6]}-{clean_val[6:13]}"
                elif len(clean_val) > 13:
                    st.session_state[key] = f"{clean_val[:6]}-{clean_val[6:13]}"
        
        # 채무자 정보
        st.markdown("**채무자**")
        debtor_col1, debtor_col2 = st.columns([2, 1])
        with debtor_col1:
            st.text_input("채무자 성명", value=st.session_state.get('t1_debtor_name', ''), key='t1_debtor_name')
        with debtor_col2:
            st.text_input("주민등록번호", value=st.session_state.get('t1_debtor_rrn', ''), key='t1_debtor_rrn', placeholder="000000-0000000", on_change=auto_format_rrn_input, args=('t1_debtor_rrn',))
        st.text_area("채무자 주소", value=st.session_state.get('t1_debtor_addr', ''), key='t1_debtor_addr', height=100)
        
        # 설정자(소유자) 정보
        st.markdown("**설정자(소유자)**")
        owner_col1, owner_col2 = st.columns([2, 1])
        with owner_col1:
            st.text_input("설정자 성명", value=st.session_state.get('t1_owner_name', ''), key='t1_owner_name')
        with owner_col2:
            st.text_input("주민등록번호", value=st.session_state.get('t1_owner_rrn', ''), key='t1_owner_rrn', placeholder="000000-0000000", on_change=auto_format_rrn_input, args=('t1_owner_rrn',))
        st.text_area("설정자 주소", value=st.session_state.get('t1_owner_addr', ''), key='t1_owner_addr', height=100)

    with st.expander("🤝 담보 및 계약 정보", expanded=True):
        st.session_state['contract_type'] = st.radio("계약서 유형", options=["개인", "3자담보", "공동담보"], horizontal=True, key='contract_type_radio')
        
        # 피담보채무 버튼 선택
        st.write("**피담보채무**")
        col_guarantee1, col_guarantee2 = st.columns(2)
        if 'guarantee' not in st.session_state:
            st.session_state['guarantee'] = "한정근담보"
        
        with col_guarantee1:
            if st.button("한정근담보", type="primary" if st.session_state['guarantee']=="한정근담보" else "secondary", use_container_width=True, key="btn_guarantee_1"):
                st.session_state['guarantee'] = "한정근담보"
                st.rerun()
        with col_guarantee2:
            if st.button("포괄근담보", type="primary" if st.session_state['guarantee']=="포괄근담보" else "secondary", use_container_width=True, key="btn_guarantee_2"):
                st.session_state['guarantee'] = "포괄근담보"
                st.rerun()
        
        def format_amount_on_change():
            raw_val = st.session_state.get('amount_raw_input', '')
            formatted = format_number_with_comma(raw_val)
            st.session_state['input_amount'] = formatted # 3탭에서 이 변수를 참조함
            st.session_state['amount_raw_input'] = formatted
        
        st.text_input("채권최고액", key='amount_raw_input', on_change=format_amount_on_change, placeholder="숫자만 입력")
        if st.session_state.get('input_amount') and st.session_state['input_amount'] != "0":
            st.info(f"💰 **{number_to_korean(remove_commas(st.session_state['input_amount']))}**")
        
        col_addr1, col_addr2 = st.columns([5, 1])
        def copy_debtor_address():
            if st.session_state.get('t1_debtor_addr'):
                st.session_state['_pending_collateral_addr'] = st.session_state['t1_debtor_addr']
        def copy_from_estate():
            # 부동산표시에서 물건지주소 추출 (도로명주소 + 동 + 호)
            estate_text = st.session_state.get('estate_text', '')
            if not estate_text.strip():
                st.session_state['_toast_msg'] = "⚠️ 부동산표시가 비어있습니다"
                return
            
            import re
            
            # 동 추출 (첫줄에서)
            lines = estate_text.strip().split('\n')
            동 = ""
            for line in lines[:3]:
                동_match = re.search(r'제?(\d+)동', line)
                if 동_match:
                    동 = f"{동_match.group(1)}동"
                    break
            
            # 호 추출 (전유부분에서)
            호 = ""
            호_match = re.search(r'제(\d+)호', estate_text)
            if 호_match:
                호 = f"{호_match.group(1)}호"
            
            # 1. 도로명주소 우선
            if '[도로명주소]' in estate_text:
                match = re.search(r'\[도로명주소\]\s*(.+?)(?:\n|$)', estate_text)
                if match:
                    주소 = match.group(1).strip()
                    물건지주소 = f"{주소} {동} {호}".strip()
                    물건지주소 = re.sub(r'\s+', ' ', 물건지주소)
                    st.session_state['_pending_collateral_addr'] = 물건지주소
                    st.session_state['_toast_msg'] = "✅ 도로명주소 추출 완료"
                    return
            
            # 2. 지번주소에서 추출 (토지의 표시 첫번째)
            지번_match = re.search(r'토지의 표시\s*\n\s*1\.\s*(.+?)(?:\n|$)', estate_text)
            if 지번_match:
                주소 = 지번_match.group(1).strip()
                물건지주소 = f"{주소} {동} {호}".strip()
                물건지주소 = re.sub(r'\s+', ' ', 물건지주소)
                st.session_state['_pending_collateral_addr'] = 물건지주소
                st.session_state['_toast_msg'] = "✅ 지번주소 추출 완료"
                return
            
            st.session_state['_toast_msg'] = "⚠️ 주소를 찾을 수 없습니다"
        
        with col_addr1:
            # pending 값이 있으면 적용 (widget key도 함께 업데이트해야 함!)
            if '_pending_collateral_addr' in st.session_state:
                new_addr = st.session_state.pop('_pending_collateral_addr')
                st.session_state['input_collateral_addr'] = new_addr
                st.session_state['collateral_addr_widget'] = new_addr  # widget key 업데이트
            
            collateral_input = st.text_area(
                "물건지주소 (수기입력가능)", 
                value=st.session_state.get('input_collateral_addr', ''),
                height=100,
                key='collateral_addr_widget'
            )
            st.session_state['input_collateral_addr'] = collateral_input
        with col_addr2:
            st.button("📋 채무자 주소복사", key='copy_debtor_addr_btn', on_click=copy_debtor_address, use_container_width=True)
            st.button("🏠 부동산표시에서 추출", key='copy_estate_addr_btn', on_click=copy_from_estate, use_container_width=True)
        
        # 토스트 메시지 표시
        if st.session_state.get('_toast_msg'):
            st.toast(st.session_state['_toast_msg'])
            del st.session_state['_toast_msg']

    st.markdown("---")
    st.markdown("### 🏠 부동산의 표시")
    
    # 등기부 PDF 업로드
    col_upload, col_help = st.columns([3, 1])
    with col_upload:
        uploaded_registry = st.file_uploader("📤 등기부등본 PDF 업로드 (인터넷등기소 열람용)", type=['pdf'], key='registry_upload_tab1')
    with col_help:
        st.caption("※ 집합건물(아파트) 등기부만 지원")
    
    if uploaded_registry:
        if st.button("📋 부동산표시 추출", key='extract_estate_btn', use_container_width=True):
            with st.spinner("등기부 분석 중..."):
                data, debug = parse_registry_pdf(uploaded_registry)
                
                # 디버그 정보를 session_state에 저장
                st.session_state['estate_debug'] = debug
                
                if debug["errors"]:
                    pass  # 오류가 있으면 추출 결과 사용 안함
                else:
                    formatted = format_estate_text(data)
                    st.session_state['estate_text'] = formatted
                    st.session_state['estate_text_area'] = formatted
                    
                    # 위택스용 물건지 주소 자동 채움 (도로명주소 + 동 + 호)
                    # 동명칭: 제1동 → 1동
                    동 = data.get("동명칭", "").replace("제", "") if data.get("동명칭") else ""
                    # 전유부분에서 호수 추출: 제5층 제507호 → 507호
                    호_match = re.search(r'제(\d+)호', data.get("전유부분", ""))
                    호 = f"{호_match.group(1)}호" if 호_match else ""
                    
                    if data["도로명주소"]:
                        물건지주소 = f"{data['도로명주소']} {동} {호}".strip()
                        # 연속 공백 제거
                        물건지주소 = re.sub(r'\s+', ' ', 물건지주소)
                        st.session_state['_pending_collateral_addr'] = 물건지주소
                    
                st.rerun()
    
    # 디버깅 정보 표시 (session_state에서)
    if 'estate_debug' in st.session_state:
        show_debug(st.session_state['estate_debug'])
    
    st.caption("※ 등기부등본 내용을 입력하세요")
    col_estate, col_pdf = st.columns([3, 1])
    with col_estate:
        st.session_state['estate_text'] = st.text_area("부동산 표시 내용", value=st.session_state['estate_text'], height=300, key='estate_text_area', label_visibility="collapsed")
    with col_pdf:
        st.markdown("#### 📑 파일 생성")
        selected_template_path = st.session_state['template_status'].get(st.session_state['contract_type'])
        if selected_template_path: st.success(f"✅ 템플릿 준비완료"); is_disabled = False
        else: st.warning(f"⚠️ 템플릿 없음"); is_disabled = True
        
        if st.button("🚀 계약서\nPDF 생성", key="generate_pdf_tab1", disabled=is_disabled or not LIBS_OK, use_container_width=True):
            if not LIBS_OK: st.error("PDF 라이브러리 미설치")
            else:
                creditor_name_for_pdf = st.session_state['input_creditor'] if st.session_state['input_creditor'] != "🖊️ 직접입력" else st.session_state.get('input_creditor_name', '')
                creditor_addr_for_pdf = CREDITORS.get(st.session_state['input_creditor'], {}).get('addr', '') if st.session_state['input_creditor'] != "🖊️ 직접입력" else st.session_state.get('input_creditor_addr', '')
                
                data = {
                    "date": format_date_korean(st.session_state['input_date']), "creditor_name": creditor_name_for_pdf, "creditor_addr": creditor_addr_for_pdf,
                    "debtor_name": st.session_state.get('t1_debtor_name', ''), "debtor_addr": st.session_state.get('t1_debtor_addr', ''),
                    "owner_name": st.session_state.get('t1_owner_name', ''), "owner_addr": st.session_state.get('t1_owner_addr', ''),
                    "guarantee_type": st.session_state['guarantee'], "claim_amount": convert_multiple_amounts_to_korean(remove_commas(st.session_state['input_amount'])),
                    "estate_list": st.session_state['estate_text'].strip().split("\n"), "contract_type": st.session_state['contract_type']
                }
                try:
                    pdf_buffer = make_pdf(selected_template_path, data)
                    filename = f"근저당권설정_{data['debtor_name']}.pdf"
                    st.download_button(
                        label="⬇️ 다운로드",
                        data=pdf_buffer,
                        file_name=filename,
                        mime="application/pdf",
                        use_container_width=True
                    )
                except Exception as e: st.error(f"오류: {e}")
    
# =============================================================================
# Tab 4: 자필서명정보 작성
# =============================================================================
with tab4:
    # [추가됨] 주민/법인번호 자동 포맷팅 함수 (13자리 숫자 입력 시 '-' 자동 삽입)
    def auto_format_rrn(key):
        if key in st.session_state:
            val = st.session_state[key]
            # 숫자만 추출
            clean_val = re.sub(r'[^0-9]', '', str(val))
            # 13자리(법인/주민)인 경우 6-7 포맷 적용
            if len(clean_val) == 13:
                st.session_state[key] = f"{clean_val[:6]}-{clean_val[6:]}"

    # 헤더와 초기화 버튼
    col_header = st.columns([6, 1])
    col_header[0].markdown("### ✍️ 자필서명정보 작성")
    with col_header[1]:
        if st.button("🔄 초기화", type="secondary", use_container_width=True, key="reset_tab2", help="모든 입력 초기화"):
            # 등기의무자 1
            st.session_state['tab2_owner1_name'] = ''
            st.session_state['tab2_owner1_name_input'] = ''
            st.session_state['tab2_owner1_rrn'] = ''
            st.session_state['tab2_owner1_rrn_input'] = ''
            
            # 등기의무자 2
            st.session_state['tab2_owner2_name'] = ''
            st.session_state['tab2_owner2_name_input'] = ''
            st.session_state['tab2_owner2_rrn'] = ''
            st.session_state['tab2_owner2_rrn_input'] = ''
            
            # 부동산 표시
            st.session_state['tab2_estate'] = ''
            st.session_state['tab2_estate_input'] = ''
            
            # 날짜
            st.session_state['tab2_date'] = datetime.now().date()
            st.session_state['tab2_date_input'] = datetime.now().date()
            
            # 신청서 구분
            st.session_state['tab2_receipt_type'] = '전자신청'
            
            st.success("✅ 초기화되었습니다!")
            st.rerun()
    
    # 대부업(전자설정)내용 가져오기 버튼
    col_btn1, col_spacer = st.columns([1, 5])
    with col_btn1:
        if st.button("📥 대부업(전자설정)내용 가져오기", type="primary", use_container_width=True, key="sync_tab2", help="2탭(대부업) 정보 불러오기"):
            # 1. 1탭 데이터 확보 (위젯 Key 기준)
            contract_type = st.session_state.get('contract_type', '개인')
            debtor_name = st.session_state.get('t1_debtor_name', '')
            owner_name = st.session_state.get('t1_owner_name', '')
            # 1탭 부동산표시 위젯의 값을 가져옵니다
            estate_info = st.session_state.get('estate_text_area', '') 
            
            # 주민번호도 1탭에서 가져오기
            debtor_rrn = st.session_state.get('t1_debtor_rrn', '')
            owner_rrn = st.session_state.get('t1_owner_rrn', '')
            date_val = st.session_state.get('input_date', datetime.now().date())

            # 2. 계약 유형별 할당 데이터 준비
            o1_name, o1_rrn = "", ""
            o2_name, o2_rrn = "", ""

            if contract_type == "개인":
                # 단독: 채무자만
                o1_name, o1_rrn = debtor_name, debtor_rrn
            elif contract_type == "3자담보":
                # 3자: 소유자만
                o1_name, o1_rrn = owner_name, owner_rrn
            elif contract_type == "공동담보":
                # 공동: 채무자 + 소유자
                o1_name, o1_rrn = debtor_name, debtor_rrn
                o2_name, o2_rrn = owner_name, owner_rrn

            # 3. [핵심] Tab 2 위젯 Key(_input)와 Value 변수를 동시에 업데이트
            
            # 등기의무자 1
            st.session_state['tab2_owner1_name_input'] = o1_name
            st.session_state['tab2_owner1_name'] = o1_name
            
            st.session_state['tab2_owner1_rrn_input'] = o1_rrn
            st.session_state['tab2_owner1_rrn'] = o1_rrn

            # 등기의무자 2
            st.session_state['tab2_owner2_name_input'] = o2_name
            st.session_state['tab2_owner2_name'] = o2_name
            
            st.session_state['tab2_owner2_rrn_input'] = o2_rrn
            st.session_state['tab2_owner2_rrn'] = o2_rrn
            
            # 부동산 표시
            st.session_state['tab2_estate_input'] = estate_info
            st.session_state['tab2_estate'] = estate_info
            
            # 작성일자
            st.session_state['tab2_date_input'] = date_val
            st.session_state['tab2_date'] = date_val
            
            st.success("✅ 2탭(대부업) 정보를 불러왔습니다!")
            st.rerun()
    
    st.markdown("---")
    
    # 신청서 구분
    st.markdown("#### 📋 신청서 구분")
    if 'tab2_receipt_type' not in st.session_state:
        st.session_state['tab2_receipt_type'] = '전자신청'
    
    col_receipt1, col_receipt2 = st.columns(2)
    with col_receipt1:
        if st.button("전자신청", 
                     type="primary" if st.session_state['tab2_receipt_type']=="전자신청" else "secondary",
                     use_container_width=True,
                     key="btn_receipt_1"):
            st.session_state['tab2_receipt_type'] = "전자신청"
            st.rerun()
    with col_receipt2:
        if st.button("서면신청",
                     type="primary" if st.session_state['tab2_receipt_type']=="서면신청" else "secondary",
                     use_container_width=True,
                     key="btn_receipt_2"):
            st.session_state['tab2_receipt_type'] = "서면신청"
            st.rerun()
    
    # 템플릿 파일 선택
    if st.session_state['tab2_receipt_type'] == "전자신청":
        template_filename = "자필서명정보_템플릿.pdf"
    else:
        template_filename = "자필서명정보_서면_템플릿.pdf"
    
    template_path = resource_path(template_filename)
    
    if os.path.exists(template_path):
        st.success(f"✅ {st.session_state['tab2_receipt_type']} 템플릿 준비완료")
    else:
        st.error(f"⚠️ {template_filename} 파일이 없습니다.")
    
    st.markdown("---")
    
    # 입력 정보
    with st.expander("📝 자필서명정보 입력", expanded=True):
        # 작성일자
        st.markdown("#### 📅 작성일자")
        if 'tab2_date' not in st.session_state:
            st.session_state['tab2_date'] = datetime.now().date()
        tab2_date = st.date_input(
            "작성일자",
            value=st.session_state.get('tab2_date', datetime.now().date()),
            key='tab2_date_input'
        )
        
        st.markdown("---")
        
        # 등기의무자 정보
        st.markdown("#### 👤 등기의무자 정보")
        st.caption("※ 대부업 가져오기: 단독(채무자), 3자(소유자), 공동(채무자+소유자)")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**등기의무자 1**")
            tab2_owner1_name = st.text_input(
                "성명(법인명 직함 성명)",
                value=st.session_state.get('tab2_owner1_name', ''),
                key='tab2_owner1_name_input',
                placeholder="주식회사대한민국 대표이사 홍길동"
            )
            
            # [수정] 붉은색 강조 라벨 (마진 조정)
            st.markdown("""
                <div style='color: #FF4B4B; font-weight: 600; font-size: 0.9rem; margin-top: 10px; margin-bottom: 5px;'>
                    ⚠️ 주민(법인)등록번호 <span style='font-size: 0.8rem; opacity: 0.8;'>(수기입력 필수)</span>
                </div>
            """, unsafe_allow_html=True)
            
            # [수정] on_change 추가 (자동 하이픈)
            tab2_owner1_rrn = st.text_input(
                "주민(법인)등록번호_라벨숨김1",
                value=st.session_state.get('tab2_owner1_rrn', ''),
                key='tab2_owner1_rrn_input',
                placeholder="숫자 13자리 입력 후 엔터 (자동 '-' 입력)",
                label_visibility="collapsed",
                on_change=auto_format_rrn,
                args=('tab2_owner1_rrn_input',)
            )
        
        with col2:
            st.markdown("**등기의무자 2** (공동명의인 경우)")
            tab2_owner2_name = st.text_input(
                "성명(법인명 직함 성명)",
                value=st.session_state.get('tab2_owner2_name', ''),
                key='tab2_owner2_name_input',
                placeholder="(선택사항)"
            )
            
            # [수정] 붉은색 강조 라벨 (마진 조정)
            st.markdown("""
                <div style='color: #FF4B4B; font-weight: 600; font-size: 0.9rem; margin-top: 10px; margin-bottom: 5px;'>
                    ⚠️ 주민(법인)등록번호 <span style='font-size: 0.8rem; opacity: 0.8;'>(수기입력 필수)</span>
                </div>
            """, unsafe_allow_html=True)
            
            # [수정] on_change 추가 (자동 하이픈)
            tab2_owner2_rrn = st.text_input(
                "주민(법인)등록번호_라벨숨김2",
                value=st.session_state.get('tab2_owner2_rrn', ''),
                key='tab2_owner2_rrn_input',
                placeholder="(선택사항) 숫자 13자리 입력 후 엔터",
                label_visibility="collapsed",
                on_change=auto_format_rrn,
                args=('tab2_owner2_rrn_input',)
            )
        
        st.markdown("---")
        
        # 부동산 표시
        st.markdown("#### 🏠 부동산의 표시")
        tab2_estate = st.text_area(
            "부동산 표시",
            value=st.session_state.get('tab2_estate', ''),
            height=170,
            key='tab2_estate_input',
            placeholder="[토지]\n서울특별시 강남구 대치동 123번지\n대 300㎡"
        )
    
    st.markdown("---")
    
    # PDF 생성
    st.markdown("#### 📄 PDF 생성")
    
    if st.button("🚀 자필서명정보 PDF 생성", type="primary", use_container_width=True, key="generate_signature_pdf"):
        if not LIBS_OK:
            st.error("PDF 라이브러리가 설치되지 않았습니다.")
        elif not os.path.exists(template_path):
            st.error(f"{template_filename} 파일이 없습니다.")
        else:
            try:
                # 자필서명정보 PDF 데이터 준비
                signature_data = {
                    "date": format_date_korean(tab2_date),
                    "debtor_name": tab2_owner1_name or "[등기의무자1]",
                    "debtor_rrn": tab2_owner1_rrn or "[주민번호1]",
                    "owner_name": tab2_owner2_name or "",
                    "owner_rrn": tab2_owner2_rrn or "",
                    "estate_text": tab2_estate or "[부동산 표시]",
                    "purpose": "근저당권설정"
                }
                
                # make_signature_pdf 함수 사용
                pdf_buffer = make_signature_pdf(template_path, signature_data)
                
                filename = f"자필서명정보_{tab2_owner1_name or '고객'}_{st.session_state['tab2_receipt_type']}.pdf"
                st.download_button(
                    label="⬇️ 다운로드",
                    data=pdf_buffer,
                    file_name=filename,
                    mime="application/pdf",
                    use_container_width=True
                )
                
            except Exception as e:
                st.error(f"PDF 생성 오류: {e}")
    
    st.info("💡 **사용 방법**: '📥 대부업(전자설정)내용 가져오기' 버튼을 눌러 계약 유형에 따라 정보를 자동으로 불러올 수 있습니다.")


# Tab 5: 비용계산(대부업)
with tab5:
    # 헤더와 초기화 버튼
    col_header = st.columns([6, 1])
    col_header[0].markdown("### 🧾 등기비용 계산기(대부업)")
    with col_header[1]:
        if st.button("🔄 초기화", type="secondary", use_container_width=True, key="reset_tab3", help="모든 입력 초기화"):
            # 동기화 방지 플래그 설정
            st.session_state['tab3_reset_flag'] = True
            
            # 기본 정보
            st.session_state['calc_data'] = {}
            st.session_state['show_fee'] = True
            st.session_state['show_fee_checkbox'] = True
            st.session_state['input_parcels'] = 1
            st.session_state['calc_parcels_input'] = 1
            st.session_state['input_rate'] = f"{get_rate()*100:.5f}"
            st.session_state['calc_rate_input'] = st.session_state['input_rate']
            
            # 주소변경
            st.session_state['use_address_change'] = False
            st.session_state['address_change_count'] = 0
            st.session_state['addr_count_input'] = 1
            
            # 금액
            st.session_state['calc_amount_input'] = ''
            st.session_state['input_amount'] = ''
            
            # 채무자, 채권자, 물건지 (위젯 키 포함)
            st.session_state['tab3_creditor_select'] = list(CREDITORS.keys())[0]
            st.session_state['tab3_direct_name'] = ''  # 직접입력 채권자명
            st.session_state['tab3_debtor_input'] = ''
            st.session_state['tab3_estate_input'] = ''
            st.session_state['input_debtor'] = ''
            st.session_state['calc_debtor_view'] = ''
            st.session_state['input_collateral_addr'] = ''
            st.session_state['calc_estate_view'] = ''
            st.session_state['calc_creditor_view'] = ''
            
            # 공과금 자동계산 항목
            st.session_state['tax_등록면허세'] = '0'
            st.session_state['tax_지방교육세'] = '0'
            st.session_state['tax_증지대'] = '0'
            st.session_state['tax_채권할인'] = '0'
            
            # base 값 초기화 (주소변경 계산용)
            st.session_state['tax_등록면허세_base'] = 0
            st.session_state['tax_지방교육세_base'] = 0
            st.session_state['tax_증지대_base'] = 0
            st.session_state['tax_채권할인_base'] = 0
            
            # 수기입력 항목
            st.session_state['cost_manual_제증명'] = '0'
            st.session_state['cost_manual_교통비'] = '0'
            st.session_state['cost_manual_원인증서'] = '0'
            st.session_state['cost_manual_주소변경'] = '0'
            st.session_state['cost_manual_확인서면'] = '0'
            st.session_state['cost_manual_선순위 말소'] = '0'
            
            # 보수 항목
            st.session_state['base_fee_val'] = '0'
            st.session_state['add_fee_val'] = '0'
            st.session_state['etc_fee_val'] = '0'
            st.session_state['disc_fee_val'] = '0'
            
            st.success("✅ 초기화되었습니다!")
            st.rerun()
    
    # 대부업(전자설정)내용 가져오기 버튼
    col_btn1, col_spacer = st.columns([2, 4])
    with col_btn1:
        if st.button("📥 대부업(전자설정)내용 가져오기", type="primary", use_container_width=True, key="sync_tab3", help="2탭(대부업) 정보 불러오기"):
            # 1탭 값 직접 가져와서 Tab3 위젯에 설정
            creditor_val = st.session_state.get('input_creditor', '')
            debtor_val = st.session_state.get('t1_debtor_name', '')
            amount_val = st.session_state.get('input_amount', '')
            estate_val = st.session_state.get('input_collateral_addr', '')
            if not estate_val:
                estate_val = extract_address_from_estate(st.session_state.get('estate_text') or "")
            
            # Tab3 위젯에 직접 설정 (selectbox는 이 값을 사용함)
            st.session_state['tab3_creditor_select'] = creditor_val
            st.session_state['tab3_debtor_input'] = debtor_val
            st.session_state['calc_amount_input'] = amount_val
            st.session_state['tab3_estate_input'] = estate_val
            
            # 수기입력 기본값 설정 (금융사에 따라)
            if "(주)유노스프레스티지대부" in creditor_val:
                st.session_state['cost_manual_제증명'] = "20,000"
                st.session_state['cost_manual_교통비'] = "0"
                st.session_state['cost_manual_원인증서'] = "0"
            else:
                st.session_state['cost_manual_제증명'] = "50,000"
                st.session_state['cost_manual_교통비'] = "100,000"
                st.session_state['cost_manual_원인증서'] = "50,000"
            
            st.success("✅ 2탭(대부업) 정보가 동기화되었습니다!")
            st.rerun()
    st.markdown("---")

    # =========================================================
    # 3탭 자체 데이터 사용 (초기화 플래그 확인)
    # =========================================================
    
    # 초기화 직후인지 확인
    if st.session_state.get('tab3_reset_flag', False):
        st.session_state['tab3_reset_flag'] = False
        # 초기화 값 유지 (덮어쓰지 않음)
        debtor_from_tab1 = ''
        creditor_from_tab1 = list(CREDITORS.keys())[0]
        amount_from_tab1 = ''
        estate_from_tab1 = ''
    else:
        # 3탭 자체 값 사용 (없으면 빈 값)
        debtor_from_tab1 = st.session_state.get('tab3_debtor_input', '')
        creditor_from_tab1 = st.session_state.get('tab3_creditor_select', list(CREDITORS.keys())[0])
        amount_from_tab1 = st.session_state.get('calc_amount_input', '')
        estate_from_tab1 = st.session_state.get('tab3_estate_input', '')
    
    # =========================================================
    # 1. 통합 기본 정보 섹션
    # =========================================================
    creditor_display = creditor_from_tab1
    if creditor_display == "🖊️ 직접입력": 
        creditor_display = st.session_state.get('input_creditor_name', '직접입력')
    
    estate_display = estate_from_tab1

    row1_c1, row1_c2, row1_c3, row1_c4 = st.columns([2, 0.5, 1, 1.2]) 
    
    with row1_c1:
        def on_tab3_amount_change():
            val = st.session_state.get('calc_amount_input', '')
            formatted = format_number_with_comma(val)
            st.session_state['calc_amount_input'] = formatted
        st.text_input("채권최고액", key='calc_amount_input', on_change=on_tab3_amount_change, placeholder="금액 입력")

    with row1_c3:
        parcels_val = st.session_state.get('input_parcels', 1)
        new_parcels = st.number_input("필지수", min_value=1, value=int(parcels_val), key='calc_parcels_input')
        st.session_state['input_parcels'] = new_parcels

    with row1_c4:
        col_rate, col_btn = st.columns([2, 0.5])
        
        # 갱신 버튼 처리 (먼저 처리)
        if col_btn.button("🔄", help="오늘 할인율 가져오기"):
            new_rate_val = f"{get_rate()*100:.5f}"
            st.session_state['input_rate'] = new_rate_val
            st.session_state['calc_rate_input'] = new_rate_val
            st.rerun()
        
        # 할인율 입력 (key만 사용)
        if 'calc_rate_input' not in st.session_state:
            st.session_state['calc_rate_input'] = st.session_state.get('input_rate', '12.00000')
        new_rate = col_rate.text_input("할인율(%)", key='calc_rate_input')
        st.session_state['input_rate'] = new_rate

    row2_c1, row2_c2 = st.columns([1, 1])
    
    # 금융사 선택 (1탭 값 기준)
    with row2_c1:
        # 기본 목록
        creditor_list = list(CREDITORS.keys())
        
        # 직접입력한 채권자 목록 관리
        if 'custom_creditors' not in st.session_state:
            st.session_state['custom_creditors'] = []
        
        # 기존 input_creditor_name이 있으면 custom_creditors로 이전
        direct_creditor_name = st.session_state.get('input_creditor_name', '')
        if direct_creditor_name:
            is_duplicate = any(direct_creditor_name in key for key in CREDITORS.keys())
            if not is_duplicate and direct_creditor_name not in st.session_state['custom_creditors']:
                st.session_state['custom_creditors'].append(direct_creditor_name)
        
        # 직접입력한 채권자들 목록에 추가
        for custom_name in st.session_state['custom_creditors']:
            creditor_list.append(f"📝 {custom_name}")
        
        creditor_list.append("🖊️ 직접입력")
        
        # 1탭 값을 우선 사용
        current_creditor = creditor_from_tab1 if creditor_from_tab1 else creditor_list[0]
        
        # "🖊️ 직접입력"이면서 이름이 있으면 → 기존 목록에서 찾거나 "📝 이름"으로 변환
        if current_creditor == "🖊️ 직접입력" and direct_creditor_name:
            # 기존 CREDITORS에서 매칭되는 키 찾기
            matched_key = next((key for key in CREDITORS.keys() if direct_creditor_name in key), None)
            if matched_key:
                current_creditor = matched_key
            elif f"📝 {direct_creditor_name}" in creditor_list:
                current_creditor = f"📝 {direct_creditor_name}"
        
        if current_creditor not in creditor_list:
            current_creditor = creditor_list[0]
        default_index = creditor_list.index(current_creditor)
        
        def on_tab3_creditor_change():
            selected = st.session_state.get('tab3_creditor_select')
            # 3탭 내부에서만 사용 (1탭 연동 제거)
            st.session_state['calc_creditor_view'] = selected
            # 수기입력 비용 설정
            if selected and "(주)유노스프레스티지대부" in selected:
                st.session_state['cost_manual_제증명'] = "20,000"
            elif selected and selected != "🖊️ 직접입력":
                # 기타 금융사는 초기화된 값이 아니면 유지
                pass
        
        # 금융사 선택과 삭제 버튼을 나란히 배치
        col_select, col_del = st.columns([5, 1])
        with col_select:
            selected_creditor_tab3 = st.selectbox("금융사", options=creditor_list, index=default_index, key='tab3_creditor_select', on_change=on_tab3_creditor_change, label_visibility="collapsed")
        with col_del:
            # 📝로 시작하는 직접입력 항목일 때만 삭제 버튼 표시
            if selected_creditor_tab3.startswith("📝 "):
                if st.button("🗑️", key="del_custom_creditor", help="직접입력 채권자 삭제"):
                    custom_name = selected_creditor_tab3[2:].strip()
                    if custom_name in st.session_state['custom_creditors']:
                        st.session_state['custom_creditors'].remove(custom_name)
                    # input_creditor_name도 초기화
                    if st.session_state.get('input_creditor_name') == custom_name:
                        st.session_state['input_creditor_name'] = ''
                    st.session_state['tab3_creditor_select'] = list(CREDITORS.keys())[0]
                    st.rerun()
        
        # 직접입력 선택 시 입력 필드 표시
        if selected_creditor_tab3 == "🖊️ 직접입력":
            st.text_input(
                "채권자 성명/상호", 
                key='tab3_direct_name',
                placeholder="채권자명 입력"
            )
        
        # 유노스프레스티지 선택 시 제증명 20,000원 자동 설정
        if "(주)유노스프레스티지대부" in selected_creditor_tab3:
            current_cert_fee = st.session_state.get('cost_manual_제증명', '0')
            try:
                cert_fee_val = int(str(current_cert_fee).replace(',', '').replace('원', '').strip() or '0')
            except:
                cert_fee_val = 0
            # 0이거나 기본값(50,000)이면 20,000으로 변경
            if cert_fee_val == 0 or cert_fee_val == 50000:
                st.session_state['cost_manual_제증명'] = "20,000"
    
    # 채무자 입력
    with row2_c2:
        def on_tab3_debtor_change():
            st.session_state['calc_debtor_view'] = st.session_state.get('tab3_debtor_input', '')
        
        st.text_input("채무자", key='tab3_debtor_input', on_change=on_tab3_debtor_change, placeholder="채무자명")
    
    # 물건지 입력
    def on_tab3_estate_change():
        st.session_state['calc_estate_view'] = st.session_state.get('tab3_estate_input', '')
    
    st.text_area("물건지", key='tab3_estate_input', on_change=on_tab3_estate_change, height=80, placeholder="물건지 주소")
    st.markdown("---")

    # =========================================================
    # 2. 계산 로직 수행
    # =========================================================
    # 주소변경 인원수 처리 (체크되어 있으면 인원수 반영, 아니면 0명)
    current_addr_use = st.session_state.get('use_address_change', False)
    current_addr_count = int(st.session_state.get('addr_count_input', 1)) if current_addr_use else 0
    st.session_state['address_change_count'] = current_addr_count
    
    # 3탭 위젯 값 사용 (사용자가 수정한 경우 그 값 반영)
    creditor_for_calc = st.session_state.get('tab3_creditor_select', creditor_from_tab1)
    if creditor_for_calc == "🖊️ 직접입력":
        creditor_for_calc = st.session_state.get('input_creditor_name', '직접입력')
    elif creditor_for_calc.startswith("📝 "):
        creditor_for_calc = creditor_for_calc[2:].strip()  # "📝 " 제거
    
    calc_input_data = {
        '채권최고액': st.session_state.get('calc_amount_input', amount_from_tab1), 
        '필지수': st.session_state['input_parcels'],
        '채권할인율': st.session_state['input_rate'],
        '금융사': creditor_for_calc,
        '채무자': st.session_state.get('tab3_debtor_input', debtor_from_tab1),
        '물건지': st.session_state.get('tab3_estate_input', estate_from_tab1),
        '기본료_val': st.session_state.get('base_fee_val', "0"),
        '추가보수_val': st.session_state.get('add_fee_val', "0"),
        '기타보수_val': st.session_state.get('etc_fee_val', "0"),
        '할인금액': st.session_state.get('disc_fee_val', "0"),
    }
    
    final_data = calculate_all(calc_input_data)
    st.session_state['calc_data'] = final_data 

    # =========================================================
    # 3. 3단 레이아웃 (보수액 / 공과금 / 결제)
    # =========================================================
    
    # [수정] 결과 표시용 함수: disabled=True인 경우 state를 강제 갱신하여 0원 표시 방지
    def make_row(label, value, key, on_change=None, disabled=False):
        c1, c2 = st.columns([1, 1.8])
        with c1: st.markdown(f"<div class='row-label'>{label}</div>", unsafe_allow_html=True)
        with c2:
            formatted_val = str(value)
            
            # disabled(수정불가) 항목은 텍스트로 표시
            if disabled:
                st.session_state[key] = formatted_val
                st.markdown(f"<div style='text-align:right; padding:8px 12px; background:#2a2a2a; border-radius:4px; color:#00d26a; font-weight:bold;'>{formatted_val} 원</div>", unsafe_allow_html=True)
            elif on_change:
                # key만 사용 (value와 key 동시 사용 시 충돌 방지)
                if key not in st.session_state or st.session_state[key] == "0":
                    st.session_state[key] = formatted_val
                st.text_input(label, key=key, on_change=on_change, args=(key,), label_visibility="collapsed")
            else:
                if key not in st.session_state or st.session_state[key] == "0":
                    st.session_state[key] = formatted_val
                st.text_input(label, key=key, label_visibility="collapsed")

    def format_cost_input(key):
        val = st.session_state[key]
        st.session_state[key] = format_number_with_comma(val)

    col_income, col_tax, col_payment = st.columns([1, 1, 1])

    # [1] 보수액 (Income)
    with col_income:
        st.markdown("<div class='section-header income-header'>💰 보수액 (Income)</div>", unsafe_allow_html=True)
        with st.container(border=True, height=650):
            # 기본료: 수기 입력 가능 (자동계산 값 표시)
            auto_base_fee = final_data.get('기본료_자동', 0)
            
            # 엘하비스트대부 체크
            current_creditor = st.session_state.get('tab3_creditor_select', '')
            is_elharvest = (current_creditor == "㈜엘하비스트대부 대표이사 김상수")
            
            # 이전 금융사 상태 추적 (금융사 변경 감지용)
            prev_creditor = st.session_state.get('_prev_creditor_for_fee', '')
            if prev_creditor != current_creditor:
                # 금융사가 변경되면 기본료를 자동계산 값으로 리셋
                st.session_state['base_fee_val'] = format_number_with_comma(auto_base_fee)
                st.session_state['_prev_creditor_for_fee'] = current_creditor
            
            c1, c2 = st.columns([1, 1.8])
            with c1: 
                st.markdown("<div class='row-label'>기본료</div>", unsafe_allow_html=True)
                st.caption(f"(자동: {format_number_with_comma(auto_base_fee)})")
            with c2:
                if is_elharvest:
                    # 엘하비스트대부는 항상 70,000원 고정 (별도 표시)
                    st.text_input("기본료", value="70,000", disabled=True, label_visibility="collapsed", key="base_fee_display_elharvest")
                    st.session_state['base_fee_val'] = "70,000"
                else:
                    # 수기입력 값이 0이면 자동계산 값으로 초기화
                    if st.session_state.get('base_fee_val', "0") == "0" and auto_base_fee > 0:
                        st.session_state['base_fee_val'] = format_number_with_comma(auto_base_fee)
                    st.text_input("기본료", key="base_fee_val", on_change=format_cost_input, args=("base_fee_val",), label_visibility="collapsed")
            
            make_row("추가보수", st.session_state['add_fee_val'], "add_fee_val", format_cost_input)
            make_row("기타보수", st.session_state['etc_fee_val'], "etc_fee_val", format_cost_input)
            make_row("할인금액", st.session_state['disc_fee_val'], "disc_fee_val", format_cost_input)
            st.markdown("---")
            
            # 공급가액, 부가세, 보수총액 - 통일된 스타일
            st.markdown(f"""
            <div style='display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;'>
                <span style='font-weight:bold;'>공급가액</span>
                <span style='color:#28a745; font-weight:bold;'>{format_number_with_comma(final_data.get('공급가액', 0))} 원</span>
            </div>
            <div style='display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;'>
                <span style='font-weight:bold;'>부가세</span>
                <span style='color:#28a745; font-weight:bold;'>{format_number_with_comma(final_data.get('부가세', 0))} 원</span>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("---")
            
            # 보수 총액
            st.markdown(f"""
            <div style='display:flex; justify-content:space-between; align-items:center;'>
                <span style='font-size:1.2rem; font-weight:bold;'>보수 총액</span>
                <span style='color:#28a745; font-size:1.3rem; font-weight:bold;'>{format_number_with_comma(final_data.get('보수총액', 0))} 원</span>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("---")
            # 참고 기준 (보수액 섹션 하단으로 이동)
            st.info("**ℹ️ 참고 기준 (주소변경)**\n* **주소변경비용**: 유노스/드림 20,000원/인, 기타 50,000원/인\n* **공과금 추가** (인원별): 등록면허세 +6,000원, 지방교육세 +1,200원, 증지대 +3,000원")

    # [2] 공과금 (Tax)
    with col_tax:
        st.markdown("<div class='section-header tax-header'>🏛️ 공과금 (Tax)</div>", unsafe_allow_html=True)
        with st.container(border=True, height=650):
            st.caption("[자동 계산] (calculate_all 결과 반영)")
            
            # calculate_all이 이미 계산해준 값(주소변경 포함)을 그대로 표시
            st.session_state['tax_등록면허세'] = format_number_with_comma(final_data.get("등록면허세", 0))
            st.session_state['tax_지방교육세'] = format_number_with_comma(final_data.get("지방교육세", 0))
            st.session_state['tax_증지대'] = format_number_with_comma(final_data.get("증지대", 0))
            st.session_state['tax_채권할인'] = format_number_with_comma(final_data.get("채권할인금액", 0))
            
            def display_tax_row(label, value):
                c1, c2 = st.columns([1, 1.5])
                c1.markdown(f"**{label}**")
                c2.markdown(f"<div style='text-align:right; padding:8px; border:1px solid #444; border-radius:4px; background:rgba(255,255,255,0.05);'>{value} 원</div>", unsafe_allow_html=True)
            
            display_tax_row("등록면허세", st.session_state['tax_등록면허세'])
            display_tax_row("지방교육세", st.session_state['tax_지방교육세'])
            display_tax_row("증지대", st.session_state['tax_증지대'])
            display_tax_row("채권할인", st.session_state['tax_채권할인'])
            
            st.markdown("---")
            st.caption("[수기 입력]")
            
            # 수기 입력 항목들
            make_row("제증명", st.session_state['cost_manual_제증명'], "cost_manual_제증명", format_cost_input)
            make_row("교통비", st.session_state['cost_manual_교통비'], "cost_manual_교통비", format_cost_input)
            make_row("원인증서", st.session_state['cost_manual_원인증서'], "cost_manual_원인증서", format_cost_input)
            make_row("주소변경", st.session_state['cost_manual_주소변경'], "cost_manual_주소변경", format_cost_input)
            make_row("확인서면", st.session_state['cost_manual_확인서면'], "cost_manual_확인서면", format_cost_input)
            make_row("선순위말소", st.session_state['cost_manual_선순위 말소'], "cost_manual_선순위 말소", format_cost_input)
            
            st.markdown("---")
            # 공과금 소계 (calculate_all에서 계산된 값 사용)
            tax_subtotal = final_data.get('공과금 총액', 0)
            st.markdown(f"""
            <div style='display:flex; justify-content:space-between; align-items:center;'>
                <span style='font-size:1.2rem; font-weight:bold;'>공과금 소계</span>
                <span style='color:#fd7e14; font-size:1.3rem; font-weight:bold;'>{format_number_with_comma(tax_subtotal)} 원</span>
            </div>
            """, unsafe_allow_html=True)

    # [3] 결제 및 청구
    with col_payment:
        st.markdown("<div class='section-header total-header'>🧾 결제 및 청구</div>", unsafe_allow_html=True)
        with st.container(border=True, height=650):
            # 총 청구금액 (calculate_all에서 계산된 값)
            grand_total = final_data.get('총 합계', 0)
            st.markdown("#### 총 청구금액")
            st.markdown(f"<div class='total-box'><div class='total-amount'>{format_number_with_comma(grand_total)} 원</div></div>", unsafe_allow_html=True)
            st.markdown("---")
            
            def toggle_show_fee(): st.session_state['show_fee'] = st.session_state['show_fee_checkbox']
            st.checkbox("보수액 포함 표시", value=st.session_state['show_fee'], key='show_fee_checkbox', on_change=toggle_show_fee)
            
            def update_manual_addr_cost():
                """체크박스/인원수 변경 시 -> 주소변경(수기입력) 비용만 업데이트"""
                if st.session_state.get('use_address_change', False):
                    count = int(st.session_state.get('addr_count_input', 1))
                    
                    # 금융사별 단가 (유노스/드림: 2만, 나머지: 5만)
                    cur_creditor = st.session_state.get('tab3_creditor_select', '')
                    unit_price = 20000 if ("유노스" in cur_creditor or "드림" in cur_creditor) else 50000
                    
                    # 수기 입력칸 값 업데이트
                    st.session_state['cost_manual_주소변경'] = format_number_with_comma(unit_price * count)
                else:
                    # 체크 해제 시 0원
                    st.session_state['cost_manual_주소변경'] = "0"

            cp1, cp2 = st.columns([1.5, 1])
            with cp1: 
                st.checkbox("주소변경 포함", key='use_address_change', on_change=update_manual_addr_cost)
            with cp2:
                st.number_input("인원", min_value=1, key='addr_count_input', label_visibility="collapsed", on_change=update_manual_addr_cost)
            st.caption("체크 시 인원별 공과금 추가 (등록면허세 +6,000 / 지방교육세 +1,200 / 증지대 +3,000)")
            
            st.markdown("---")
            
            # PDF 데이터 준비
            pdf_creditor = st.session_state.get('tab3_creditor_select', creditor_from_tab1)
            if pdf_creditor == "🖊️ 직접입력":
                pdf_creditor = st.session_state.get('input_creditor_name', '직접입력')
            elif pdf_creditor and pdf_creditor.startswith("📝 "):
                pdf_creditor = pdf_creditor[2:].strip()
            
            debtor_name = st.session_state.get('tab3_debtor_input', debtor_from_tab1)
            if not debtor_name or debtor_name.strip() == '':
                debtor_name = '고객'
            
            # PDF 다운로드 버튼 (클릭 시 바로 다운로드)
            if FPDF_OK:
                try:
                    pdf_data = {
                        'date_input': format_date_korean(st.session_state.get('input_date', datetime.now().date())),
                        'client': {
                            '채권최고액': format_number_with_comma(st.session_state.get('calc_amount_input', amount_from_tab1)),
                            '필지수': str(st.session_state.get('input_parcels', 1)),
                            '금융사': pdf_creditor,
                            '채무자': debtor_name,
                            '물건지': st.session_state.get('tab3_estate_input', estate_from_tab1)
                        },
                        'fee_items': {
                            '기본료': parse_int_input(final_data.get('기본료', 0)),
                            '추가보수': parse_int_input(st.session_state.get('add_fee_val', 0)),
                            '기타보수': parse_int_input(st.session_state.get('etc_fee_val', 0)),
                            '할인금액': parse_int_input(st.session_state.get('disc_fee_val', 0))
                        },
                        'fee_totals': {'보수총액': final_data.get('보수총액', 0)},
                        'cost_items': {
                            '등록면허세': final_data.get('등록면허세', 0),
                            '지방교육세': final_data.get('지방교육세', 0),
                            '증지대': final_data.get('증지대', 0),
                            '채권할인': final_data.get('채권할인금액', 0),
                            '제증명': parse_int_input(st.session_state.get('cost_manual_제증명', 0)),
                            '교통비': parse_int_input(st.session_state.get('cost_manual_교통비', 0)),
                            '원인증서': parse_int_input(st.session_state.get('cost_manual_원인증서', 0)),
                            '주소변경': parse_int_input(st.session_state.get('cost_manual_주소변경', 0)),
                            '확인서면': parse_int_input(st.session_state.get('cost_manual_확인서면', 0)),
                            '선순위 말소': parse_int_input(st.session_state.get('cost_manual_선순위 말소', 0))
                        },
                        'cost_totals': {'공과금 총액': final_data.get('공과금 총액', 0)},
                        'cost_section_title': '2. 공과금' if st.session_state.get('show_fee', True) else '1. 공과금',
                        'grand_total': final_data.get('총 합계', 0)
                    }
                    pdf_converter = PDFConverter(show_fee=st.session_state.get('show_fee', True))
                    pdf_buffer = pdf_converter.output_pdf(pdf_data)
                    
                    st.download_button(
                        label="📄 비용내역 PDF 다운로드",
                        data=pdf_buffer,
                        file_name=f"근저당설정_비용내역_{debtor_name}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="btn_pdf_download"
                    )
                except Exception as e:
                    st.button("📄 비용내역 PDF 다운로드", disabled=True, use_container_width=True)
                    st.error(f"PDF 준비 오류: {e}")
            else:
                st.button("📄 비용내역 PDF 다운로드", disabled=True, use_container_width=True)
            
            # Excel 다운로드 버튼 (클릭 시 바로 다운로드)
            if EXCEL_OK:
                try:
                    receipt_template = st.session_state['template_status'].get('영수증')
                    excel_data = {
                        'date_input': format_date_korean(st.session_state.get('input_date', datetime.now().date())),
                        'client': {
                            '채권최고액': format_number_with_comma(st.session_state.get('calc_amount_input', amount_from_tab1)),
                            '필지수': str(st.session_state.get('input_parcels', 1)),
                            '금융사': pdf_creditor,
                            '채무자': debtor_name,
                            '물건지': st.session_state.get('tab3_estate_input', estate_from_tab1)
                        },
                        'fee_items': {
                            '기본료': parse_int_input(final_data.get('기본료', 0)),
                            '추가보수': parse_int_input(st.session_state.get('add_fee_val', 0)),
                            '기타보수': parse_int_input(st.session_state.get('etc_fee_val', 0)),
                            '할인금액': parse_int_input(st.session_state.get('disc_fee_val', 0))
                        },
                        'fee_totals': {'보수총액': final_data.get('보수총액', 0)},
                        'cost_items': {
                            '등록면허세': final_data.get('등록면허세', 0),
                            '지방교육세': final_data.get('지방교육세', 0),
                            '증지대': final_data.get('증지대', 0),
                            '채권할인': final_data.get('채권할인금액', 0),
                            '제증명': parse_int_input(st.session_state.get('cost_manual_제증명', 0)),
                            '교통비': parse_int_input(st.session_state.get('cost_manual_교통비', 0)),
                            '원인증서': parse_int_input(st.session_state.get('cost_manual_원인증서', 0)),
                            '주소변경': parse_int_input(st.session_state.get('cost_manual_주소변경', 0)),
                            '확인서면': parse_int_input(st.session_state.get('cost_manual_확인서면', 0)),
                            '선순위 말소': parse_int_input(st.session_state.get('cost_manual_선순위 말소', 0))
                        },
                        'cost_totals': {'공과금 총액': final_data.get('공과금 총액', 0)}
                    }
                    excel_buffer = create_receipt_excel(excel_data, receipt_template)
                    
                    if excel_buffer:
                        st.download_button(
                            label="🏦 영수증 Excel 다운로드",
                            data=excel_buffer,
                            file_name=f"영수증_{debtor_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="btn_excel_download"
                        )
                    else:
                        st.button("🏦 영수증 Excel 다운로드", disabled=True, use_container_width=True)
                except Exception as e:
                    st.button("🏦 영수증 Excel 다운로드", disabled=True, use_container_width=True)
            else:
                st.button("🏦 영수증 Excel 다운로드", disabled=True, use_container_width=True)


# Tab 3: 말소문서
# =============================================================================
with tab3:
    # 말소 임시저장 목록 초기화
    if 'malso_saved_list' not in st.session_state:
        st.session_state['malso_saved_list'] = []
    
    # 헤더와 초기화 버튼
    col_header = st.columns([6, 1])
    col_header[0].markdown("### 🗑️ 말소 문서 작성")
    with col_header[1]:
        if st.button("🔄 초기화", type="secondary", use_container_width=True, key="reset_tab4", help="모든 입력 초기화"):
            # 말소 유형
            st.session_state['malso_type'] = '근저당권'
            
            # 등기의무자 (채권자)
            st.session_state['malso_obligor_name'] = ''
            st.session_state['malso_obligor_id'] = ''
            st.session_state['malso_obligor_addr'] = ''
            st.session_state['malso_obligor_rep'] = ''
            st.session_state['malso_obligor_branch'] = ''
            st.session_state['malso_obligor_select'] = "직접입력"
            st.session_state['malso_obligor_name_input'] = ''
            st.session_state['malso_obligor_id_input'] = ''
            st.session_state['malso_obligor_addr_input'] = ''
            st.session_state['malso_obligor_rep_input'] = ''
            st.session_state['malso_obligor_branch_input'] = ''
            
            # 등기권리자 1
            st.session_state['malso_holder1_name'] = ''
            st.session_state['malso_holder1_rrn'] = ''
            st.session_state['malso_holder1_addr'] = ''
            
            # 등기권리자 2
            st.session_state['malso_holder2_name'] = ''
            st.session_state['malso_holder2_rrn'] = ''
            st.session_state['malso_holder2_addr'] = ''
            
            # 부동산 표시
            st.session_state['malso_estate_detail'] = ''
            
            # 물건지 주소
            st.session_state['malso_property_addr'] = ''
            
            # 말소 내역
            st.session_state['malso_cancel_text'] = ''
            
            # 이전등기소
            st.session_state['malso_from_branch'] = ''
            st.session_state['malso_to_branch'] = ''
            
            # 원인일자
            st.session_state['malso_cause_date'] = datetime.now().date()
            st.session_state['malso_cause_date_input'] = datetime.now().date()
            
            # 임시저장 목록도 초기화
            st.session_state['malso_saved_list'] = []
            
            st.success("✅ 초기화되었습니다!")
            st.rerun()
    
    # 대부업(전자설정)내용 가져오기 및 임시저장 버튼
    col_btn1, col_btn2, col_spacer = st.columns([1, 1, 4])
    with col_btn1:
        if st.button("📥 대부업(전자설정)내용 가져오기", type="primary", use_container_width=True, key="sync_tab4", help="2탭(대부업) 정보 불러오기"):
            # 2탭 데이터 동기화
            contract_type = st.session_state.get('contract_type', '개인')
            
            # 등기권리자 (소유자) - 3자담보면 소유자, 아니면 채무자
            if contract_type == "3자담보":
                st.session_state['malso_holder1_name'] = st.session_state.get('t1_owner_name', '')
                st.session_state['malso_holder1_rrn'] = st.session_state.get('t1_owner_rrn', '')
                st.session_state['malso_holder1_addr'] = st.session_state.get('t1_owner_addr', '')
            else:
                st.session_state['malso_holder1_name'] = st.session_state.get('t1_debtor_name', '')
                st.session_state['malso_holder1_rrn'] = st.session_state.get('t1_debtor_rrn', '')
                st.session_state['malso_holder1_addr'] = st.session_state.get('t1_debtor_addr', '')
            
            # 등기의무자 (채권자)
            creditor = st.session_state.get('input_creditor', '')
            if creditor and creditor != "🖊️ 직접입력":
                creditor_info = CREDITORS.get(creditor, {})
                # 금융사명에서 회사명과 대표자 분리
                obligor_name = ''
                obligor_rep = ''
                if '대표이사' in creditor:
                    parts = creditor.split('대표이사')
                    obligor_name = parts[0].strip()
                    obligor_rep = '대표이사 ' + parts[1].strip() if len(parts) > 1 else ''
                elif '사내이사' in creditor:
                    parts = creditor.split('사내이사')
                    obligor_name = parts[0].strip()
                    obligor_rep = '사내이사 ' + parts[1].strip() if len(parts) > 1 else ''
                else:
                    obligor_name = creditor
                    obligor_rep = ''
                
                obligor_id = creditor_info.get('corp_num', '')
                obligor_addr = creditor_info.get('addr', '')
                
                # session_state 업데이트
                st.session_state['malso_obligor_name'] = obligor_name
                st.session_state['malso_obligor_rep'] = obligor_rep
                st.session_state['malso_obligor_id'] = obligor_id
                st.session_state['malso_obligor_addr'] = obligor_addr
                st.session_state['malso_obligor_branch'] = ''
                
                # input 키들도 업데이트
                st.session_state['malso_obligor_name_input'] = obligor_name
                st.session_state['malso_obligor_rep_input'] = obligor_rep
                st.session_state['malso_obligor_id_input'] = obligor_id
                st.session_state['malso_obligor_addr_input'] = obligor_addr
                st.session_state['malso_obligor_branch_input'] = ''
                
                # 선택 드롭다운을 "직접입력"으로 (2탭에서 가져온 값이므로)
                st.session_state['malso_obligor_select'] = "직접입력"
            
            # 부동산 표시
            st.session_state['malso_estate_detail'] = st.session_state.get('estate_text', '')
            
            # 물건지 주소
            st.session_state['malso_property_addr'] = st.session_state.get('input_collateral_addr', '')
            
            st.success("✅ 2탭(대부업) 정보를 불러왔습니다!")
            st.rerun()
    with col_btn2:
        if st.button("💾 임시저장", type="secondary", use_container_width=True, key="save_malso_temp", help="현재 말소건 임시저장"):
            # 현재 입력된 말소 데이터 수집
            malso_type = st.session_state.get('malso_type', '근저당권')
            holder_name = st.session_state.get('malso_holder1_name', '')
            holder_rrn = st.session_state.get('malso_holder1_rrn', '')
            holder_addr = st.session_state.get('malso_holder1_addr', '')
            property_addr = st.session_state.get('malso_property_addr', '')
            
            if not holder_name:
                st.error("❌ 등기권리자(소유자) 성명을 입력하세요!")
            else:
                # 중복 체크 (같은 유형 + 같은 이름이면 업데이트)
                existing_idx = None
                for idx, item in enumerate(st.session_state['malso_saved_list']):
                    if item['type'] == malso_type and item['holder_name'] == holder_name:
                        existing_idx = idx
                        break
                
                malso_data = {
                    'type': malso_type,
                    'holder_name': holder_name,
                    'holder_rrn': holder_rrn,
                    'holder_addr': holder_addr,
                    'property_addr': property_addr
                }
                
                if existing_idx is not None:
                    st.session_state['malso_saved_list'][existing_idx] = malso_data
                    st.success(f"✅ {malso_type}말소 - {holder_name} 업데이트 완료!")
                else:
                    st.session_state['malso_saved_list'].append(malso_data)
                    st.success(f"✅ {malso_type}말소 - {holder_name} 저장 완료!")
                st.rerun()
    
    # 임시저장된 말소건 목록 표시
    if st.session_state.get('malso_saved_list'):
        st.markdown("---")
        st.markdown("#### 📦 저장된 말소 건")
        
        with st.container(border=True):
            for idx, item in enumerate(st.session_state['malso_saved_list']):
                col_item, col_del = st.columns([5, 1])
                with col_item:
                    st.markdown(f"✅ **{item['type']}말소** - {item['holder_name']}")
                with col_del:
                    if st.button("🗑️", key=f"del_malso_{idx}", help="삭제"):
                        st.session_state['malso_saved_list'].pop(idx)
                        st.rerun()
            
            st.caption(f"총 {len(st.session_state['malso_saved_list'])}건 저장됨 → 위택스신고 탭에서 불러오기 가능")
    
    st.markdown("---")
    
    # 1. 말소 유형 선택
    st.markdown("#### 📋 말소 유형 선택")
    
    if 'malso_type' not in st.session_state:
        st.session_state['malso_type'] = "근저당권"
    
    malso_type_cols = st.columns(3)
    with malso_type_cols[0]:
        if st.button("근저당권", 
                     type="primary" if st.session_state['malso_type']=="근저당권" else "secondary",
                     use_container_width=True,
                     key="btn_malso_type_1"):
            st.session_state['malso_type'] = "근저당권"
            st.rerun()
    with malso_type_cols[1]:
        if st.button("질권",
                     type="primary" if st.session_state['malso_type']=="질권" else "secondary",
                     use_container_width=True,
                     key="btn_malso_type_2"):
            st.session_state['malso_type'] = "질권"
            st.rerun()
    with malso_type_cols[2]:
        if st.button("전세권",
                     type="primary" if st.session_state['malso_type']=="전세권" else "secondary",
                     use_container_width=True,
                     key="btn_malso_type_3"):
            st.session_state['malso_type'] = "전세권"
            st.rerun()
    
    # 권리자 라벨 설정
    malso_type = st.session_state.get('malso_type', '근저당권')
    if malso_type == "근저당권":
        obligor_label = "근저당권자"
    elif malso_type == "질권":
        obligor_label = "질권자"
    else:
        obligor_label = "전세권자"
    
    st.info(f"✅ 선택된 유형: **{malso_type}말소** ({obligor_label})")
    st.markdown("---")
    
    # 2. 등기의무자 / 등기권리자 입력
    col_input1, col_input2 = st.columns(2)
    
    # 콜백 함수 정의
    def on_obligor_change():
        selected = st.session_state.get('malso_obligor_select', '직접입력')
        if selected and selected != "직접입력":
            obligor_info = OBLIGORS.get(selected, {})
            st.session_state['malso_obligor_name_input'] = selected
            st.session_state['malso_obligor_id_input'] = obligor_info.get("corp_num", "")
            st.session_state['malso_obligor_addr_input'] = obligor_info.get("addr", "")
            st.session_state['malso_obligor_rep_input'] = obligor_info.get("rep", "")
            st.session_state['malso_obligor_branch_input'] = obligor_info.get("branch", "")
        else:
            st.session_state['malso_obligor_name_input'] = ""
            st.session_state['malso_obligor_id_input'] = ""
            st.session_state['malso_obligor_addr_input'] = ""
            st.session_state['malso_obligor_rep_input'] = ""
            st.session_state['malso_obligor_branch_input'] = ""
    
    with col_input1:
        st.markdown(f"#### 1️⃣ 등기의무자 ({obligor_label})")
        with st.container(border=True):
            # 등기의무자 선택
            obligor_options = list(OBLIGORS.keys())
            
            selected_obligor = st.selectbox(
                "등기의무자 선택",
                options=obligor_options,
                key="malso_obligor_select",
                index=0,
                on_change=on_obligor_change
            )
            
            # 입력 필드 (수정 가능)
            obligor_name = st.text_input("성명(법인명)", key="malso_obligor_name_input", placeholder="주식회사티플레인대부")
            st.session_state['malso_obligor_name'] = obligor_name
            
            obligor_id = st.text_input("법인등록번호", key="malso_obligor_id_input", placeholder="110111-7350161")
            st.session_state['malso_obligor_id'] = obligor_id
            
            obligor_addr = st.text_area("주소", key="malso_obligor_addr_input", height=80,
                        placeholder="서울특별시 마포구 삼개로 16, 2신관 1층 103호")
            st.session_state['malso_obligor_addr'] = obligor_addr
            
            # 취급지점 & 대표자 (같은 줄)
            rep_cols = st.columns([1, 1])
            with rep_cols[0]:
                obligor_branch = st.text_input("취급지점", key="malso_obligor_branch_input", placeholder="")
                st.session_state['malso_obligor_branch'] = obligor_branch
            with rep_cols[1]:
                obligor_rep = st.text_input("대표자 직위/성명", key="malso_obligor_rep_input", placeholder="대표이사 윤웅원")
                st.session_state['malso_obligor_rep'] = obligor_rep
    
    with col_input2:
        st.markdown("#### 2️⃣ 등기권리자 (소유자)")
        with st.container(border=True):
            st.markdown("**권리자 1**")
            h1_col1, h1_col2 = st.columns([2, 1])
            with h1_col1:
                st.text_input("성명", key="malso_holder1_name", placeholder="홍길동")
            with h1_col2:
                st.text_input("주민등록번호", key="malso_holder1_rrn", placeholder="000000-0000000")
            st.text_area("주소", key="malso_holder1_addr", height=60, placeholder="서울특별시 송파구...")
            
            st.markdown("**권리자 2** (있는 경우)")
            h2_col1, h2_col2 = st.columns([2, 1])
            with h2_col1:
                st.text_input("성명", key="malso_holder2_name", placeholder="김철수", label_visibility="collapsed")
            with h2_col2:
                st.text_input("주민등록번호", key="malso_holder2_rrn", placeholder="000000-0000000", label_visibility="collapsed")
            st.text_area("주소", key="malso_holder2_addr", height=60, placeholder="", label_visibility="collapsed")
    
    st.markdown("---")
    
    # 3. 등기원인 및 부동산 정보
    col_info = st.columns(2)
    with col_info[0]:
        st.markdown("#### 3️⃣ 등기원인과 그 년월일")
        if 'malso_cause_date' not in st.session_state:
            st.session_state['malso_cause_date'] = datetime.now().date()
        st.date_input("등기원인일", value=st.session_state['malso_cause_date'], key="malso_cause_date_input")
        st.session_state['malso_cause_date'] = st.session_state.get('malso_cause_date_input', datetime.now().date())
    
    with col_info[1]:
        st.markdown("#### 4️⃣ 등기목적")
        malso_purpose = f"{malso_type}말소"
        st.text_input("등기목적", value=malso_purpose, disabled=True, key="malso_purpose_display")
    
    st.markdown("#### 5️⃣ 부동산의 표시")
    with st.container(border=True):
        st.text_area(
            "부동산 상세 (등기부등본에서 복사)",
            key="malso_estate_detail",
            height=200,
            placeholder="1동의 건물의 표시\n서울특별시 송파구 문정동 150\n..."
        )
    
    # 물건지 주소 (위택스용)
    st.markdown("#### 📍 물건지 주소 (위택스 신고용)")
    if 'malso_property_addr' not in st.session_state:
        st.session_state['malso_property_addr'] = ''
    st.text_input(
        "물건지 주소",
        key="malso_property_addr",
        placeholder="서울특별시 강남구 테헤란로 123"
    )
    
    st.markdown("#### 6️⃣ 말소할 등기")
    st.text_input(
        "말소할 등기 (접수번호 등)",
        key="malso_cancel_text",
        placeholder="2025년09월30일 접수 제5201489호(으)로 경료한 근저당권설정"
    )
    
    st.markdown("---")
    
    # 4. 이관 정보 (이관증명서용)
    st.markdown("#### 🏦 이관 정보 (이관증명서용)")
    col_transfer = st.columns(2)
    with col_transfer[0]:
        st.text_input("이관 전", key="malso_from_branch", placeholder="취급지점명")
    with col_transfer[1]:
        st.text_input("이관 후", key="malso_to_branch", placeholder="본점")
    
    st.markdown("---")
    
    # 5. PDF 생성 - 체크박스 선택 시 바로 다운로드 버튼 표시
    st.markdown("### 📥 문서 생성")
    
    # 체크박스로 문서 선택
    col_chk = st.columns(4)
    with col_chk[0]:
        chk_sig = st.checkbox("📄 자필서명정보", value=True, key="chk_malso_sig")
    with col_chk[1]:
        chk_power = st.checkbox("📄 위임장", value=True, key="chk_malso_power")
    with col_chk[2]:
        chk_term = st.checkbox("📄 해지증서", value=True, key="chk_malso_term")
    with col_chk[3]:
        chk_transfer = st.checkbox("📄 이관증명서", value=False, key="chk_malso_transfer")
    
    # 말소타입 약어 및 holder_name 미리 설정
    malso_prefix = {"근저당권": "근말", "질권": "질말", "전세권": "전말"}.get(malso_type, "말소")
    holder_name = st.session_state.get('malso_holder1_name', '고객') or '고객'
    
    # 체크된 항목에 대해 바로 다운로드 버튼 표시
    if chk_sig and PDF_OK:
        try:
            sig_template = resource_path("자필서명정보_서면_템플릿.pdf")
            if os.path.exists(sig_template):
                holders = []
                if st.session_state.get('malso_holder1_name'):
                    holders.append({
                        'name': st.session_state.get('malso_holder1_name', ''),
                        'rrn': st.session_state.get('malso_holder1_rrn', ''),
                        'addr': st.session_state.get('malso_holder1_addr', '')
                    })
                if st.session_state.get('malso_holder2_name'):
                    holders.append({
                        'name': st.session_state.get('malso_holder2_name', ''),
                        'rrn': st.session_state.get('malso_holder2_rrn', ''),
                        'addr': st.session_state.get('malso_holder2_addr', '')
                    })
                sig_data = {
                    'date': format_date_korean(st.session_state.get('malso_cause_date', datetime.now().date())),
                    'estate_list': st.session_state.get('malso_estate_detail', '').strip().split('\n'),
                    'holders': holders,
                    'purpose': f"{malso_type}말소"
                }
                pdf_buffer = make_malso_signature_pdf(sig_template, sig_data)
                st.download_button(
                    label="⬇️ 자필서명정보 다운로드",
                    data=pdf_buffer,
                    file_name=f"{malso_prefix}_{holder_name}_자필서명정보.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="dl_malso_sig"
                )
        except Exception as e:
            st.error(f"자필서명정보 생성 오류: {e}")
    
    if chk_power and PDF_OK:
        try:
            power_template_path = resource_path("말소_위임장.pdf")
            if os.path.exists(power_template_path):
                power_data = {
                    'date': format_date_korean(st.session_state.get('malso_cause_date', datetime.now().date())),
                    'malso_type': malso_type,
                    'obligor_label': obligor_label,
                    'obligor_name': st.session_state.get('malso_obligor_name', ''),
                    'obligor_id': st.session_state.get('malso_obligor_id', ''),
                    'obligor_addr': st.session_state.get('malso_obligor_addr', ''),
                    'obligor_rep': st.session_state.get('malso_obligor_rep', ''),
                    'obligor_branch': st.session_state.get('malso_obligor_branch', ''),
                    'holder1_name': st.session_state.get('malso_holder1_name', ''),
                    'holder1_addr': st.session_state.get('malso_holder1_addr', ''),
                    'holder2_name': st.session_state.get('malso_holder2_name', ''),
                    'holder2_addr': st.session_state.get('malso_holder2_addr', ''),
                    'estate_text': st.session_state.get('malso_estate_detail', ''),
                    'cancel_text': st.session_state.get('malso_cancel_text', '')
                }
                pdf_buffer = make_malso_power_pdf(power_template_path, power_data)
                st.download_button(
                    label="⬇️ 위임장 다운로드",
                    data=pdf_buffer,
                    file_name=f"{malso_prefix}_{holder_name}_위임장.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="dl_malso_power"
                )
        except Exception as e:
            st.error(f"위임장 생성 오류: {e}")
    
    if chk_term and PDF_OK:
        try:
            term_data = {
                'date': format_date_korean(st.session_state.get('malso_cause_date', datetime.now().date())),
                'malso_type': malso_type,
                'obligor_label': obligor_label,
                'obligor_name': st.session_state.get('malso_obligor_name', ''),
                'obligor_id': st.session_state.get('malso_obligor_id', ''),
                'obligor_addr': st.session_state.get('malso_obligor_addr', ''),
                'obligor_rep': st.session_state.get('malso_obligor_rep', ''),
                'obligor_branch': st.session_state.get('malso_obligor_branch', ''),
                'holder1_name': st.session_state.get('malso_holder1_name', ''),
                'holder2_name': st.session_state.get('malso_holder2_name', ''),
                'estate_text': st.session_state.get('malso_estate_detail', ''),
                'cancel_text': st.session_state.get('malso_cancel_text', '')
            }
            pdf_buffer = make_malso_termination_pdf(term_data)
            st.download_button(
                label="⬇️ 해지증서 다운로드",
                data=pdf_buffer,
                file_name=f"{malso_prefix}_{holder_name}_해지증서.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="dl_malso_term"
            )
        except Exception as e:
            st.error(f"해지증서 생성 오류: {e}")
    
    if chk_transfer and PDF_OK:
        try:
            transfer_data = {
                'date': format_date_korean(st.session_state.get('malso_cause_date', datetime.now().date())),
                'malso_type': malso_type,
                'obligor_label': obligor_label,
                'obligor_name': st.session_state.get('malso_obligor_name', ''),
                'obligor_id': st.session_state.get('malso_obligor_id', ''),
                'obligor_addr': st.session_state.get('malso_obligor_addr', ''),
                'obligor_rep': st.session_state.get('malso_obligor_rep', ''),
                'obligor_branch': st.session_state.get('malso_obligor_branch', ''),
                'estate_text': st.session_state.get('malso_estate_detail', ''),
                'cancel_text': st.session_state.get('malso_cancel_text', ''),
                'from_branch': st.session_state.get('malso_from_branch', ''),
                'to_branch': st.session_state.get('malso_to_branch', '')
            }
            pdf_buffer = make_malso_transfer_pdf(transfer_data)
            st.download_button(
                label="⬇️ 이관증명서 다운로드",
                data=pdf_buffer,
                file_name=f"{malso_prefix}_{holder_name}_이관증명서.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="dl_malso_transfer"
            )
        except Exception as e:
            st.error(f"이관증명서 생성 오류: {e}")
    
    # 안내 메시지
    st.info("💡 **사용 방법**: '📥 대부업(전자설정)내용 가져오기' 버튼을 눌러 소유자 정보와 부동산 표시를 자동으로 불러올 수 있습니다.")

# =============================================================================
# Tab 1: 시중은행
# =============================================================================
with tab1:
    col_header = st.columns([6, 1])
    col_header[0].markdown("### 🏦 시중은행 서류 (설정계약서/위임장/자필서명정보)")
    with col_header[1]:
        if st.button("🔄 초기화", key="reset_tab5", use_container_width=True, type="secondary"):
            st.session_state['tab5_bank'] = "하나은행"
            st.session_state['tab5_estate_input'] = ""
            st.session_state['tab5_date'] = datetime.now().date()
            st.session_state['tab5_amount'] = ""
            st.session_state['tab5_property_addr'] = ""
            for i in range(1, 4):
                st.session_state[f'tab5_owner{i}_name_input'] = ""
                st.session_state[f'tab5_owner{i}_rrn_input'] = ""
                st.session_state[f'tab5_owner{i}_addr_input'] = ""
            st.rerun()
    
    # 세션 상태 초기화
    if 'tab5_bank' not in st.session_state:
        st.session_state['tab5_bank'] = "하나은행"
    if 'tab5_estate' not in st.session_state:
        st.session_state['tab5_estate'] = ""
    if 'tab5_date' not in st.session_state:
        st.session_state['tab5_date'] = datetime.now().date()
    if 'tab5_amount' not in st.session_state:
        st.session_state['tab5_amount'] = ""
    if 'tab5_property_addr' not in st.session_state:
        st.session_state['tab5_property_addr'] = ""
    
    # 등기의무자 3명 초기화 (이름, 주민번호, 주소)
    for i in range(1, 4):
        if f'tab5_owner{i}_name_input' not in st.session_state:
            st.session_state[f'tab5_owner{i}_name_input'] = ''
        if f'tab5_owner{i}_rrn_input' not in st.session_state:
            st.session_state[f'tab5_owner{i}_rrn_input'] = ''
        if f'tab5_owner{i}_addr_input' not in st.session_state:
            st.session_state[f'tab5_owner{i}_addr_input'] = ''
    
    # 물건지 주소 복사 pending 처리 (위젯 렌더링 전에 실행)
    for i in range(1, 4):
        pending_key = f'_pending_tab5_owner{i}_addr'
        if st.session_state.get(pending_key):
            st.session_state[f'tab5_owner{i}_addr_input'] = st.session_state[pending_key]
            del st.session_state[pending_key]
    
    st.markdown("---")
    
    # 거래은행 선택
    st.markdown("#### 🏛️ 거래은행 선택")
    bank_cols = st.columns(3)
    bank_list = ["하나은행", "신한은행", "우리은행"]
    current_bank = st.session_state.get('tab5_bank', '하나은행')
    
    for i, bank in enumerate(bank_list):
        with bank_cols[i]:
            btn_type = "primary" if current_bank == bank else "secondary"
            if st.button(bank, key=f"tab5_bank_{bank}", type=btn_type, use_container_width=True):
                st.session_state['tab5_bank'] = bank
                st.rerun()
    
    selected_bank = st.session_state.get('tab5_bank', '하나은행')
    
    st.markdown("---")
    
    # 작성일자 & 채권최고액
    col_date, col_amount = st.columns(2)
    
    with col_date:
        st.markdown("#### 📅 작성일자")
        tab5_date = st.date_input(
            "작성일자",
            value=st.session_state.get('tab5_date', datetime.now().date()),
            key="tab5_date_input",
            label_visibility="collapsed"
        )
        st.session_state['tab5_date'] = tab5_date
    
    with col_amount:
        st.markdown("#### 💰 채권최고액")
        tab5_amount = st.text_input(
            "채권최고액",
            value=st.session_state.get('tab5_amount', ''),
            key="tab5_amount_input",
            placeholder="130,000,000",
            label_visibility="collapsed"
        )
        st.session_state['tab5_amount'] = tab5_amount
    
    st.markdown("---")
    
    # 등기부등본 업로드 및 부동산표시 추출
    st.markdown("#### 📋 부동산 표시")
    
    uploaded_registry = st.file_uploader(
        "📤 등기부등본 PDF 업로드 (인터넷등기소 열람용)", 
        type=['pdf'], 
        key='registry_upload_tab5'
    )
    
    # PDF 업로드 시 추출 버튼
    if uploaded_registry is not None:
        if st.button("📋 부동산표시 추출", key='extract_estate_btn_tab5', use_container_width=True):
            with st.spinner("등기부 분석 중..."):
                try:
                    data, debug = parse_registry_pdf(uploaded_registry)
                    
                    if debug["errors"]:
                        for err in debug["errors"]:
                            st.error(f"❌ {err}")
                    else:
                        formatted = format_estate_text(data)
                        st.session_state['tab5_estate_input'] = formatted
                        st.session_state['_tab5_extract_done'] = True
                        st.rerun()
                except Exception as e:
                    st.error(f"❌ PDF 파싱 오류: {e}")
    
    # 추출 완료 메시지
    if st.session_state.get('_tab5_extract_done'):
        st.success("✅ 부동산표시 추출 완료!")
        st.session_state['_tab5_extract_done'] = False
    
    # 초기값 설정
    if 'tab5_estate_input' not in st.session_state:
        st.session_state['tab5_estate_input'] = ''
    
    estate_text = st.text_area(
        "부동산 표시 (자동 추출 또는 직접 입력)",
        height=200,
        key='tab5_estate_input'
    )
    st.session_state['tab5_estate'] = estate_text
    
    # 물건지 주소 입력
    st.markdown("#### 📍 물건지 주소")
    tab5_property = st.text_input(
        "물건지 주소 (위택스 신고용)",
        value=st.session_state.get('tab5_property_addr', ''),
        key='tab5_property_addr_input',
        placeholder="서울특별시 강남구 테헤란로 123"
    )
    st.session_state['tab5_property_addr'] = tab5_property
    
    st.markdown("---")
    
    # 등기의무자 정보 입력 (3명까지)
    st.markdown("#### 👤 등기의무자 정보 (최대 3명)")
    
    # 주민번호 자동 하이픈 삽입 함수
    def auto_format_rrn_tab5(key):
        """6자리 입력 시 자동으로 '-' 삽입"""
        if key in st.session_state:
            val = st.session_state[key]
            clean_val = re.sub(r'[^0-9]', '', str(val))
            if len(clean_val) >= 6 and '-' not in val:
                st.session_state[key] = f"{clean_val[:6]}-{clean_val[6:13]}"
            elif len(clean_val) > 13:
                st.session_state[key] = f"{clean_val[:6]}-{clean_val[6:13]}"
    
    owner_cols = st.columns(3)
    
    for i in range(1, 4):
        with owner_cols[i-1]:
            st.markdown(f"**등기의무자 {i}**")
            st.text_input("성명", key=f'tab5_owner{i}_name_input', placeholder="홍길동")
            st.text_input(
                "주민등록번호", 
                key=f'tab5_owner{i}_rrn_input', 
                placeholder="000000-0000000",
                on_change=auto_format_rrn_tab5,
                args=(f'tab5_owner{i}_rrn_input',)
            )
            addr_col1, addr_col2 = st.columns([3, 2])
            with addr_col1:
                st.text_input("주소", key=f'tab5_owner{i}_addr_input', placeholder="서울시 강남구 테헤란로 123")
            with addr_col2:
                st.markdown("<div style='height: 28px'></div>", unsafe_allow_html=True)
                if st.button("📍 물건지 주소복사", key=f'tab5_addr_copy_{i}', use_container_width=True):
                    st.session_state[f'_pending_tab5_owner{i}_addr'] = st.session_state.get('tab5_property_addr', '')
                    st.rerun()
    
    st.markdown("---")
    
    # PDF 생성 버튼
    st.markdown("#### 📥 PDF 생성")
    
    btn_cols = st.columns(3)
    
    current_estate = st.session_state.get('tab5_estate', '')
    
    # 설정계약서 PDF 생성
    with btn_cols[0]:
        if st.button("📄 설정계약서", key="gen_contract_tab5", use_container_width=True, type="primary"):
            if current_estate:
                try:
                    pdf_buffer = make_bank_estate_pdf(selected_bank, "설정계약서", current_estate)
                    st.download_button(
                        label="⬇️ 설정계약서 다운로드",
                        data=pdf_buffer.getvalue(),
                        file_name=f"{selected_bank}_설정계약서.pdf",
                        mime="application/pdf",
                        key="dl_contract_tab5"
                    )
                except Exception as e:
                    st.error(f"❌ PDF 생성 오류: {e}")
            else:
                st.warning("⚠️ 부동산 표시를 입력하세요.")
    
    # 위임장 PDF 생성
    with btn_cols[1]:
        if st.button("📝 위임장", key="gen_power_tab5", use_container_width=True, type="primary"):
            if current_estate:
                try:
                    pdf_buffer = make_bank_estate_pdf(selected_bank, "위임장", current_estate)
                    st.download_button(
                        label="⬇️ 위임장 다운로드",
                        data=pdf_buffer.getvalue(),
                        file_name=f"{selected_bank}_위임장.pdf",
                        mime="application/pdf",
                        key="dl_power_tab5"
                    )
                except Exception as e:
                    st.error(f"❌ PDF 생성 오류: {e}")
            else:
                st.warning("⚠️ 부동산 표시를 입력하세요.")
    
    # 자필서명정보 PDF 생성 (입력된 등기의무자 수만큼)
    with btn_cols[2]:
        if st.button("✍️ 자필서명정보", key="gen_sig_tab5", use_container_width=True, type="primary"):
            if current_estate:
                try:
                    template_path = resource_path("자필서명정보_서면_템플릿.pdf")
                    if not os.path.exists(template_path):
                        st.error("❌ 자필서명정보 템플릿 파일이 없습니다.")
                    else:
                        # 날짜 포맷
                        date_val = st.session_state.get('tab5_date', datetime.now().date())
                        date_str = date_val.strftime("%Y년 %m월 %d일")
                        
                        # 입력된 등기의무자 수집
                        owners = []
                        for i in range(1, 4):
                            name = st.session_state.get(f'tab5_owner{i}_name_input', '').strip()
                            rrn = st.session_state.get(f'tab5_owner{i}_rrn_input', '').strip()
                            if name:  # 이름이 입력된 경우만
                                owners.append({'name': name, 'rrn': rrn})
                        
                        if not owners:
                            st.warning("⚠️ 등기의무자를 1명 이상 입력하세요.")
                        else:
                            st.session_state['_tab5_sig_pdfs'] = []
                            
                            # 각 등기의무자별로 PDF 생성
                            for idx, owner in enumerate(owners):
                                sig_data = {
                                    "estate_text": current_estate,
                                    "purpose": "근저당권설정",
                                    "debtor_name": owner['name'],
                                    "debtor_rrn": owner['rrn'],
                                    "owner_name": "",
                                    "owner_rrn": "",
                                    "date": date_str
                                }
                                pdf_buffer = make_bank_signature_pdf(template_path, sig_data)
                                st.session_state['_tab5_sig_pdfs'].append({
                                    'name': owner['name'],
                                    'data': pdf_buffer.getvalue()
                                })
                            
                            st.session_state['_tab5_sig_ready'] = True
                            st.rerun()
                            
                except Exception as e:
                    st.error(f"❌ PDF 생성 오류: {e}")
            else:
                st.warning("⚠️ 부동산 표시를 입력하세요.")
    
    # 자필서명정보 다운로드 버튼들 표시
    if st.session_state.get('_tab5_sig_ready') and st.session_state.get('_tab5_sig_pdfs'):
        st.markdown("---")
        st.success(f"✅ 자필서명정보 {len(st.session_state['_tab5_sig_pdfs'])}건 생성 완료!")
        
        dl_cols = st.columns(len(st.session_state['_tab5_sig_pdfs']))
        for idx, pdf_info in enumerate(st.session_state['_tab5_sig_pdfs']):
            with dl_cols[idx]:
                st.download_button(
                    label=f"⬇️ {pdf_info['name']}",
                    data=pdf_info['data'],
                    file_name=f"{selected_bank}_자필서명정보_{pdf_info['name']}.pdf",
                    mime="application/pdf",
                    key=f"dl_sig_tab5_{idx}",
                    use_container_width=True
                )
        
        st.session_state['_tab5_sig_ready'] = False
    
    # 입력된 등기의무자 수 카운트
    owner_count = sum(1 for i in range(1, 4) if st.session_state.get(f'tab5_owner{i}_name_input', '').strip())
    
    st.info(f"💡 **선택된 은행:** {selected_bank} | **작성일자:** {st.session_state.get('tab5_date', datetime.now().date())} | **등기의무자:** {owner_count}명")

# =============================================================================
# Tab 6: 위택스 등록면허세 신고
# =============================================================================
with tab6:
    col_header = st.columns([6, 1])
    col_header[0].markdown("### 🏛️ 위택스신고 (등록면허세)")
    with col_header[1]:
        if st.button("🔄 초기화", key="wetax_reset", use_container_width=True, type="secondary"):
            # 위택스 관련 세션 상태 초기화
            st.session_state['wetax_data_source'] = ''
            st.session_state['wetax_report_type'] = '설정'
            st.session_state['wetax_manual_list'] = []
            st.session_state['wetax_malso_list'] = []
            st.session_state['wetax_creditor_name'] = ''
            st.session_state['wetax_creditor_corp_num'] = ''
            st.session_state['wetax_creditor_addr'] = ''
            st.session_state['wetax_debtor_name'] = ''
            st.session_state['wetax_debtor_rrn'] = ''
            st.session_state['wetax_debtor_addr'] = ''
            st.session_state['wetax_owner_name'] = ''
            st.session_state['wetax_owner_rrn'] = ''
            st.session_state['wetax_owner_addr'] = ''
            st.session_state['wetax_property_addr'] = ''
            st.session_state['wetax_amount'] = ''
            st.session_state['wetax_contract_type'] = '개인'
            st.session_state['wetax_tab1_owners'] = []
            st.success("✅ 초기화되었습니다!")
            st.rerun()
    
    # =========================================================================
    # 터널 연결 설정
    # =========================================================================
    st.markdown("#### ⚙️ 터널 연결 설정")
    
    if 'wetax_server_url' not in st.session_state:
        st.session_state['wetax_server_url'] = ''
    
    with st.container(border=True):
        st.caption("wetax_launcher.exe 실행 후 생성된 URL을 붙여넣으세요")
        
        col_url, col_test = st.columns([4, 1])
        with col_url:
            wetax_url = st.text_input(
                "서버 URL",
                value=st.session_state.get('wetax_server_url', ''),
                placeholder="https://xxxx.trycloudflare.com",
                key='tab6_wetax_url_input',
                label_visibility='collapsed'
            )
            if wetax_url != st.session_state.get('wetax_server_url', ''):
                st.session_state['wetax_server_url'] = wetax_url
        
        with col_test:
            if st.button("🔗 연결 테스트", key='tab6_wetax_test_conn', use_container_width=True):
                if wetax_url:
                    try:
                        test_url = wetax_url.rstrip('/') + "/"
                        resp = requests.get(test_url, timeout=15)
                        if resp.status_code == 200:
                            st.success("✅ 연결 성공!")
                        else:
                            st.error(f"❌ 연결 실패 (상태코드: {resp.status_code})")
                    except Exception as e:
                        st.error(f"❌ 연결 실패: {e}")
                else:
                    st.warning("⚠️ URL을 입력하세요")
        
        # 연결 상태 표시
        if st.session_state.get('wetax_server_url'):
            st.success(f"🟢 연결됨: {st.session_state['wetax_server_url']}")
        else:
            st.warning("🔴 연결되지 않음")
    
    st.markdown("---")
    
    # =========================================================================
    # 데이터 불러오기
    # =========================================================================
    st.markdown("#### 📥 데이터 불러오기")
    
    # 위택스 신고용 세션 초기화
    if 'wetax_report_type' not in st.session_state:
        st.session_state['wetax_report_type'] = '설정'
    if 'wetax_data_source' not in st.session_state:
        st.session_state['wetax_data_source'] = ''
    if 'wetax_manual_list' not in st.session_state:
        st.session_state['wetax_manual_list'] = []
    
    btn_cols = st.columns(4)
    
    with btn_cols[0]:
        if st.button("🏦 1탭 가져오기\n(시중은행 설정)", key="wetax_load_tab1", use_container_width=True, type="primary"):
            # 1탭(시중은행)에서 데이터 가져오기
            st.session_state['wetax_data_source'] = 'tab1'
            st.session_state['wetax_report_type'] = '설정'
            
            # 은행 정보
            selected_bank = st.session_state.get('tab5_bank', '하나은행')
            BANK_INFO = {
                "하나은행": {"name": "주식회사 하나은행", "corp_num": "110111-0672538", "addr": "서울특별시 중구 을지로 35(을지로1가)"},
                "신한은행": {"name": "주식회사 신한은행", "corp_num": "110111-0012809", "addr": "서울특별시 중구 세종대로9길 20(태평로2가)"},
                "우리은행": {"name": "주식회사 우리은행", "corp_num": "110111-0023393", "addr": "서울특별시 중구 소공로51(회현동1가)"}
            }
            bank_info = BANK_INFO.get(selected_bank, {})
            
            st.session_state['wetax_creditor_name'] = bank_info.get('name', selected_bank)
            st.session_state['wetax_creditor_corp_num'] = bank_info.get('corp_num', '')
            st.session_state['wetax_creditor_addr'] = bank_info.get('addr', '')
            
            # 등기의무자 정보 (최대 3명)
            owners = []
            for i in range(1, 4):
                name = st.session_state.get(f'tab5_owner{i}_name_input', '').strip()
                rrn = st.session_state.get(f'tab5_owner{i}_rrn_input', '').strip()
                addr = st.session_state.get(f'tab5_owner{i}_addr_input', '').strip()
                if name:
                    owners.append({'name': name, 'rrn': rrn, 'addr': addr})
            st.session_state['wetax_tab1_owners'] = owners
            
            # 물건지, 채권최고액
            st.session_state['wetax_property_addr'] = st.session_state.get('tab5_property_addr', '')
            st.session_state['wetax_amount'] = st.session_state.get('tab5_amount', '')
            
            st.success("✅ 1탭(시중은행) 데이터 불러오기 완료!")
            st.rerun()
    
    with btn_cols[1]:
        if st.button("📄 2탭 가져오기\n(대부업 설정)", key="wetax_load_tab2", use_container_width=True, type="primary"):
            # 2탭(대부업)에서 데이터 가져오기
            st.session_state['wetax_data_source'] = 'tab2'
            st.session_state['wetax_report_type'] = '설정'
            
            # 채권자(근저당권자) 정보
            st.session_state['wetax_creditor_name'] = st.session_state.get('input_creditor_name', '') or st.session_state.get('input_creditor', '')
            st.session_state['wetax_creditor_corp_num'] = st.session_state.get('input_creditor_corp_num', '')
            st.session_state['wetax_creditor_addr'] = st.session_state.get('input_creditor_addr', '')
            
            # 채무자/소유자 정보
            st.session_state['wetax_debtor_name'] = st.session_state.get('t1_debtor_name', '')
            st.session_state['wetax_debtor_rrn'] = st.session_state.get('t1_debtor_rrn', '')
            st.session_state['wetax_debtor_addr'] = st.session_state.get('t1_debtor_addr', '')
            st.session_state['wetax_owner_name'] = st.session_state.get('t1_owner_name', '')
            st.session_state['wetax_owner_rrn'] = st.session_state.get('t1_owner_rrn', '')
            st.session_state['wetax_owner_addr'] = st.session_state.get('t1_owner_addr', '')
            
            # 물건지, 채권최고액, 계약유형
            st.session_state['wetax_property_addr'] = st.session_state.get('input_collateral_addr', '')
            st.session_state['wetax_amount'] = st.session_state.get('input_amount', '')
            st.session_state['wetax_contract_type'] = st.session_state.get('contract_type', '개인')
            
            st.success("✅ 2탭(대부업) 데이터 불러오기 완료!")
            st.rerun()
    
    with btn_cols[2]:
        if st.button("🗑️ 3탭 가져오기\n(말소)", key="wetax_load_tab3", use_container_width=True, type="primary"):
            # 3탭(말소)에서 임시저장 목록 가져오기
            st.session_state['wetax_data_source'] = 'tab3'
            st.session_state['wetax_report_type'] = '말소'
            
            # 임시저장 목록 가져오기
            saved_list = st.session_state.get('malso_saved_list', [])
            
            if saved_list:
                st.session_state['wetax_malso_list'] = saved_list.copy()
                st.success(f"✅ 3탭(말소) {len(saved_list)}건 불러오기 완료!")
            else:
                # 임시저장이 없으면 현재 입력된 1건만 가져오기
                holder_name = st.session_state.get('malso_holder1_name', '')
                if holder_name:
                    st.session_state['wetax_malso_list'] = [{
                        'type': st.session_state.get('malso_type', '근저당권'),
                        'holder_name': holder_name,
                        'holder_rrn': st.session_state.get('malso_holder1_rrn', ''),
                        'holder_addr': st.session_state.get('malso_holder1_addr', ''),
                        'property_addr': st.session_state.get('malso_property_addr', '')
                    }]
                    st.success("✅ 3탭(말소) 1건 불러오기 완료!")
                else:
                    st.session_state['wetax_malso_list'] = []
                    st.warning("⚠️ 3탭에 저장된 말소 건이 없습니다.")
            st.rerun()
    
    with btn_cols[3]:
        if st.button("✍️ 수기입력\n(직접 입력)", key="wetax_load_manual", use_container_width=True, type="secondary"):
            st.session_state['wetax_data_source'] = 'manual'
            st.session_state['wetax_report_type'] = '수기'
            st.session_state['wetax_manual_list'] = []
            st.rerun()
    
    st.markdown("---")
    
    # =========================================================================
    # 신고 내용 확인 및 실행
    # =========================================================================
    data_source = st.session_state.get('wetax_data_source', '')
    report_type = st.session_state.get('wetax_report_type', '설정')
    
    if data_source:
        st.markdown(f"#### 📋 신고 내용 확인 (데이터 출처: **{data_source.upper()}**)")
        
        with st.container(border=True):
            # 미리보기 데이터 수집
            preview_rows = []
            
            if data_source == 'tab2':
                # 1탭: 설정/주소변경
                contract_type = st.session_state.get('wetax_contract_type', '개인')
                creditor_name = st.session_state.get('wetax_creditor_name', '')
                creditor_corp_num = st.session_state.get('wetax_creditor_corp_num', '')
                creditor_addr = st.session_state.get('wetax_creditor_addr', '')
                property_addr = st.session_state.get('wetax_property_addr', '')
                amount = st.session_state.get('wetax_amount', '')
                
                # 근저당설정 (필수)
                preview_rows.append({
                    '선택': True,
                    '신고유형': '근저당설정',
                    '납세자': creditor_name,
                    '주민/법인번호': creditor_corp_num,
                    '주소': creditor_addr[:30] + '...' if len(creditor_addr) > 30 else creditor_addr,
                    '물건지': property_addr[:25] + '...' if len(property_addr) > 25 else property_addr,
                    '채권최고액': amount
                })
                
                # 미리보기 테이블 표시
                st.markdown("**📊 신고 미리보기**")
                st.dataframe(
                    preview_rows,
                    column_config={
                        "선택": st.column_config.CheckboxColumn("선택", default=True),
                        "신고유형": "신고유형",
                        "납세자": "납세자",
                        "주민/법인번호": "주민/법인번호",
                        "주소": "주소",
                        "물건지": "물건지",
                        "채권최고액": "채권최고액"
                    },
                    hide_index=True,
                    use_container_width=True
                )
                
                st.markdown("---")
                
                # 주소변경 옵션
                include_addr = st.checkbox("📍 **주소변경 포함**", key='wetax_include_addr_change')
                
                if include_addr:
                    addr_preview = []
                    if contract_type == "개인":
                        debtor_name = st.session_state.get('wetax_debtor_name', '')
                        debtor_rrn = st.session_state.get('wetax_debtor_rrn', '')
                        debtor_addr = st.session_state.get('wetax_debtor_addr', '')
                        addr_preview.append({
                            '납세자': debtor_name, '주민번호': debtor_rrn, 
                            '주소': debtor_addr[:30] + '...' if len(debtor_addr) > 30 else debtor_addr,
                            '물건지': property_addr[:25] + '...' if len(property_addr) > 25 else property_addr,
                            '신고유형': '주소변경'
                        })
                        correction = st.checkbox("      └─ 경정 포함 (2건 신고)", key='wetax_include_correction')
                        if correction:
                            addr_preview.insert(0, {
                                '납세자': debtor_name, '주민번호': debtor_rrn,
                                '주소': debtor_addr[:30] + '...' if len(debtor_addr) > 30 else debtor_addr,
                                '물건지': property_addr[:25] + '...' if len(property_addr) > 25 else property_addr,
                                '신고유형': '경정'
                            })
                    elif contract_type == "3자담보":
                        owner_name = st.session_state.get('wetax_owner_name', '')
                        owner_rrn = st.session_state.get('wetax_owner_rrn', '')
                        owner_addr = st.session_state.get('wetax_owner_addr', '')
                        addr_preview.append({
                            '납세자': owner_name, '주민번호': owner_rrn,
                            '주소': owner_addr[:30] + '...' if len(owner_addr) > 30 else owner_addr,
                            '물건지': property_addr[:25] + '...' if len(property_addr) > 25 else property_addr,
                            '신고유형': '주소변경'
                        })
                        correction = st.checkbox("      └─ 경정 포함 (2건 신고)", key='wetax_include_correction')
                        if correction:
                            addr_preview.insert(0, {
                                '납세자': owner_name, '주민번호': owner_rrn,
                                '주소': owner_addr[:30] + '...' if len(owner_addr) > 30 else owner_addr,
                                '물건지': property_addr[:25] + '...' if len(property_addr) > 25 else property_addr,
                                '신고유형': '경정'
                            })
                    else:  # 공동담보
                        owner_name = st.session_state.get('wetax_owner_name', '')
                        owner_rrn = st.session_state.get('wetax_owner_rrn', '')
                        owner_addr = st.session_state.get('wetax_owner_addr', '')
                        debtor_name = st.session_state.get('wetax_debtor_name', '')
                        debtor_rrn = st.session_state.get('wetax_debtor_rrn', '')
                        debtor_addr = st.session_state.get('wetax_debtor_addr', '')
                        
                        col_o, col_d = st.columns(2)
                        with col_o:
                            addr_owner = st.checkbox(f"소유자 ({owner_name})", key='wetax_addr_owner')
                            if addr_owner:
                                addr_preview.append({'납세자': owner_name, '주민번호': owner_rrn, '주소': owner_addr[:25]+'...' if len(owner_addr)>25 else owner_addr, '물건지': property_addr[:20]+'...' if len(property_addr)>20 else property_addr, '신고유형': '주소변경'})
                                if st.checkbox("   └─ 경정 포함", key='wetax_owner_correction'):
                                    addr_preview.insert(len(addr_preview)-1, {'납세자': owner_name, '주민번호': owner_rrn, '주소': owner_addr[:25]+'...' if len(owner_addr)>25 else owner_addr, '물건지': property_addr[:20]+'...' if len(property_addr)>20 else property_addr, '신고유형': '경정'})
                        with col_d:
                            addr_debtor = st.checkbox(f"채무자 ({debtor_name})", key='wetax_addr_debtor')
                            if addr_debtor:
                                addr_preview.append({'납세자': debtor_name, '주민번호': debtor_rrn, '주소': debtor_addr[:25]+'...' if len(debtor_addr)>25 else debtor_addr, '물건지': property_addr[:20]+'...' if len(property_addr)>20 else property_addr, '신고유형': '주소변경'})
                                if st.checkbox("   └─ 경정 포함", key='wetax_debtor_correction'):
                                    addr_preview.insert(len(addr_preview)-1, {'납세자': debtor_name, '주민번호': debtor_rrn, '주소': debtor_addr[:25]+'...' if len(debtor_addr)>25 else debtor_addr, '물건지': property_addr[:20]+'...' if len(property_addr)>20 else property_addr, '신고유형': '경정'})
                    
                    if addr_preview:
                        st.dataframe(addr_preview, hide_index=True, use_container_width=True)
            
            elif data_source == 'tab3':
                # 4탭: 말소 (여러 건 지원)
                malso_list = st.session_state.get('wetax_malso_list', [])
                
                if malso_list:
                    st.markdown(f"**📊 말소 신고 미리보기 ({len(malso_list)}건)**")
                    
                    # 테이블 형태로 표시
                    malso_preview = []
                    for idx, item in enumerate(malso_list):
                        malso_preview.append({
                            '신고유형': f"{item['type']}말소",
                            '납세자': item['holder_name'],
                            '주민번호': item['holder_rrn'],
                            '주소': item['holder_addr'][:30] + '...' if len(item.get('holder_addr', '')) > 30 else item.get('holder_addr', ''),
                            '물건지': item['property_addr'][:25] + '...' if len(item.get('property_addr', '')) > 25 else item.get('property_addr', '')
                        })
                    
                    st.dataframe(malso_preview, hide_index=True, use_container_width=True)
                    
                    st.markdown("---")
                    st.markdown("**신고 대상 선택**")
                    
                    # 각 말소건에 대해 체크박스 생성
                    chk_cols = st.columns(min(len(malso_list), 3))
                    for idx, item in enumerate(malso_list):
                        with chk_cols[idx % 3]:
                            st.checkbox(
                                f"{item['type']}말소 - {item['holder_name']}", 
                                value=True, 
                                key=f'wetax_malso_chk_{idx}'
                            )
                else:
                    st.warning("⚠️ 말소 건이 없습니다. 3탭에서 임시저장 후 다시 불러오세요.")
            
            elif data_source == 'tab1':
                # 5탭: 1금융권 설정/주소변경
                creditor_name = st.session_state.get('wetax_creditor_name', '')
                creditor_corp_num = st.session_state.get('wetax_creditor_corp_num', '')
                creditor_addr = st.session_state.get('wetax_creditor_addr', '')
                property_addr = st.session_state.get('wetax_property_addr', '')
                amount = st.session_state.get('wetax_amount', '')
                owners = st.session_state.get('wetax_tab1_owners', [])
                
                # 근저당설정 미리보기
                st.markdown("**📊 신고 미리보기**")
                setting_preview = [{
                    '신고유형': '근저당설정',
                    '납세자': creditor_name,
                    '주민/법인번호': creditor_corp_num,
                    '주소': creditor_addr[:30] + '...' if len(creditor_addr) > 30 else creditor_addr,
                    '물건지': property_addr[:25] + '...' if len(property_addr) > 25 else property_addr,
                    '채권최고액': amount
                }]
                st.dataframe(setting_preview, hide_index=True, use_container_width=True)
                
                st.markdown("---")
                
                # 주소변경 옵션
                include_addr = st.checkbox("📍 **주소변경 포함** (등기의무자 대상)", key='wetax_include_addr_tab5')
                
                if include_addr and owners:
                    addr_preview = []
                    for idx, owner in enumerate(owners):
                        addr_preview.append({
                            '신고유형': '주소변경',
                            '납세자': owner['name'],
                            '주민번호': owner['rrn'],
                            '주소': owner['addr'][:30] + '...' if len(owner.get('addr', '')) > 30 else owner.get('addr', ''),
                            '물건지': property_addr[:25] + '...' if len(property_addr) > 25 else property_addr
                        })
                    
                    st.dataframe(addr_preview, hide_index=True, use_container_width=True)
                    
                    correction = st.checkbox("      └─ 경정 포함 (각 2건씩 신고)", key='wetax_correction_tab5')
                    
                    if correction:
                        st.caption(f"   → 경정 {len(owners)}건 + 주소변경 {len(owners)}건 = 총 {len(owners)*2}건 추가")
            
            elif data_source == 'manual':
                # 수기입력 모드
                st.markdown("**✍️ 수기입력 모드**")
                
                # 주민번호 자동 포맷팅 함수
                def auto_format_manual_rrn(key):
                    if key in st.session_state:
                        val = st.session_state[key]
                        clean_val = re.sub(r'[^0-9]', '', str(val))
                        if len(clean_val) >= 6 and '-' not in val:
                            st.session_state[key] = f"{clean_val[:6]}-{clean_val[6:13]}"
                
                # 신고 건 추가 폼
                with st.expander("➕ 신고 건 추가", expanded=True):
                    # 신고 유형 선택
                    report_type_options = ["설정", "전세권설정", "주소변경", "경정", "말소", "기타"]
                    col_type, col_taxpayer = st.columns(2)
                    
                    with col_type:
                        manual_report_type = st.selectbox("신고유형", report_type_options, key='manual_report_type')
                    
                    with col_taxpayer:
                        taxpayer_type = st.radio("납세자 유형", ["개인", "법인"], horizontal=True, key='manual_taxpayer_type')
                    
                    # 납세자 정보
                    st.markdown("**납세자 정보**")
                    col1, col2 = st.columns(2)
                    with col1:
                        manual_name = st.text_input("납세자명", key='manual_taxpayer_name', placeholder="홍길동 / ㈜OO저축은행")
                    with col2:
                        if taxpayer_type == "개인":
                            manual_rrn = st.text_input("주민등록번호", key='manual_taxpayer_rrn', placeholder="800101-1234567",
                                                       on_change=auto_format_manual_rrn, args=('manual_taxpayer_rrn',))
                        else:
                            manual_rrn = st.text_input("법인등록번호", key='manual_taxpayer_rrn', placeholder="110111-1234567",
                                                       on_change=auto_format_manual_rrn, args=('manual_taxpayer_rrn',))
                    
                    manual_addr = st.text_input("납세자 주소", key='manual_taxpayer_addr', placeholder="서울특별시 강남구 테헤란로 123")
                    
                    # 물건지 및 채권최고액/과세표준
                    st.markdown("**물건지 정보**")
                    manual_property = st.text_input("물건지 주소", key='manual_property_addr', placeholder="서울특별시 서초구 서초대로 456")
                    
                    # 설정, 전세권설정은 과세표준(채권최고액) 입력 필요
                    if manual_report_type in ["설정", "전세권설정"]:
                        amount_label = "전세금" if manual_report_type == "전세권설정" else "채권최고액"
                        manual_amount = st.text_input(amount_label, key='manual_amount', placeholder="130,000,000")
                    else:
                        manual_amount = ""
                    
                    # 추가 버튼
                    if st.button("➕ 신고 건 추가", key='manual_add_btn', use_container_width=True):
                        if not manual_name:
                            st.error("❌ 납세자명을 입력하세요!")
                        else:
                            new_item = {
                                'report_type': manual_report_type,
                                'taxpayer_type': '01' if taxpayer_type == '개인' else '02',
                                'taxpayer_type_name': taxpayer_type,
                                'name': manual_name,
                                'rrn': st.session_state.get('manual_taxpayer_rrn', ''),
                                'addr': manual_addr,
                                'property_addr': manual_property,
                                'amount': manual_amount if manual_report_type in ["설정", "전세권설정"] else ''
                            }
                            st.session_state['wetax_manual_list'].append(new_item)
                            
                            # 입력 필드 초기화
                            st.session_state['manual_taxpayer_name'] = ''
                            st.session_state['manual_taxpayer_rrn'] = ''
                            st.session_state['manual_taxpayer_addr'] = ''
                            st.session_state['manual_property_addr'] = ''
                            st.session_state['manual_amount'] = ''
                            
                            st.success(f"✅ {manual_report_type} - {manual_name} 추가 완료!")
                            st.rerun()
                
                # 추가된 신고 건 목록
                manual_list = st.session_state.get('wetax_manual_list', [])
                
                if manual_list:
                    st.markdown("---")
                    st.markdown(f"**📊 신고 목록 ({len(manual_list)}건)**")
                    
                    # 테이블 형태로 표시
                    manual_preview = []
                    for idx, item in enumerate(manual_list):
                        manual_preview.append({
                            '신고유형': item['report_type'],
                            '납세자유형': item['taxpayer_type_name'],
                            '납세자': item['name'],
                            '주민/법인번호': item['rrn'],
                            '주소': item['addr'][:25] + '...' if len(item.get('addr', '')) > 25 else item.get('addr', ''),
                            '물건지': item['property_addr'][:20] + '...' if len(item.get('property_addr', '')) > 20 else item.get('property_addr', ''),
                            '과세표준': item.get('amount', '')
                        })
                    
                    st.dataframe(manual_preview, hide_index=True, use_container_width=True)
                    
                    # 삭제 버튼들
                    st.markdown("**삭제**")
                    del_cols = st.columns(min(len(manual_list), 5))
                    for idx, item in enumerate(manual_list):
                        with del_cols[idx % 5]:
                            if st.button(f"🗑️ {item['name']}", key=f'manual_del_{idx}'):
                                st.session_state['wetax_manual_list'].pop(idx)
                                st.rerun()
                else:
                    st.info("💡 위 폼에서 신고 건을 추가하세요.")
            
            st.markdown("---")
            
            # 신고 실행 버튼
            if st.button("🚀 위택스 신고 실행", type="primary", use_container_width=True, key='wetax_final_submit'):
                wetax_url = st.session_state.get('wetax_server_url', '')
                
                if not wetax_url:
                    st.error("❌ 터널 연결을 먼저 설정하세요!")
                else:
                    cases = []
                    
                    if data_source == 'tab2':
                        # 1탭: 설정/주소변경
                        creditor_name = st.session_state.get('wetax_creditor_name', '')
                        creditor_corp_num = st.session_state.get('wetax_creditor_corp_num', '')
                        creditor_addr = st.session_state.get('wetax_creditor_addr', '')
                        property_addr = st.session_state.get('wetax_property_addr', '')
                        tax_base = remove_commas(st.session_state.get('wetax_amount', '0'))
                        contract_type = st.session_state.get('wetax_contract_type', '개인')
                        
                        front, back = parse_corp_num(creditor_corp_num)
                        road_addr, detail_addr = extract_road_address(creditor_addr)
                        prop_road, prop_detail = extract_road_address(property_addr)
                        
                        # 근저당설정
                        cases.append({
                            "type": "설정", "taxpayer_type": "02", "taxpayer_name": creditor_name,
                            "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                            "address": road_addr, "address_detail": detail_addr,
                            "property_address": prop_road, "property_detail": prop_detail,
                            "tax_base": int(tax_base) if tax_base else 0
                        })
                        
                        # 주소변경
                        if st.session_state.get('wetax_include_addr_change'):
                            if contract_type == "개인":
                                debtor_name = st.session_state.get('wetax_debtor_name', '')
                                debtor_rrn = st.session_state.get('wetax_debtor_rrn', '')
                                debtor_addr = st.session_state.get('wetax_debtor_addr', '')
                                front, back = parse_rrn(debtor_rrn)
                                road_addr, detail_addr = extract_road_address(debtor_addr)
                                
                                if st.session_state.get('wetax_include_correction'):
                                    cases.append({"type": "변경", "taxpayer_type": "01", "taxpayer_name": debtor_name,
                                        "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                                        "address": road_addr, "address_detail": detail_addr,
                                        "property_address": prop_road, "property_detail": prop_detail, "tax_base": None})
                                cases.append({"type": "주소변경", "taxpayer_type": "01", "taxpayer_name": debtor_name,
                                    "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                                    "address": road_addr, "address_detail": detail_addr,
                                    "property_address": prop_road, "property_detail": prop_detail, "tax_base": None})
                            
                            elif contract_type == "3자담보":
                                owner_name = st.session_state.get('wetax_owner_name', '')
                                owner_rrn = st.session_state.get('wetax_owner_rrn', '')
                                owner_addr = st.session_state.get('wetax_owner_addr', '')
                                front, back = parse_rrn(owner_rrn)
                                road_addr, detail_addr = extract_road_address(owner_addr)
                                
                                if st.session_state.get('wetax_include_correction'):
                                    cases.append({"type": "변경", "taxpayer_type": "01", "taxpayer_name": owner_name,
                                        "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                                        "address": road_addr, "address_detail": detail_addr,
                                        "property_address": prop_road, "property_detail": prop_detail, "tax_base": None})
                                cases.append({"type": "주소변경", "taxpayer_type": "01", "taxpayer_name": owner_name,
                                    "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                                    "address": road_addr, "address_detail": detail_addr,
                                    "property_address": prop_road, "property_detail": prop_detail, "tax_base": None})
                            
                            else:  # 공동담보
                                if st.session_state.get('wetax_addr_owner'):
                                    owner_name = st.session_state.get('wetax_owner_name', '')
                                    owner_rrn = st.session_state.get('wetax_owner_rrn', '')
                                    owner_addr = st.session_state.get('wetax_owner_addr', '')
                                    front, back = parse_rrn(owner_rrn)
                                    road_addr, detail_addr = extract_road_address(owner_addr)
                                    
                                    if st.session_state.get('wetax_owner_correction'):
                                        cases.append({"type": "변경", "taxpayer_type": "01", "taxpayer_name": owner_name,
                                            "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                                            "address": road_addr, "address_detail": detail_addr,
                                            "property_address": prop_road, "property_detail": prop_detail, "tax_base": None})
                                    cases.append({"type": "주소변경", "taxpayer_type": "01", "taxpayer_name": owner_name,
                                        "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                                        "address": road_addr, "address_detail": detail_addr,
                                        "property_address": prop_road, "property_detail": prop_detail, "tax_base": None})
                                
                                if st.session_state.get('wetax_addr_debtor'):
                                    debtor_name = st.session_state.get('wetax_debtor_name', '')
                                    debtor_rrn = st.session_state.get('wetax_debtor_rrn', '')
                                    debtor_addr = st.session_state.get('wetax_debtor_addr', '')
                                    front, back = parse_rrn(debtor_rrn)
                                    road_addr, detail_addr = extract_road_address(debtor_addr)
                                    
                                    if st.session_state.get('wetax_debtor_correction'):
                                        cases.append({"type": "변경", "taxpayer_type": "01", "taxpayer_name": debtor_name,
                                            "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                                            "address": road_addr, "address_detail": detail_addr,
                                            "property_address": prop_road, "property_detail": prop_detail, "tax_base": None})
                                    cases.append({"type": "주소변경", "taxpayer_type": "01", "taxpayer_name": debtor_name,
                                        "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                                        "address": road_addr, "address_detail": detail_addr,
                                        "property_address": prop_road, "property_detail": prop_detail, "tax_base": None})
                    
                    elif data_source == 'tab3':
                        # 4탭: 말소 (여러 건 처리)
                        malso_list = st.session_state.get('wetax_malso_list', [])
                        
                        for idx, item in enumerate(malso_list):
                            # 체크된 항목만 신고
                            if st.session_state.get(f'wetax_malso_chk_{idx}', True):
                                front, back = parse_rrn(item.get('holder_rrn', ''))
                                road_addr, detail_addr = extract_road_address(item.get('holder_addr', ''))
                                prop_road, prop_detail = extract_road_address(item.get('property_addr', ''))
                                
                                cases.append({
                                    "type": "말소", 
                                    "taxpayer_type": "01", 
                                    "taxpayer_name": item.get('holder_name', ''),
                                    "resident_no_front": front, 
                                    "resident_no_back": back, 
                                    "phone": "0218335482",
                                    "address": road_addr, 
                                    "address_detail": detail_addr,
                                    "property_address": prop_road, 
                                    "property_detail": prop_detail, 
                                    "tax_base": None
                                })
                    
                    elif data_source == 'tab1':
                        # 5탭: 1금융권 설정/주소변경
                        creditor_name = st.session_state.get('wetax_creditor_name', '')
                        creditor_corp_num = st.session_state.get('wetax_creditor_corp_num', '')
                        creditor_addr = st.session_state.get('wetax_creditor_addr', '')
                        property_addr = st.session_state.get('wetax_property_addr', '')
                        tax_base = remove_commas(st.session_state.get('wetax_amount', '0'))
                        owners = st.session_state.get('wetax_tab1_owners', [])
                        
                        front, back = parse_corp_num(creditor_corp_num)
                        road_addr, detail_addr = extract_road_address(creditor_addr)
                        prop_road, prop_detail = extract_road_address(property_addr)
                        
                        # 근저당설정
                        cases.append({
                            "type": "설정", "taxpayer_type": "02", "taxpayer_name": creditor_name,
                            "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                            "address": road_addr, "address_detail": detail_addr,
                            "property_address": prop_road, "property_detail": prop_detail,
                            "tax_base": int(tax_base) if tax_base else 0
                        })
                        
                        # 주소변경
                        if st.session_state.get('wetax_include_addr_tab5') and owners:
                            for owner in owners:
                                front, back = parse_rrn(owner.get('rrn', ''))
                                road_addr, detail_addr = extract_road_address(owner.get('addr', ''))
                                
                                if st.session_state.get('wetax_correction_tab5'):
                                    cases.append({"type": "변경", "taxpayer_type": "01", "taxpayer_name": owner['name'],
                                        "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                                        "address": road_addr, "address_detail": detail_addr,
                                        "property_address": prop_road, "property_detail": prop_detail, "tax_base": None})
                                cases.append({"type": "주소변경", "taxpayer_type": "01", "taxpayer_name": owner['name'],
                                    "resident_no_front": front, "resident_no_back": back, "phone": "0218335482",
                                    "address": road_addr, "address_detail": detail_addr,
                                    "property_address": prop_road, "property_detail": prop_detail, "tax_base": None})
                    
                    elif data_source == 'manual':
                        # 수기입력: 추가된 모든 건 처리
                        manual_list = st.session_state.get('wetax_manual_list', [])
                        
                        for item in manual_list:
                            # 주민/법인번호 파싱
                            if item['taxpayer_type'] == '01':  # 개인
                                front, back = parse_rrn(item.get('rrn', ''))
                            else:  # 법인
                                front, back = parse_corp_num(item.get('rrn', ''))
                            
                            road_addr, detail_addr = extract_road_address(item.get('addr', ''))
                            prop_road, prop_detail = extract_road_address(item.get('property_addr', ''))
                            
                            # 신고 유형에 따른 처리
                            report_type = item.get('report_type', '설정')
                            tax_base = None
                            
                            if report_type == '설정':
                                tax_base = int(remove_commas(item.get('amount', '0'))) if item.get('amount') else 0
                                api_type = "설정"
                            elif report_type == '전세권설정':
                                tax_base = int(remove_commas(item.get('amount', '0'))) if item.get('amount') else 0
                                api_type = "전세권설정"
                            elif report_type == '기타':
                                # 기타는 정액등록세 (과세표준 불필요)
                                api_type = "기타"
                            elif report_type == '주소변경':
                                api_type = "주소변경"
                            elif report_type == '경정':
                                api_type = "변경"
                            elif report_type == '말소':
                                api_type = "말소"
                            else:
                                api_type = report_type
                            
                            cases.append({
                                "type": api_type,
                                "taxpayer_type": item['taxpayer_type'],
                                "taxpayer_name": item['name'],
                                "resident_no_front": front,
                                "resident_no_back": back,
                                "phone": "0218335482",
                                "address": road_addr,
                                "address_detail": detail_addr,
                                "property_address": prop_road,
                                "property_detail": prop_detail,
                                "tax_base": tax_base
                            })
                    
                    # API 호출
                    if cases:
                        st.info(f"📤 총 {len(cases)}건 신고 중...")
                        result, error = call_wetax_api(cases, base_url=wetax_url)
                        
                        if error:
                            st.error(f"❌ 오류: {error}")
                        else:
                            st.success(f"✅ 위택스 신고 완료! ({len(cases)}건)")
                            st.json(result)
                    else:
                        st.warning("⚠️ 신고할 내용이 없습니다.")
    else:
        st.info("💡 위 버튼을 눌러 각 탭에서 데이터를 불러오세요.")

# =============================================================================
# Tab 7: 법인등기
# =============================================================================
with tab7:
    st.markdown("### 🏢 법인등기 비용계산기")
    
    # 법인등기 종류별 비용 데이터
    CORP_REGISTRY_FEES = {
        "대표자 주소변경": {"fee": 55000, "tax": 52240, "note": "", "docs": "초본 / 법인인감도장 / 대표자 금융인증서"},
        "임원변경": {"fee": 55000, "tax": 52240, "note": "공증 여부 확인", "docs": "정관 / 주주명부 / 등기부등본 / 금융인증서(주주전원, 사/취임원) / 법인인감도장 / 사/취 초본 / 전자증명서"},
        "본점이전(관내)": {"fee": 99000, "tax": 156000, "note": "서울의 경우 산업단지에서 관내인지 확인", "docs": "정관 / 등기부등본 / 전자증명서 / 공증서류(필요시)"},
        "상호변경": {"fee": 99000, "tax": 52240, "note": "공증 여부 확인, 관할내 동일상호 확인", "docs": "정관 / 주주명부 / 등기부등본 / 금융인증서(주주전원) / 공증서류(필요시) / 전자증명서"},
        "목적변경": {"fee": 99000, "tax": 52240, "note": "공증 여부 확인", "docs": "정관 / 주주명부 / 등기부등본 / 금융인증서(주주전원) / 공증서류(필요시) / 전자증명서"},
        "지점설치": {"fee": 99000, "tax": 52240, "note": "타관할: 148,720원", "docs": "정관 / 등기부등본 / 전자증명서"},
        "공고방법변경": {"fee": 99000, "tax": 52240, "note": "공증 여부 확인", "docs": "정관 / 주주명부 / 등기부등본 / 금융인증서(주주전원) / 공증서류(필요시) / 전자증명서"},
        "주식매수선택권 규정": {"fee": 154000, "tax": 52240, "note": "공증 여부 확인", "docs": "정관 / 주주명부 / 등기부등본 / 금융인증서(주주전원) / 공증서류(필요시) / 전자증명서"},
        "본점이전(타관)": {"fee": 154000, "tax": 0, "note": "이전 지역에 따라 공과금 상이", "docs": "정관 / 주주명부 / 등기부등본 / 금융인증서(주주전원) / 전자증명서 / 공증서류(필요시)"},
        "법인설립": {"fee": 165000, "tax": 0, "note": "자본금/본점에 따라 공과금 상이 (과밀:550,000 / 비과밀:280,000)", "docs": "전원 인증서 / 초본 / 잔고증명서"},
        "법인설립(제휴)": {"fee": 99000, "tax": 0, "note": "자본금/본점에 따라 공과금 상이", "docs": "전원 인증서 / 초본 / 잔고증명서"},
        "주식매수선택권 행사": {"fee": 165000, "tax": 0, "note": "설립연도/지역/증자금액에 따라 공과금 상이", "docs": "정관 / 주주명부 / 등기부등본 / 청구인 정보 / 전자증명서(OTP)"},
        "유상증자(보통주)": {"fee": 209000, "tax": 0, "note": "설립연도/지역/증자금액에 따라 공과금 상이", "docs": "정관 / 주주명부 / 등기부등본 / 신주인수인 정보 / 금융인증서(주주전원) / 전자증명서 / 잔고증명서"},
        "유상증자(우선주)": {"fee": 275000, "tax": 0, "note": "설립연도/지역/증자금액에 따라 공과금 상이", "docs": "정관 / 주주명부 / 등기부등본 / 투자계약서 / 금융인증서(주주전원) / 전자증명서 / 잔고증명서"},
        "유상증자(가수금)": {"fee": 220000, "tax": 0, "note": "설립연도/지역/증자금액에 따라 공과금 상이", "docs": "정관 / 주주명부 / 등기부등본 / 금융인증서(주주전원) / 전자증명서 / 계정별원장(가수금)"},
        "감자등기": {"fee": 275000, "tax": 52240, "note": "공증 여부 확인 / 공고 별도", "docs": "-"},
        "사채발행": {"fee": 275000, "tax": 52240, "note": "공증 여부 확인", "docs": "정관 / 주주명부 / 등기부등본 / 투자계약서 / 금융인증서(주주전원) / 전자증명서"},
        "무상증자": {"fee": 363000, "tax": 0, "note": "설립연도/지역/증자금액에 따라 공과금 상이, 공증 여부 확인", "docs": "-"},
        "해산/청산": {"fee": 660000, "tax": 156720, "note": "공증 여부 확인", "docs": "-"},
    }
    
    # 추가 옵션
    CORP_EXTRA_OPTIONS = {
        "공증료 추가": {"fee": 110000, "tax_options": [30000, 60000]},
        "전자증명서 발급대행": {"fee": 55000, "tax": 0},
        "목적 추가 (10개당)": {"fee": 22000, "tax": 0},
    }
    
    col_left, col_right = st.columns([1.5, 1])
    
    with col_left:
        st.markdown("#### 📋 등기 종류 선택")
        
        # 등기 종류 선택
        selected_type = st.selectbox(
            "등기 종류",
            options=list(CORP_REGISTRY_FEES.keys()),
            key="corp_reg_type"
        )
        
        # 선택된 등기 정보
        selected_info = CORP_REGISTRY_FEES[selected_type]
        
        # 추가 옵션
        st.markdown("#### ➕ 추가 옵션")
        
        col_opt1, col_opt2 = st.columns(2)
        with col_opt1:
            add_notary = st.checkbox("공증료 추가", key="corp_add_notary")
            if add_notary:
                notary_type = st.radio("공증 유형", ["일반 (30,000원)", "특별 (60,000원)"], key="corp_notary_type", horizontal=True)
                notary_tax = 30000 if "30,000" in notary_type else 60000
            else:
                notary_tax = 0
        
        with col_opt2:
            add_cert = st.checkbox("전자증명서 발급대행 (+55,000원)", key="corp_add_cert")
        
        # 목적 추가 (법인설립 시)
        extra_purpose_fee = 0
        if "법인설립" in selected_type:
            st.markdown("---")
            purpose_count = st.number_input("목적 개수 (기본 10개 포함)", min_value=10, value=10, step=10, key="corp_purpose_count")
            if purpose_count > 10:
                extra_sets = (purpose_count - 10) // 10
                extra_purpose_fee = extra_sets * 22000
                st.caption(f"※ 추가 목적 {purpose_count - 10}개 → +{extra_purpose_fee:,}원")
        
        # 공과금 직접 입력 (변동 항목용)
        if selected_info["tax"] == 0:
            st.markdown("---")
            manual_tax = st.number_input(
                "공과금 직접입력", 
                min_value=0, 
                value=0, 
                step=10000,
                key="corp_manual_tax",
                help=selected_info["note"]
            )
        else:
            manual_tax = selected_info["tax"]
        
        # 비고
        if selected_info["note"]:
            st.info(f"📌 {selected_info['note']}")
    
    with col_right:
        st.markdown("#### 💰 비용 계산")
        
        # 계산
        base_fee = selected_info["fee"]
        base_tax = manual_tax
        
        notary_fee = 110000 if add_notary else 0
        cert_fee = 55000 if add_cert else 0
        
        total_fee = base_fee + notary_fee + cert_fee + extra_purpose_fee
        total_tax = base_tax + (notary_tax if add_notary else 0)
        grand_total = total_fee + total_tax
        
        # 결과 표시
        with st.container(border=True):
            st.markdown(f"**등기 종류:** {selected_type}")
            st.markdown("---")
            
            st.markdown("**📑 대행료**")
            st.markdown(f"- 기본 대행료: **{base_fee:,}원**")
            if notary_fee > 0:
                st.markdown(f"- 공증료: **{notary_fee:,}원**")
            if cert_fee > 0:
                st.markdown(f"- 전자증명서 발급: **{cert_fee:,}원**")
            if extra_purpose_fee > 0:
                st.markdown(f"- 목적 추가: **{extra_purpose_fee:,}원**")
            st.markdown(f"**대행료 소계: {total_fee:,}원**")
            
            st.markdown("---")
            st.markdown("**🏛️ 공과금**")
            st.markdown(f"- 등록면허세 등: **{base_tax:,}원**")
            if notary_tax > 0:
                st.markdown(f"- 공증 공과금: **{notary_tax:,}원**")
            st.markdown(f"**공과금 소계: {total_tax:,}원**")
            
            st.markdown("---")
            st.markdown(f"""
            <div style='background-color: #ff0033; color: white; padding: 15px; text-align: center; border-radius: 8px; margin-top: 10px;'>
                <div style='font-size: 1rem;'>총 합계</div>
                <div style='font-size: 1.8rem; font-weight: 800;'>{grand_total:,} 원</div>
            </div>
            """, unsafe_allow_html=True)
        
        # 필요서류
        st.markdown("#### 📎 필요서류")
        with st.expander("필요서류 보기", expanded=False):
            docs = selected_info["docs"]
            if docs and docs != "-":
                for doc in docs.split(" / "):
                    st.markdown(f"- {doc.strip()}")
            else:
                st.markdown("- 별도 문의")
        
        # 입금 안내
        st.markdown("#### 🏦 입금 안내")
        with st.container(border=True):
            st.markdown("""
            **입금자:** 법인명  
            **은 행:** 신한은행  
            **계 좌:** 100-035-852291  
            **예금주:** 법무법인시화
            """)

# =============================================================================
# 하단 푸터
# =============================================================================
st.markdown("---")
st.markdown("""<div style='text-align: center; color: #6c757d; padding: 20px; background-color: white; border-radius: 10px; border: 2px solid #e1e8ed;'>
    <p style='margin: 0; font-size: 1rem;'><strong><span style="color: #00428B;">DG-Form</span></strong> <span style="color: #6c757d;">전자설정 자동화시스템</span> | <strong><span style="color: #FDD000;">등기온</span></strong></p></div>""", unsafe_allow_html=True)